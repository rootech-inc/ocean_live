

def safe_add_image_to_pdf(pdf, image_path, x, y, w, h=None):
    """
    Safely add an image to PDF with fallback options
    """
    import os
    
    # Try the original image path first
    if os.path.exists(image_path):
        try:
            if h:
                pdf.image(image_path, x=x, y=y, w=w, h=h)
            else:
                pdf.image(image_path, x=x, y=y, w=w)
            return True
        except Exception as e:
            print(f"Warning: Could not load image {image_path}: {e}")
    
    # If original fails, try fallback options
    fallback_paths = [
        'static/general/img/logo.png',
        'static/general/img/logo-blue.png',
        'static/general/img/sneda_motors_logo.png'
    ]
    
    for fallback_path in fallback_paths:
        if os.path.exists(fallback_path) and fallback_path != image_path:
            try:
                if h:
                    pdf.image(fallback_path, x=x, y=y, w=w, h=h)
                else:
                    pdf.image(fallback_path, x=x, y=y, w=w)
                print(f"Used fallback image: {fallback_path}")
                return True
            except Exception as e:
                print(f"Warning: Could not load fallback image {fallback_path}: {e}")
                continue
    
    # If all images fail, just skip the logo
    print("Warning: No logo images could be loaded, skipping logo")
    return False

from django.views.decorators.csrf import csrf_exempt
@csrf_exempt
def interface(request):
    import json
    import sys
    from decimal import Decimal
    from django.http import JsonResponse

    from django.db.models import Q
    from django.contrib.auth.models import User
    from ocean.settings import BASE_URL

    from ssml.models import Cardex, Location, RequiredReturn, InvoiceHD, InvoiceTransactions, Contractor, Expense, \
        InventoryGroup, InventoryMaterial, Issue, IssueTransaction, Ledger, MaterialOrderItem, Meter, PaySlipHD, Plot, \
        RedeemTransactions, Reedem, Service, ServiceMaterials, ServiceOrder, ServiceOrderItem, ServiceOrderReturns, \
        ServiceType, Supplier, Grn, GrnTransaction, ContractorError, TransferHd, TransferTr
    from admin_panel.models import Emails, Locations, BusinessEntityTypes, MailSenders, MailQueues, MailAttachments, \
        Sms, SmsApi
    from admin_panel.anton import make_md5_hash
    response = {
        'status_code': 0,
        'message': ""
    }

    success_response = {
        'status_code': 200,
        'message': "Procedure Completed Successfully"
    }

    # get method
    method = request.method

    try:

        body = json.loads(request.body)
        module = body.get('module')
        data = body.get('data')


        if method == 'PUT':
            if module == 'contractor':
                link = data.get('link')
                company = data.get('company')
                owner = data.get('owner')
                phone = data.get('phone')
                email = data.get('email')
                country = data.get('country')
                city = data.get('city')
                postal_code = data.get('postal_code')
                gh_post_code = data.get('gh_post_code')
                gh_card_no = data.get('gh_card_no')
                created_by = User.objects.get(id=data.get('mypk'))
                code = data.get('code')

                
                Contractor.objects.create(code=code,link=link,company=company, owner=owner, phone=phone,email=email,country=country,
                                          city=city,postal_code=postal_code,gh_card_no=gh_card_no,gh_post_code=gh_post_code,created_by=created_by)

                success_response['message'] = "Contractor Created Successfully"

            elif module == 'daily_sms_report':
                from django.utils import timezone
                from datetime import datetime
                print(data)
                day_date = data.get('date',datetime.now().date())
                print(day_date)

                import openpyxl
                
                for contractor in Contractor.objects.all():
                    today_jobs = ServiceOrder.objects.filter(contractor=contractor,service_date=day_date)
                    total_jobs = ServiceOrder.objects.filter(contractor=contractor).count()

                    if today_jobs.count() > 0:
                        book = openpyxl.Workbook()
                        sheet = book.active
                        sheet.title = str(day_date)
                        sheet.append(['METER NUMBER','SERVICE TYPE'])
                        for job in today_jobs:
                            sheet.append([job.new_meter,job.service_type.name])
                        
                        
                        file_name = f"static/general/tmp/daily_report_{contractor.company.replace(' ','_')}_{str(day_date).replace('-','_')}.xlsx"
                        file_url = f"{BASE_URL}{file_name}"

                        import requests

                        long_url = file_url
                        api_url = "https://is.gd/create.php"

                        params = {
                            "format": "simple",  # "simple" returns just the short URL as text
                            "url": long_url
                        }

                        response = requests.get(api_url, params=params)
                        print(response)

                        if response.status_code == 200:
                            file_url = response.text
                        else:
                            print("Error:", response.text)


                        book.save(file_name)
                        sms_message = f"Daily SMS Report for {contractor.company} on {day_date}\nToday Jobs: {today_jobs.count()}\nTotal Jobs: {total_jobs}\n\nDownload the report: \n{file_url}"
                        print(sms_message)
                        to = contractor.phone.replace('233','0').replace('+233','0')
                        if len(to) == 10:
                            Sms(
                                api=SmsApi.objects.get(is_default=True),
                                to=to,
                                message=sms_message
                            ).save()
                    
                    
                    

            elif module == 'return_qty':
                service_id = data.get('service_id')
                transactions = data.get('transactions')
                st = ServiceType.objects.get(id=service_id)
                qty_as_of = data.get('qty_as_of')

                for tran in transactions:
                    idd = tran.get('id')
                    qty = tran.get('qty')

                    print(tran)

                    RequiredReturn.objects.create(
                        service_type=st,
                        material_id=idd,
                        as_of=qty_as_of,
                        quantity=qty
                    )

                success_response['message'] = "Operated"
                response = success_response

            elif module == 'service_materials':
                barcode = data.get('barcode')
                material = InventoryMaterial.objects.get(barcode=barcode)
                ser_id = data.get('service')
                qty = data.get('qty')

                ServiceMaterials.objects.create(
                    service_id=ser_id,
                    material=material,
                    qty=qty
                )

                success_response['message'] = "Created"
                response = success_response

            elif module == 'location':
                loc_id = data.get('loc_id')
                loc_name = data.get('loc_name')
                loc_desc = data.get('loc_desc')
                country = data.get('country')
                city = data.get('city')
                street = data.get('street')
                address = data.get('address')
                phone = data.get('phone')
                email = data.get('email')
                created_by = User.objects.get(pk=data.get('mypk'))


                Location.objects.create(
                    loc_id=loc_id,
                    loc_name=loc_name,
                    loc_desc = loc_desc,
                    country=country,
                    street=street,
                    address=address,
                    phone=phone,
                    email=email,
                    city=city
                )



            elif module == 'finance':
                direction = data.get('direction')
                date = data.get('date')
                category = data.get('category')
                transaction_type = data.get('transaction_type')
                amount = Decimal(data.get('amount'))
                reference = data.get('reference')
                description = data.get('description')
                created_by = User.objects.get(id=data.get('mypk'))

                print(data)

                if direction == 'out':
                    amount = amount * -1
                Expense.objects.create(direction=direction, date=date, category=category, transaction_type=transaction_type, amount=amount, reference=reference, description=description, created_by=created_by)
                success_response['message'] = "Expense Created Successfully"

            elif module == 'invoice':

                if  InvoiceHD.objects.filter(is_approved=False).exists():
                    raise Exception("You have unposted entries")
               
                try:
                    entry_no = f"SSML-INV-1" if InvoiceHD.objects.all().count() == 0 else f"SSML-INV-{InvoiceHD.objects.all().last().id + 1}"
                    contractor_id = data.get('contractor')
                    contractor = Contractor.objects.get(id=contractor_id)
                    entry_date = data.get('entry_date')
                    remarks = data.get('remarks')
                    transactions = data.get('transactions')
                    created_by = User.objects.get(pk=data.get('mypk'))

                    # save hd
                    InvoiceHD(
                        entry_no=entry_no,
                        entry_date=entry_date,
                        contractor=contractor,
                        remarks=remarks,
                        created_by=created_by
                    ).save()

                    entry = InvoiceHD.objects.get(entry_no=entry_no)

                    print(transactions)

                    total = 0
                    # loop transactions
                    for transaction in transactions:
                        InvoiceTransactions(
                            entry=entry,
                            service_date=transaction.get('date'),
                            asset=transaction.get('meter'),
                            amount=transaction.get('amount'),
                            remarks=transaction.get('service')
                        ).save()
                        print(transaction)

                        total += Decimal(transaction.get('amount'))

                    success_response['message'] = "Invoice Created"
                    entry.total_amount = total
                    entry.save()
                except Exception as e:
                    # Roll back any Issue transactions created
                    if 'entry' in locals():
                        InvoiceHD.objects.filter(entry=entry).delete()
                        entry.delete()  
                    success_response['status_code'] = 400
                    success_response['message'] = f"Error on line {sys.exc_info()[2].tb_lineno}: {e}"

            elif module == 'approve_invoice':
                id = data.get('id')
                for invoice in InvoiceHD.objects.filter(id=id):
                    # mark meters
                    for transaction in invoice.transactions():
                        print(transaction)
                        service = ServiceOrder.objects.get(new_meter=transaction.get('asset'))
                        status = service.status
                        service.status = 'invoiced'
                        service.save()
                        print(status)

                    invoice.is_approved = True
                    invoice.approved_by = User.objects.get(pk=data.get('mypk'))
                    invoice.save()

            elif module == 'post_invoice':
                id = data.get('id')
                approved_by = User.objects.get(pk=data.get('mypk'))
                for invoice in InvoiceHD.objects.filter(id=id):
                    # mark meters
                    for transaction in invoice.transactions():
                        print(transaction)       

            elif module == 'meters':
                meters = data.get('meters')
                meter_type = data.get('meter_type')
                created_by = User.objects.get(id=data.get('mypk'))
                contractor = Contractor.objects.get(id=data.get('contractor'))

                for meter in meters:
                    try:
                        Meter.objects.create(meter_no=meter, meter_type=meter_type, created_by=created_by, contractor=contractor)
                    except Exception as e:
                        print(e)
                        pass
                success_response['message'] = "Meters Created Successfully" 

            elif module == 'group':
                name = data.get('name')
                InventoryGroup.objects.create(name=name)
                success_response['message'] = "Group Created Successfully"
            elif module == 'material':
                barcode = data.get('barcode',f"MAT-{InventoryMaterial.objects.all().count() + 1}")
                name = data.get('name')
                group = InventoryGroup.objects.get(id=data.get('group'))
                reorder_qty = data.get('reorder_qty')
                value = data.get('value')


                # check name and barcode
                if InventoryMaterial.objects.filter(name=name).exists():
                    success_response['status_code'] = 400
                    success_response['message'] = "Material Name Already Exists"
                elif InventoryMaterial.objects.filter(barcode=barcode).exists():
                    success_response['status_code'] = 400
                    success_response['message'] = "Material Barcode Already Exists"
                else:
                    InventoryMaterial.objects.create(
                        barcode=barcode,
                        name=name,
                        group=group,
                        reorder_qty=reorder_qty,
                        value=value
                    )
                    success_response['message'] = "Material Created Successfully"

            elif module == 'supplier':
                supplier_name = data.get('supplier_name')
                contact_person = data.get('contact_person')
                supplier_address = data.get('supplier_address')
                supplier_phone = data.get('supplier_phone')
                supplier_email = data.get('supplier_email')
                Supplier.objects.create(supplier_name=supplier_name, contact_person=contact_person, supplier_address=supplier_address, supplier_phone=supplier_phone, supplier_email=supplier_email)
                success_response['message'] = "Supplier Created Successfully"

            elif module == 'grn':
                try:
                    supplier = Supplier.objects.get(id=data.get('supplier'))
                    grn_date = data.get('grn_date')
                    remarks = data.get('remarks')
                    grn_type = data.get('grn_type')
                    created_by = User.objects.get(id=data.get('created_by'))
                    location = Location.objects.get(pk=data.get('location'))
                    grn = Grn.objects.create(
                        supplier=supplier, 
                        grn_date=grn_date, 
                        remarks=remarks, 
                        grn_type=grn_type, 
                        created_by=created_by,
                        total_amount=0,
                        total_qty=0,
                        grn_no = f"GRN-{Grn.objects.all().count() + 1}",
                        location=location
                        )
                    total_amount = 0
                    total_qty = 0
                    for transaction in data.get('transactions'):
                        material = InventoryMaterial.objects.get(barcode=transaction.get('barcode'))
                        barcode = transaction.get('barcode')
                        name = transaction.get('name')
                        amount = transaction.get('amount')
                        pack_qty = transaction.get('pack_qty')
                        uom = transaction.get('uom')
                        rate = transaction.get('rate')
                        qty = transaction.get('qty')
                        total_qty = Decimal(qty) * Decimal(pack_qty)
                        GrnTransaction.objects.create(
                            grn=grn, 
                            material=material, 
                            barcode=barcode, 
                            name=name, 
                            amount=amount, 
                            pack_qty=pack_qty, 
                            uom=uom, 
                            rate=rate, 
                            qty=qty, 
                            total_qty=total_qty)
                        total_amount += Decimal(amount)
                        total_qty += Decimal(total_qty)

                    grn.total_amount = total_amount
                    grn.total_qty = total_qty
                    grn.save()
                        
                    success_response['message'] = "GRN Created Successfully"
                except Exception as e:
                    # Roll back any GRN transactions created
                    if 'grn' in locals():
                        GrnTransaction.objects.filter(grn=grn).delete()
                        grn.delete()
                    success_response['status_code'] = 400
                    success_response['message'] = f"Error on line {sys.exc_info()[2].tb_lineno}: {e}"
                
            elif module == 'issue':
                next_no = Issue.objects.all().last()
                if Issue.objects.all().count() > 0:
                    issue_no = f"IS-{next_no.id + 1}"
                else:
                    issue_no = f"IS-1"
                try:
                    contractor = Contractor.objects.get(id=data.get('contractor'))
                    issue_date = data.get('issue_date')
                    remarks = data.get('remarks')
                    issue_type = data.get('issue_type')
                    created_by = User.objects.get(id=data.get('created_by'))
                    car_no = data.get('car_no')
                    loc_id = data.get('loc_id')
                    location = Location.objects.get(pk=loc_id) if loc_id else Location.objects.get(loc_id='002')
                        
                    issue = Issue.objects.create(
                        contractor=contractor,
                        issue_date=issue_date,
                        remarks=remarks,
                        issue_type=issue_type,
                        created_by=created_by,
                        issue_no=issue_no,
                        car_no=car_no,
                        location=location
                    )
                    total_amount = 0
                    total_qty = 0
                    for transaction in data.get('transactions'):
                        material = InventoryMaterial.objects.get(barcode=transaction.get('barcode'))
                        barcode = transaction.get('barcode')
                        name = transaction.get('name')
                        amount = transaction.get('amount')
                        pack_qty = transaction.get('pack_qty')
                        uom = transaction.get('uom')
                        rate = transaction.get('rate')
                        qty = transaction.get('qty')
                        total_qty = Decimal(qty) * Decimal(pack_qty)
                        IssueTransaction.objects.create(
                            issue=issue,
                            material=material,
                            barcode=barcode,
                            name=name,
                            amount=amount,
                            pack_qty=pack_qty,
                            uom=uom,
                            rate=rate,
                            qty=qty,
                            total_qty=total_qty)

                    issue.total_amount = total_amount
                    issue.total_qty = total_qty

                    for meter in data.get('meters'):
                        if Meter.objects.filter(meter_no=meter).exists():
                            # change contractor
                            Meter.objects.filter(meter_no=meter).update(contractor=contractor)
                        else:
                            Meter.objects.create(meter_no=meter, contractor=contractor, created_by=created_by, issue=issue,location=location)

                    issue.save()

                    success_response['message'] = "Issue Created Successfully"
                except Exception as e:
                    # Roll back any Issue transactions created
                    if 'issue' in locals():
                        Issue.objects.filter(issue_no=issue_no).delete()
                        issue.delete()  
                    success_response['status_code'] = 400
                    success_response['message'] = f"Error on line {sys.exc_info()[2].tb_lineno}: {e}"
            
            elif module == 'plot':
                plot_no = data.get('plot-no')
                contractor = Contractor.objects.get(id=data.get('contractor'))
                plot_size = data.get('plot-size')
                created_by = User.objects.get(id=data.get('mypk'))
                Plot.objects.create(plot_no=plot_no, contractor=contractor, plot_size=plot_size,created_by=created_by)
                success_response['message'] = "Plot Created Successfully"

            elif module == 'service_type':
                name = data.get('service_name')
                description = data.get('service_description')
                created_by = User.objects.get(id=data.get('mypk'))
                ServiceType.objects.create(name=name, description=description, created_by=created_by)
                success_response['message'] = "Service Type Created Successfully"

            elif module == 'service':
                code = f"SRV-{Service.objects.all().count() + 1}"
                name = data.get('service_name')
                description = data.get('service_description')
                rate = data.get('service_rate')
                uom = data.get('service_uom')
                created_by = User.objects.get(id=data.get('mypk'))
                Service.objects.create(code=code, name=name, description=description, rate=rate, uom=uom, created_by=created_by)
                success_response['message'] = "Service Created Successfully"

            elif module == 'service_order':
                service_type = ServiceType.objects.get(id=data.get('service-type'))
                service_date = data.get('service-date')
                contractor = Contractor.objects.get(id=data.get('contractor'))
                plot = Plot.objects.get(plot_no=data.get('plot'))
                geo_code = data.get('geo-code')
                customer = data.get('customer')
                customer_no = data.get('customer_no')
                old_meter_no = data.get('old_meter_no')
                new_meter_no = data.get('new_meter_no')
                meter = Meter.objects.get(meter_no=new_meter_no)
                # check if meter belongs to contractor
                if meter.issue.contractor != contractor:
                    raise Exception("Meter Does Not Belong To Contractor")

                # check if meter is already issued
                if meter.is_issued:
                    raise Exception("Meter Already Issued")

                created_by = User.objects.get(id=data.get('mypk'))
                service_order = ServiceOrder.objects.create(service_type=service_type, service_date=service_date, contractor=contractor, plot=plot, geo_code=geo_code, customer=customer, customer_no=customer_no, old_meter_no=old_meter_no, new_meter=meter, created_by=created_by)
                meter.is_issued = True
                meter.service_order = service_order
                meter.save()
                success_response['message'] = "Service Order Created Successfully"

            elif module == 'service_order_item':
                service_order = ServiceOrder.objects.get(id=data.get('order_id'))
                service = Service.objects.get(id=data.get('service_id'))
                quantity = data.get('serv_qty')
                rate = service.rate
                amount = Decimal(quantity) * Decimal(rate)
                ServiceOrderItem.objects.create(service_order=service_order, service=service, quantity=quantity, rate=rate, amount=amount)
                success_response['message'] = "Service Order Item Created Successfully"
                
            elif module == 'material_order_item':
                service_order = ServiceOrder.objects.get(id=data.get('order_id'))
                material = InventoryMaterial.objects.get(id=data.get('material_id'))
                quantity = data.get('mat_qty')
                material_type = data.get('mat_type')
                
                rate = 0
                amount = Decimal(quantity) * Decimal(rate)
                MaterialOrderItem.objects.create(service_order=service_order, material=material, quantity=quantity, rate=rate, amount=amount, material_type=material_type)
                success_response['message'] = "Material Order Item Created Successfully"

            elif module == 'service_order_return':
                service_order = ServiceOrder.objects.get(id=data.get('server_order_id'))
                material = InventoryMaterial.objects.get(id=data.get('material_id'))
                quantity = data.get('ret_qty')
                user = User.objects.get(id=data.get('mypk'))
                ServiceOrderReturns.objects.create(service_order=service_order, material=material, quantity=quantity, created_by=user, modified_by=user)
                success_response['message'] = "Service Order Return Created Successfully"

            elif module == 'service_order_new':
                header = data.get('header')
                mats = data.get('materials')
                services = data.get('services')
                returns = data.get('returns')
                user = User.objects.get(id=header.get('mypk'))

                # print(mats)
                # print(services)
                # print(returns)

                # raise Exception("Test End")

                # print(mats)
                # raise Exception("Test End")

                # validate meter
                new_meter_no = header.get('new_meter_no')
                contractor = Contractor.objects.get(id=header.get('contractor'))
                print(new_meter_no)
                meter = Meter.objects.get(meter_no=header.get('new_meter_no'))



                # check if meter belongs to contractor
                if meter.contractor != contractor:
                    raise Exception("Meter Does Not Belong To Contractor")

                # check if meter is already issued
                if meter.is_issued:
                    raise Exception("Meter Already Issued")



                try:
                    # create service order
                    service_order = ServiceOrder.objects.create(
                        service_type_id=header.get('service_type'),
                        service_date=header.get('service_date'),
                        contractor_id=header.get('contractor'),
                        plot_id=header.get('plot'),
                        geo_code=header.get('cust_code'),
                        customer=header.get('customer'),
                        customer_no=header.get('customer_no'),
                        old_meter_no=header.get('old_meter_no'),
                        new_meter=meter.meter_no,
                        created_by=user,
                        old_meter_no_reading=header.get('old_meter_no_reading'),
                        new_meter_no_reading=header.get('new_meter_no_reading')
                    )

                    # create service order items
                    for service in services:
                        r_service = Service.objects.get(code=service.get('code'))
                        ServiceOrderItem.objects.create(
                            service_order=service_order,
                            service=Service.objects.get(code=service.get('code')),
                            rate=r_service.rate,
                            quantity=service.get('quantity'),
                            amount=Decimal(service.get('quantity')) * Decimal(r_service.rate)
                        )
                    
                    # # create material order items
                    for mat in mats:
                        print(mat)
                        mt = InventoryMaterial.objects.get(barcode=mat.get('barcode'))
                        MaterialOrderItem.objects.create(
                            service_order=service_order,
                            material=InventoryMaterial.objects.get(barcode=mat.get('barcode')),
                            quantity=mat.get('quantity'),
                            material_type='is',
                            rate=mt.value,
                            amount=Decimal(mt.value) * Decimal(mat.get('quantity'))
                        )

                    # # automate materials add
                    # for ret_item in InventoryMaterial.objects.filter(is_return=True):
                    #     try:
                    #         MaterialOrderItem.objects.create(
                    #             service_order=service_order,
                    #             material=ret_item,
                    #             quantity=1,
                    #             material_type='rt',
                    #             rate=ret_item.value,
                    #             amount=Decimal(mt.value) * Decimal(mat.get('quantity'))
                    #         )
                    #     except Exception as e:
                    #         pass
                    
                    # create service order returns
                    for rt in returns:
                        ServiceOrderReturns.objects.create(
                            service_order=service_order,
                            material=InventoryMaterial.objects.get(barcode=rt.get('barcode')),
                            quantity=rt.get('quantity'),
                            created_by=user,
                            modified_by=user
                        )

                    # add automatic returns
                    for ret_item in InventoryMaterial.objects.filter(is_return=True):
                        try:
                            ServiceOrderReturns.objects.create(
                                service_order=service_order,
                                material=ret_item,
                                quantity=1,
                                created_by=user,
                                modified_by=user
                            )
                        except Exception as e:
                            pass

                    

                    meter.is_issued = True
                    meter.service_order = service_order
                    meter.save()

                    success_response['message'] = "Service Order Created Successfully"

                

                except Exception as e:
                    # delete service order if created
                    if 'service_order' in locals():
                        service_order.delete()
                    success_response['status_code'] = 400
                    success_response['message'] = f"Error on line {sys.exc_info()[2].tb_lineno}: {e}"

            elif module == 'redeem':
                

                try:
                    header = data.get('header')
                    transactions = data.get('transactions')

                    contractor = Contractor.objects.get(id=header.get('contractor'))
                    entry_date = header.get('date')
                    user = User.objects.get(pk=header.get('mypk'))
                    remarks = header.get('remarks')
                    typ = header.get('type')
                    entry_no = f"RED-{Reedem.objects.all().last().pk + 1}" if Reedem.objects.all().count() > 0 else "RED-1"


                    entry = Reedem.objects.create(
                        entry_no=entry_no,
                        entry_date=entry_date,
                        contractor=contractor,created_by=user,
                        transaction_type=typ,remarks=remarks
                    )

                    # enter transactions
                    for transaction in transactions:
                        
                        barcode = transaction.get('barcode')
                        mt = InventoryMaterial.objects.get(barcode=barcode)
                        pack_um = transaction.get('pack_um')
                        balance = transaction.get('balance')
                        reason = transaction.get('reason')
                        qty=transaction.get('qty')
                        RedeemTransactions.objects.create(
                            redeem=entry,
                            material=mt,
                            pack_um=pack_um,
                            balance=balance,
                            reason=reason,
                            qty=qty
                        )





                except Exception as e:
                    # delete service order if created
                    if 'entry' in locals():
                        entry.delete()
                    success_response['status_code'] = 400
                    success_response['message'] = f"Error on line {sys.exc_info()[2].tb_lineno}: {e}"

            elif module == 'transfer':
                hdd = data.get('header')
                print(hdd)
                transactions = data.get('transactions')

                entry_date = hdd.get('entry_date')
                loc_from = hdd.get('loc_from')
                loc_to = hdd.get('loc_to')
                mypk = hdd.get('mypk')
                remarks = hdd.get('remarks')
                tran_type = hdd.get('tran_type')

                if loc_from == loc_to:
                    raise Exception("Location from and to cannot be same")

                owner = User.objects.get(pk=mypk)
                loc_fr = Location.objects.get(pk=loc_from)
                loc_t = Location.objects.get(pk=loc_to)

                try:
                
                    next_no = TransferHd.objects.all().last().pk + 1 if TransferHd.objects.all().count() > 0 else 1

                    from datetime import datetime

                    entry_date_str = hdd.get('entry_date')  # e.g. "2025-10-02"

                    # Convert string to datetime (assuming format YYYY-MM-DD)
                    entry_date = datetime.strptime(entry_date_str, "%Y-%m-%d").date()

                    # Extract month and day
                    month = entry_date.month
                    day = entry_date.day

                    # Add month and year to entry_no
                    entry_no = f'TR{month}{day}-{next_no}'

                    print(entry_no)


                    TransferHd.objects.create(
                        loc_fr=loc_fr,
                        loc_to=loc_t,
                        entry_date=entry_date,
                        remarks=remarks,
                        created_by=owner,
                        tran_type=tran_type,
                        entry_no=entry_no
                    )

                    hd = TransferHd.objects.get(entry_no=entry_no)

                    for transaction in transactions:
                        print(transaction)
                        mt = InventoryMaterial.objects.get(barcode=transaction.get('barcode'))


                        if TransferTr.objects.filter(entry=hd, material=mt).exists():
                            pass
                        else:

                            TransferTr.objects.create(
                                entry=hd,
                                material=mt,
                                name=mt.name,
                                barcode=mt.barcode,
                                oum=transaction.get('oum'),
                                pack_qty=transaction.get('pack_qty'),
                                sent_qty=transaction.get('sent_qty')
                            )

                    
                    success_response['message'] = "Transfer Created Successfully"
                    success_response['status_code'] = 200

                    
                    # todo: save tran hd
                
                except Exception as e:
                    # If TransferHd was created, delete it to avoid orphaned header
                    try:
                        if 'hd' in locals() and isinstance(hd, TransferHd):
                            hd.delete()
                    except Exception:
                        pass
                    success_response['status_code'] = 400
                    success_response['message'] = f"Error on line {sys.exc_info()[2].tb_lineno}: {e}"


            else:
                success_response['status_code'] = 400
                success_response['message'] = f"Module Not Found"



        elif method == 'VIEW':
                
                if module == 'contractor':
                    id = data.get('id','*') 
                    if id == '*':
                        contractors = Contractor.objects.all().order_by('company')
                        success_response['message'] = [contractor.obj() for contractor in contractors]
                    else:
                        contractor = Contractor.objects.get(id=id)
                        success_response['message'] = contractor.obj()

                elif module == 'foc_otp':
                    phone = data.get('phone')
                    if Contractor.objects.filter(phone=phone).exists():
                        # generate opt
                        import random
                        otp = random.randint(101010, 999999)

                        contractor = Contractor.objects.get(phone=phone)
                        contractor.otp = otp
                        contractor.save()

                        # Send OTP via SMS
                        to = phone.replace('233', '0').replace('+233', '0')
                        if len(to) == 10:
                            sms_message = f"Your OTP code is: {otp}"
                            Sms(
                                api=SmsApi.objects.get(sender_id='SNEDA SML') if SmsApi.objects.filter(sender_id='SNEDA SML').exists() else SmsApi.objects.get(is_default=True),
                                to=to,
                                message=sms_message
                            ).save()

                        success_response['message'] = otp
                    else:
                        raise Exception("Contractor does not exist")

                elif module == 'transfer':
                    id = data.get('id','*')
                    if id == '*':
                        transfers = TransferHd.objects.all().order_by('entry_date')
                        success_response['message'] = [transfer.obj() for transfer in transfers]
                    else:
                        if isinstance(id, int) or (isinstance(id, str) and id.isdigit()):
                            transfer = TransferHd.objects.filter(id=id)
                        else:
                            transfer = TransferHd.objects.filter(entry_no=id)

                        success_response['message'] = [t.obj() for t in transfer]
                        

                elif module == 'job_material_details':
                    contractor_code = data.get('code','*')
                    contractor = Contractor.objects.get(code=contractor_code)

                    import openpyxl
                    book = openpyxl.Workbook()
                    sheet = book.active

                    for job in ServiceOrder.objects.filter(contractor=contractor):
                        jb = job.obj()
                        meter = job.new_meter
                        customer = job.customer
                        customer_no = job.customer_no
                        sheet.append(['#############################'])
                        sheet.append(["METER","CUSTOMER"])
                        sheet.append([meter,f"[{customer} - {customer_no}]"])

                        sheet.append(['MATERIAL',"QUANTITY"])
                        mats = MaterialOrderItem.objects.filter(service_order=job,material_type='is')
                        print(mats)
                        for material in mats:
                            
                            sheet.append([material.material.name,material.quantity])
                        sheet.append(['#############################'])
                        sheet.append([''])

                    file_name = f'static/general/tmp/{contractor.company}.xlsx'
                    book.save(file_name)

                        

                    success_response['message'] = file_name

                    response = success_response

                elif module == 'month_walk':
                    
                    from datetime import datetime
                    location = '*'
                    if data.get('location'):
                        location = Location.objects.get(id=data.get('location'))
                    now = datetime.now()
                    year = now.year
                    current_month = now.month
                    months = [["Month","Stock","Installed"]]
                    import openpyxl
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    header = ['MONTH',"STOCK",'INSTALLED','%']
                    ws.append(header)
                    
                    for month in range(1, current_month + 1):
                        import calendar
                        month_name = calendar.month_name[month]
                        # Get stock of sum of material in cardex, within this month
                        from django.db.models import Sum
                        # Calculate the first and last day of the month
                        from datetime import datetime
                        first_day = datetime(year, month, 1)
                        if month == 12:
                            last_day = datetime(year + 1, 1, 1)
                        else:
                            last_day = datetime(year, month + 1, 1)
                        # Aggregate sum of material quantity in cardex for this month
                        
                        if location == '*':
                            stock_sum = Cardex.objects.filter(
                                created_at__lte=last_day,
                                material__item_type='pri'
                                ).aggregate(total=Sum('qty'))['total'] or 0
                        else:
                            stock_sum = Cardex.objects.filter(
                                
                                created_at__lte=last_day,
                                material__item_type='pri',
                                location=location
                                ).aggregate(total=Sum('qty'))['total'] or 0

                        from decimal import Decimal
                        stock_sum = int(stock_sum)
                        print(stock_sum, type(stock_sum))

                        from ssml.models import ServiceOrder
                        if location == '*':
                            total_service_orders = ServiceOrder.objects.filter(service_date__gte=first_day, service_date__lt=last_day).count()
                        else:
                            total_service_orders = ServiceOrder.objects.filter(service_date__gte=first_day, service_date__lt=last_day,location=location).count()


                        stock_sum = round(stock_sum, 2)
                        months.append([month_name,stock_sum,total_service_orders])
                        # INSERT_YOUR_CODE
                        percent = 0
                        if stock_sum > 0:
                            percent = round((total_service_orders / stock_sum) * 100, 2)
                        else:
                            percent = 0
                        ws.append([month_name,stock_sum,total_service_orders,percent])
                    
                    file = fr"static/general/tmp/status.xlsx"
                    wb.save(file)
                    success_response['message'] = months
                    success_response['file'] = file
                    response = success_response

                elif module == 'location':
                    id = data.get('id','*')
                    if id == '*':
                        success_response['message'] = [loc.obj() for loc in Location.objects.all().order_by('loc_name')]
                    else:
                        success_response['message'] =[loc.obj() for loc in Location.objects.get(id=id)]
                    

                elif module == 'expense':
                    from django.db.models import Sum
                    from datetime import datetime
                    id = data.get('id','*')
                    limit = data.get('limit',10000)
                    as_of = data.get('as_of', f'{datetime.now().year}-01-01')
                    doc = data.get('doc','json')
                    direction = data.get('direction')
                    print(data)
                    if id == '*':
                        filt = data.get('filter','pending')
                        print(filt)
                        if filt == 'approved':
                            expenses = Expense.objects.filter(is_approved=True)
                        elif filt == 'pending':
                            expenses = Expense.objects.filter(is_approved=False)
                        else:
                            expenses = Expense.objects.all()
                        # if direction == '*':
                        #     direction = 'all'
                        #     expenses = expenses.filter(is_approved=False).order_by('date')
                        # else:
                        #     expenses = expenses.filter(direction=direction,is_approved=False).order_by('date')
                        # expenses.filter(date=as_of)
                        # expenses.order_by('-date')[:limit]
                        if doc == 'json':
                            arr  = [expense.obj() for expense in expenses]
                        
                        if doc == 'excel':
                            import openpyxl
                            wb = openpyxl.Workbook()
                            ws = wb.active
                            header = ['DIRECTION',"TYPE",'AMOUNT','DATE','REF',"DESCRIPTION"]
                            ws.append(header)
                            

                            for expense in expenses:
                                tr = [expense.direction,expense.category,expense.amount,expense.date,expense.reference,expense.description]
                                ws.append(tr)

                            file_name = f'static/general/tmp/{direction}_expensis.xlsx'
                            wb.save(file_name)
                            arr = file_name

                    else:
                        expense = Expense.objects.get(id=id)
                        arr = expense.obj()

                    if filt == 'approved':
                        success_response['message'] = {
                            'total': "{:,.2f}".format(Expense.objects.filter(is_approved=True).aggregate(total=Sum('amount'))['total'] or 0),
                            'total_in': "{:,.2f}".format(Expense.objects.filter(direction='in',is_approved=True).aggregate(total=Sum('amount'))['total'] or 0),
                            'total_out': "{:,.2f}".format(Expense.objects.filter(direction='out',is_approved=True).aggregate(total=Sum('amount'))['total'] or 0),
                            'expenses':arr
                        }
                    else:
                        success_response['message'] = {
                            'total': "{:,.2f}".format(Expense.objects.filter(is_approved=False).aggregate(total=Sum('amount'))['total'] or 0),
                            'total_in': "{:,.2f}".format(Expense.objects.filter(direction='in',is_approved=False).aggregate(total=Sum('amount'))['total'] or 0),
                            'total_out': "{:,.2f}".format(Expense.objects.filter(direction='out',is_approved=False).aggregate(total=Sum('amount'))['total'] or 0),
                            'expenses':arr
                        }

                elif module == 'invoice':
                    
                    id = data.get('id','*')
                    if id == '*':
                        arr = [inv.obj() for inv in InvoiceHD.objects.all()]
                    else:
                        arr = [inv.obj() for inv in InvoiceHD.objects.filter(id=id)]

                    success_response['message'] = arr

                elif module == 'print_invoice':
                    id = data.get('id')
                    invoice = InvoiceHD.objects.get(id=id)

                    from fpdf import FPDF
                    pdf = FPDF()
                    pdf.add_page()  
                    
                    pdf.set_font("Arial","",10)
                    pdf.cell(190, 5, txt="Sneda Smart Meters Limited", ln=True, border=0, align="C")
                    pdf.cell(190, 5, txt="P.O. Box 1024, Accra, Ghana", ln=True, border=0, align="C")
                    pdf.cell(190, 5, txt="Tel: +233 54 631 0011", ln=True, border=0, align="C")
                    pdf.cell(190, 5, txt="Email: solomon@snedaghana.com", ln=True, border=0, align="C")
                    pdf.ln(10)

                    pdf.set_font("Arial","B", size=20)
                    pdf.cell(190, 5, txt="JOBS COMPLETED DOCUMENT", ln=True, border=0, align="C")
                    pdf.ln(10)

                    pdf.set_font("Arial","B", size=12)
                    pdf.cell(95, 5, txt=invoice.contractor.company, ln=False, align="L",border=0)
                    
                    
                    pdf.set_font("Arial", "B", size=12)
                    pdf.cell(95, 5, txt=str(f"GHS {invoice.total_amount:,.2f}"), ln=True, align="R",border=0)

                    pdf.set_font("Arial","B", size=10)
                    pdf.cell(50, 5, txt="SERVICE", ln=False, align="L",border=1)
                    pdf.cell(50, 5, txt="Date.", ln=False, align="L",border=1)
                    pdf.cell(50, 5, txt="Asset", ln=False, align="L",border=1)
                    pdf.cell(40, 5, txt="Amount", ln=True, align="L",border=1)

                    pdf.set_font("Arial","", size=8)
                    for transaction in InvoiceTransactions.objects.filter(entry=invoice):
                        pdf.cell(50, 5, txt=transaction.remarks, ln=False, align="L",border=1)
                        pdf.cell(50, 5, txt=str(transaction.service_date), ln=False, align="L",border=1)
                        pdf.cell(50, 5, txt=transaction.asset, ln=False, align="L",border=1)
                        pdf.cell(40, 5, txt=str(transaction.amount), ln=True, align="L",border=1)

                    pdf.multi_cell(190, 10, txt="This document is subject to review and is not a payslip. After review, an official invoice will be sent to you to commerce payment",border=False, align="L")


                    pdf.ln(5)
                    pdf.set_font("Arial", size=8)
                    pdf.cell(190 / 6, 8, txt=f"{invoice.created_by.first_name} {invoice.created_by.last_name}", ln=False, align="C",border=1)
                    pdf.cell(190 / 6, 8, txt="", ln=False, align="R",border=0)
                    pdf.cell(190 / 12, 8, txt="", ln=False, align="R",border=0)
                    pdf.cell(190 / 6, 8, txt=f"{invoice.approved_by.get_full_name() if invoice.approved_by else ''}", ln=False, align="C",border=1)
                    pdf.cell(190 / 12, 8, txt="", ln=False, align="R",border=0)
                    pdf.cell(190 / 6, 8, txt="", ln=False, align="R",border=0)
                    pdf.cell(190 / 6, 8, txt="", ln=True, align="R",border=1)


                    pdf.cell(190 / 6, 8, txt="Created By", ln=False, align="C",border=0)
                    pdf.cell(190 / 6, 8, txt="", ln=False, align="R",border=0)
                    pdf.cell(190 / 12, 8, txt="", ln=False, align="R",border=0)
                    pdf.cell(190 / 6, 8, txt="Posted By", ln=False, align="C",border=0)
                    pdf.cell(190 / 12, 8, txt="", ln=False, align="R",border=0)
                    pdf.cell(190 / 6, 8, txt="", ln=False, align="R",border=0)
                    pdf.cell(190 / 6, 8, txt="Confirmed By", ln=True, align="C",border=0)



                    file_name = f'static/general/tmp/{invoice.entry_no}.pdf'
                    pdf.output(file_name)

                    success_response['message'] = f"/{file_name}"


                elif module == 'print_tr':
                    entry_no = data.get('entry_no')
                    
                    entry = TransferHd.objects.get(entry_no=entry_no)

                    from fpdf import FPDF
                    pdf = FPDF()
                    pdf.add_page() 

                # INSERT_YOUR_CODE
                    pdf = FPDF(orientation='L', unit='mm', format='A4')
                    pdf.add_page()
                    pdf.set_auto_page_break(auto=True, margin=15)

                    # Define a function to draw the header on each page
                    def draw_waybill_header(pdf):
                        # Logo
                        safe_add_image_to_pdf(pdf, 'static/general/logo.png', x=10, y=8, w=35)

                        # Title
                        pdf.set_font("Arial", 'B', 20)
                        pdf.set_xy(70, 10)
                        pdf.cell(0, 10, "SNEDA SMART METERS", ln=1, align="L")

                        # WAYBILL
                        pdf.set_font("Arial", 'B', 18)
                        pdf.set_xy(230, 10)
                        pdf.cell(0, 10, f"WAYBILL: {entry.entry_no}", ln=1, align="L")

                        # TO, ADDRESS, STATION
                        pdf.set_font("Arial", '', 12)
                        pdf.set_xy(70, 20)
                        pdf.cell(0, 8, f"FROM:  {entry.loc_fr.loc_name}", ln=1, align="L")
                        pdf.set_xy(70, 28)
                        pdf.cell(0, 8, f"TO:  {entry.loc_to.loc_name}   DISPATCH:  {entry.car_no}", ln=1, align="L")

                        # Table header
                        pdf.set_xy(10, 40)
                        pdf.set_font("Arial", 'B', 11)
                        col_widths = [15, 35, 137, 30, 30, 30]
                        headers = ["SN", "ITEM CODE", "DESCRIPTION","UOM", "SENT", "RECIEVED"]
                        for i, header in enumerate(headers):
                            pdf.cell(col_widths[i], 10, header, border=1, align="C")
                        pdf.ln()
                        return col_widths

                    # Custom class to override header
                    from fpdf import FPDF as _FPDF
                    class WaybillPDF(_FPDF):
                        def header(self):
                            # Save current y to restore after header
                            y = self.get_y()
                            draw_waybill_header(self)
                            # Set y to just below the table header
                            self.set_y(50)
                        def footer(self):
                            # Move 40mm from bottom
                            
                            self.set_y(-40)
                            self.set_font("Arial", 'i', 8)
                            page_width = self.w - self.l_margin - self.r_margin
                            col_footer_width = page_width / 3
                            # Draw signature lines
                            self.cell(page_width,8,"Acraa, Osu, +233 24 350 0687, bharat@snedaghana.com",ln=1,align='C')
                            self.set_font("Arial", 'B', 11)
                            self.cell(col_footer_width, 8, "Sender", border="LTR", align="C")
                            self.cell(col_footer_width, 8, "Delivered By", border="LTR", align="C")
                            self.cell(col_footer_width, 8, "Received By", border="LTR", align="C")
                            self.ln(8)
                            
                            self.set_font("Arial", '', 11)
                            self.cell(col_footer_width, 8, f"{entry.obj().get('sent_by')}", border="LR", align="C")
                            self.cell(col_footer_width, 8, f"{entry.obj().get('transporter')}", border="LR", align="C")
                            self.cell(col_footer_width, 8, "", border="LR", align="C")
                            self.ln(1)
                            
                            self.set_font("Arial", '', 11)
                            self.cell(col_footer_width, 12, "_________________________", border="LRB", align="C")
                            self.cell(col_footer_width, 12, "_________________________", border="LRB", align="C")
                            self.cell(col_footer_width, 12, "_________________________", border="LRB", align="C")
                            self.ln(12)

                    pdf = WaybillPDF(orientation='L', unit='mm', format='A4')
                    pdf.set_auto_page_break(auto=True, margin=40)
                    pdf.add_page()
                    col_widths = [15, 35, 137, 30, 30, 30]  # For use in data rows

                    # Dummy data rows
                    pdf.set_font("Arial", '', 11)
                    
                    dummy_data = [
                        # [1, "SM-1001", "Single Phase Smart Meter", 10, "10", ""],
                    ]

                    line = 1
                    for tran in entry.transactions():
                        dummy_data.append(
                            [line,tran.get('barcode'),tran.get('name'),tran.get('oum'),tran.get('total_qty'),'']
                        )
                        line += 1
                    for row in dummy_data:
                        for i, item in enumerate(row):
                            align = "C" if i in [0,3,4] else "L"
                            pdf.cell(col_widths[i], 10, str(item), border=1, align=align)
                        pdf.ln()

                    # Fill up to 12 rows for visual
                    for _ in range(12 - len(dummy_data)):
                        for w in col_widths:
                            pdf.cell(w, 10, "", border=1)
                        pdf.ln()

                    # Footer lines
                    

                    # Footer fields
                    # Footer with three columns: Sender, Delivered By, Received By, on every page
                    class WaybillPDF(_FPDF):
                        def header(self):
                            y = self.get_y()
                            draw_waybill_header(self)
                            self.set_y(50)
                        
                    

                    file_name = f"{entry_no}.pdf"
                    file = f"static/general/tmp/{file_name}"
                    pdf.output(file)

                    success_response['message'] = file


                elif module == 'cont_issued':
                    cont_pk = data.get('contractor','*')
                    if cont_pk == '*':
                        issueds = Issue.objects.all().order_by('issue_date')
                    else:
                        issueds = Issue.objects.filter(contractor_id=cont_pk).order_by('issue_date')

                    

                    if data.get('type') == 'ISS':
                        issueds = issueds.filter(issue_type='ISS')
                    elif data.get('type') == 'RET':
                        issueds = issueds.filter(issue_type='RET')

                    success_response['message'] = [
                        {"pk":issue.id,'entry':issue.issue_no,'remarks':issue.remarks,'date':issue.issue_date,'is_posted':issue.is_posted,'deleted':issue.is_deleted,'total_qty':issue.total_qty()}
                     for issue in issueds]

                elif module == 'export_installations':
                    service_type = data.get('service_type','*')
                    print(data)
                    if service_type == '*':
                        installations = ServiceOrder.objects.all()
                    else:
                        st = ServiceType.objects.get(id=service_type)
                        installations = ServiceOrder.objects.filter(service_type=st)
                    
                    doc = data.get('doc','JSON')
                    if doc == 'JSON':
                        success_response['message'] = [installation.obj() for installation in installations]
                    elif doc == 'EXCEL':
                        import openpyxl
                        book = openpyxl.Workbook()
                        sheet = book.active
                        sheet.title = "Installations"
                        sheet.append(['Service Type', 'Service Date', 'Contractor', 'Plot', 'Geo Code', 'Customer', 'Customer No', 'Old Meter No', 'New Meter No', 'Old Meter No Reading', 'New Meter No Reading'])
                        for installation in installations:
                            sheet.append([installation.service_type.name, installation.service_date, installation.contractor.company, installation.plot.plot_no, installation.geo_code, installation.customer, installation.customer_no, installation.old_meter_no, installation.new_meter, installation.old_meter_no_reading, installation.new_meter_no_reading])
                        file_name = f"static/general/tmp/Installations-{st.name}.xlsx"
                        book.save(file_name)
                        success_response['message'] = f'{file_name}'

                elif module == 'cardex':
                    pk = data.get('pk')
                    item = InventoryMaterial.objects.get(id=pk)
                    loc_id = data.get('location','*')
                    if loc_id == '*':
                        file_name = 'all_locations'
                    else:
                        file_name = Location.objects.get(id=loc_id).loc_name
                    
                    cd = item.cardex(loc_id)

                    # make excel
                    import openpyxl
                    book = openpyxl.Workbook()
                    sheet = book.active
                    sheet.title = "Cardex"
                    sheet.append(['Doc No', 'Doc Type', 'Qty',"Balance", 'Created At'])
                    for row in cd:
                        sheet.append([row.get('doc_no'),row.get('doc_type'),row.get('qty'),row.get('balance'),row.get('created_at')])

                    import hashlib
                    name = hashlib.md5(item.name.encode()).hexdigest()
                    file_name = f"static/general/tmp/{file_name}-Cardex-{name}.xlsx"
                    book.save(file_name)
                   

                    arr = {
                        'json':cd,
                        'excel':file_name
                    }
                    success_response['message'] = arr
                    response = success_response

                elif module == 'group':
                    id = data.get('id','*')
                    if id == '*':
                        groups = InventoryGroup.objects.all()
                        success_response['message'] = [group.obj() for group in groups]
                    else:
                        group = InventoryGroup.objects.get(id=id)
                        success_response['message'] = group.obj()
                elif module == 'material':
                    id = data.get('id','*')
                    if id == '*':
                        materials = InventoryMaterial.objects.all().order_by('name')
                        success_response['message'] = [material.obj() for material in materials]
                    else:
                        if str(id).isdigit():
                            material = InventoryMaterial.objects.get(id=id)
                            success_response['message'] = material.obj()
                        else:
                            material = InventoryMaterial.objects.filter(Q(barcode__icontains=id) | Q(name__icontains=id))
                            filta = data.get('filter', None)
                            if filta is not None:
                                    if filta == 'is_return':
                                        material = material.filter(is_return=True)
                            if material.exists():
                                
                                success_response['message'] = [m.obj() for m in material]
                            else:
                                success_response['status_code'] = 404
                                success_response['message'] = f"No material with barcode or name matching {id} with filter {filta}"
                       

                elif module == 'meter':
                    meter_no=data.get('meter_no')
                    meter = Meter.objects.get(meter_no=meter_no)
                    arr = meter.obj()
                    print(meter.obj())

                    success_response['message'] = arr

                elif module == 'supplier':
                    id = data.get('id','*')
                    if id == '*':
                        suppliers = Supplier.objects.all()
                        success_response['message'] = [supplier.obj() for supplier in suppliers]
                    else:
                        supplier = Supplier.objects.get(id=id)
                        success_response['message'] = supplier.obj()

                elif module == 'grn':
                    id = data.get('id','*')
                    if id == '*':
                        grns = Grn.objects.all()
                        success_response['message'] = [grn.obj() for grn in grns]
                    else:
                        grn = Grn.objects.get(id=id)
                        success_response['message'] = grn.obj()

                elif module == 'plot':
                    id = data.get('id','*')
                    if id == '*':
                        plots = Plot.objects.all()
                        success_response['message'] = [plot.obj() for plot in plots]
                    else:
                        plot = Plot.objects.get(id=id)
                        success_response['message'] = plot.obj()

                elif module == 'service_type':
                    id = data.get('id','*')
                    if id == '*':
                        service_types = ServiceType.objects.all()
                        success_response['message'] = [service_type.obj() for service_type in service_types]
                    else:
                        service_type = ServiceType.objects.get(id=id)
                        success_response['message'] = service_type.obj()

                elif module == 'service':
                    id = data.get('id','*')
                    if id == '*':
                        services = Service.objects.all()
                        success_response['message'] = [service.obj() for service in services]
                    else:
                        service = Service.objects.get(id=id)
                        success_response['message'] = service.obj() 
                elif module == 'print_grn':
                    grn = Grn.objects.get(id=data.get('id'))

                    from fpdf import FPDF
                    pdf = FPDF()
                    pdf.add_page()

                    pdf.set_font("Arial","B", size=20)
                    pdf.cell(190, 5, txt="GOODS RECIEVED NOTE", ln=True, border=0, align="C")
                    pdf.ln(10)

                    pdf.set_font("Arial","B", size=12)
                    pdf.cell(20, 5, txt=f"Entry No: ", ln=False, align="L",border=0)
                    pdf.set_font("Arial", size=10)
                    pdf.cell(75, 5, txt=f"{grn.grn_no}", ln=False, align="L",border=0)
                    
                    pdf.set_font("Arial", "B", size=10)
                    pdf.cell(75, 5, txt="Supplier Code:", ln=False, align="R",border=0)
                    pdf.set_font("Arial", size=10)
                    pdf.cell(20, 5, txt=f"{grn.supplier.id}", ln=True, align="R",border=0)

                    pdf.set_font("Arial", "B", size=10)
                    pdf.cell(20, 5, txt=f"GRN Date: ", ln=False, align="L",border=0)
                    pdf.set_font("Arial", size=10)
                    pdf.cell(75, 5, txt=f"{grn.grn_date}", ln=False, align="L",border=0)
                    
                    pdf.set_font("Arial", "B", size=10)
                    pdf.cell(75, 5, txt="", ln=False, align="R",border=0)
                    pdf.set_font("Arial", size=10)
                    pdf.cell(20, 5, txt=f"{grn.supplier.supplier_name}", ln=True, align="R",border=0)

                    pdf.set_font("Arial", "B", size=10)
                    pdf.cell(20, 5, txt=f"Remarks: ", ln=False, align="L",border=0)
                    pdf.set_font("Arial", size=10)
                    pdf.cell(75, 5, txt=f"{grn.remarks}", ln=False, align="L",border=0)

                    pdf.set_font("Arial", "B", size=10)
                    pdf.cell(75, 5, txt=f"", ln=False, align="R",border=0)
                    pdf.set_font("Arial", size=10)
                    pdf.cell(20, 5, txt=f"{grn.supplier.supplier_phone}", ln=True, align="R",border=0)

                    pdf.ln(10)
                    

                   

                    pdf.set_font("Arial", "B", size=10)
                    pdf.cell(10, 5, txt="LN", ln=False, border=1, align="L")
                    pdf.cell(25, 5, txt="Barcode", ln=False, border=1, align="L")
                    pdf.cell(50, 5, txt="Name", ln=False, border=1, align="L")
                    pdf.cell(18, 5, txt="UOM", ln=False, border=1, align="C")
                    pdf.cell(18, 5, txt="Pak. Qty", ln=False, border=1, align="C")
                    pdf.cell(18, 5, txt="Qty", ln=False, border=1, align="C")
                    pdf.cell(18, 5, txt="Tot. Qty", ln=False, border=1, align="C")
                    pdf.cell(18, 5, txt="Rate", ln=False, border=1, align="C")
                    pdf.cell(18, 5, txt="Tot. Amt", ln=True, border=1, align="C")
                    line = 1
                    total_qty = 0
                    total_amount = 0
                    pdf.set_font("Arial", size=10)
                    for transaction in GrnTransaction.objects.filter(grn=grn):
                        pdf.cell(10, 5, txt=str(line), ln=False, border=1, align="L")
                        pdf.cell(25, 5, txt=transaction.barcode, ln=False, border=1, align="L")
                        replacements = {
                            '\u2018': "'",
                            '\u2019': "'",
                            '\u201C': '"',
                            '\u201D': '"',
                            '\u2014': '-',
                        }


                        for orig, repl in replacements.items():
                            transaction.name = transaction.name.replace(orig, repl)

                        pdf.cell(50, 5, txt=transaction.name[:20], ln=False, border=1, align="L")
                        pdf.cell(18, 5, txt=transaction.uom, ln=False, border=1, align="C")
                        pdf.cell(18, 5, txt=str(transaction.pack_qty), ln=False, border=1, align="C")
                        pdf.cell(18, 5, txt=str(transaction.qty), ln=False, border=1, align="C")
                        pdf.cell(18, 5, txt=str(transaction.total_qty), ln=False, border=1, align="C")
                        pdf.cell(18, 5, txt=str(transaction.rate), ln=False, border=1, align="C")
                        pdf.cell(18, 5, txt=str(transaction.amount), ln=True, border=1, align="C")
                        line += 1
                        total_qty += transaction.total_qty
                        total_amount += transaction.amount

                    
                    pdf.ln(10)
                    pdf.set_font("Arial", size=10)
                    pdf.cell(170, 5, txt="Total Amount: ", ln=False, align="R",border=0)
                    pdf.cell(20, 5, txt=str(total_amount), ln=True, align="R",border=0)

                    
                    pdf.set_font("Arial", size=10)
                    pdf.cell(170, 5, txt="Total Qty: ", ln=False, align="R",border=0)
                    pdf.cell(20, 5, txt=str(total_qty), ln=True, align="R",border=0)

                    pdf.ln(10)
                    pdf.set_font("Arial", size=10)
                    pdf.cell(190 / 6, 10, txt=f"{grn.created_by.first_name} {grn.created_by.last_name}", ln=False, align="C",border=1)
                    pdf.cell(190 / 6, 10, txt="", ln=False, align="R",border=0)
                    pdf.cell(190 / 12, 10, txt="", ln=False, align="R",border=0)
                    pdf.cell(190 / 6, 10, txt=f"{grn.obj().get('posted_by')}", ln=False, align="R",border=1)
                    pdf.cell(190 / 12, 10, txt="", ln=False, align="R",border=0)
                    pdf.cell(190 / 6, 10, txt="", ln=False, align="R",border=0)
                    pdf.cell(190 / 6, 10, txt="", ln=True, align="R",border=1)
                    
                    pdf.cell(190 / 6, 10, txt="Created By", ln=False, align="C",border=0)
                    pdf.cell(190 / 6, 10, txt="", ln=False, align="R",border=0)
                    pdf.cell(190 / 12, 10, txt="", ln=False, align="R",border=0)
                    pdf.cell(190 / 6, 10, txt="Posted By", ln=False, align="C",border=0)
                    pdf.cell(190 / 12, 10, txt="", ln=False, align="R",border=0)
                    pdf.cell(190 / 6, 10, txt="", ln=False, align="R",border=0)
                    pdf.cell(190 / 6, 10, txt="Confirmed By", ln=True, align="C",border=0)

                        
                    file_name = f"GRN-{grn.grn_no}.pdf"
                    file = f"static/general/tmp/{file_name}"
                    pdf.output(file)
                    success_response['message'] = file

                elif module == 'print_issue':
                    issue = Issue.objects.get(id=data.get('id'))

                    from fpdf import FPDF
                    pdf = FPDF()
                    pdf.add_page()  
                    
                    pdf.set_font("Arial","",10)
                    pdf.cell(190, 5, txt="Sneda Smart Meters Limited", ln=True, border=0, align="C")
                    pdf.cell(190, 5, txt="P.O. Box 1024, Accra, Ghana", ln=True, border=0, align="C")
                    pdf.cell(190, 5, txt="Tel: +233 54 200 3000 / 55 505 9999", ln=True, border=0, align="C")
                    pdf.cell(190, 5, txt="Email: bharat@snedaghana.com", ln=True, border=0, align="C")
                    pdf.ln(10)

                    pdf.set_font("Arial","B", size=20)
                    pdf.cell(190, 5, txt="GOODS ISSUE WAYBILL", ln=True, border=0, align="C")
                    pdf.ln(10)

                    pdf.set_font("Arial","B", size=12)
                    pdf.cell(90, 5, txt=f"Document", ln=False, align="L",border=0)
                    
                    
                    pdf.set_font("Arial", "B", size=12)
                    pdf.cell(100, 5, txt="Contactor", ln=True, align="R",border=0)

                    pdf.set_font("Arial",'B', size=10)
                    pdf.cell(20, 5, txt=f"Entry: ", ln=False, align="L",border=0)
                    pdf.set_font("Arial",'', size=10)
                    pdf.cell(75, 5, txt=f"{issue.issue_no}", ln=False, align="L",border=0)

                    pdf.set_font("Arial",'B', size=10)
                    pdf.cell(75, 5, txt=f"", ln=False, align="R",border=0)
                    pdf.set_font("Arial",'', size=10)
                    pdf.cell(20, 5, txt=f"{issue.contractor.company}", ln=True, align="R",border=0)

                    pdf.set_font("Arial",'B', size=10)
                    pdf.cell(20, 5, txt=f"Date: ", ln=False, align="L",border=0)
                    pdf.set_font("Arial",'', size=10)
                    pdf.cell(75, 5, txt=f"{issue.issue_date}", ln=False, align="L",border=0)

                    pdf.set_font("Arial",'B', size=10)
                    pdf.cell(75, 5, txt=f"", ln=False, align="R",border=0)
                    pdf.set_font("Arial",'', size=10)
                    pdf.cell(20, 5, txt=f"{issue.contractor.owner}", ln=True, align="R",border=0)

                    pdf.set_font("Arial",'B', size=10)
                    pdf.cell(20, 5, txt=f"", ln=False, align="L",border=0)
                    pdf.set_font("Arial",'', size=10)
                    pdf.cell(75, 5, txt=f"", ln=False, align="L",border=0)

                    pdf.set_font("Arial",'B', size=10)
                    pdf.cell(75, 5, txt=f"Phone: ", ln=False, align="R",border=0)
                    pdf.set_font("Arial",'', size=10)
                    pdf.cell(20, 5, txt=f"{issue.contractor.phone}", ln=True, align="R",border=0)

                    pdf.set_font("Arial",'B', size=10)
                    pdf.cell(20, 5, txt=f"", ln=False, align="L",border=0)
                    pdf.set_font("Arial",'', size=10)
                    pdf.cell(75, 5, txt=f"", ln=False, align="L",border=0)

                    pdf.set_font("Arial",'B', size=10)
                    pdf.cell(75, 5, txt=f"Car No: ", ln=False, align="R",border=0)
                    pdf.set_font("Arial",'', size=10)
                    pdf.cell(20, 5, txt=f"{issue.car_no}", ln=True, align="R",border=0)

                    pdf.set_font("Arial",'B', size=10)
                    pdf.cell(20, 5, txt=f"", ln=False, align="L",border=0)
                    pdf.set_font("Arial",'', size=10)
                    pdf.cell(75, 5, txt=f"", ln=False, align="L",border=0)

                   
                    
                    pdf.ln(10)

                    pdf.set_font("Arial", "B", size=8)
                    pdf.cell(10, 5, txt="LN", ln=False, border=1, align="L")
                    pdf.cell(25, 5, txt="Barcode", ln=False, border=1, align="L")
                    pdf.cell(65, 5, txt="Name", ln=False, border=1, align="L")
                    pdf.cell(15, 5, txt="UOM", ln=False, border=1, align="C")
                    pdf.cell(15, 5, txt="Pak. Qty", ln=False, border=1, align="C")
                    pdf.cell(15, 5, txt="Qty", ln=False, border=1, align="C")
                    pdf.cell(15, 5, txt="Tot. Qty", ln=False, border=1, align="C")
                    pdf.cell(15, 5, txt="Rate", ln=False, border=1, align="C")
                    pdf.cell(15, 5, txt="Tot. Amt", ln=True, border=1, align="C")

                    pdf.set_font("Arial", size=8)
                    ln = 1
                    for transaction in IssueTransaction.objects.filter(issue=issue):
                        pdf.cell(10, 5, txt=str(ln), ln=False, border=1, align="L")
                        pdf.cell(25, 5, txt=transaction.barcode, ln=False, border=1, align="L")
                        pdf.cell(65, 5, txt=transaction.name[:60], ln=False, border=1, align="L")
                        pdf.cell(15, 5, txt=transaction.uom, ln=False, border=1, align="C")
                        pdf.cell(15, 5, txt=str(transaction.pack_qty), ln=False, border=1, align="C")
                        pdf.cell(15, 5, txt=str(transaction.qty), ln=False, border=1, align="C")
                        pdf.cell(15, 5, txt=str(transaction.total_qty), ln=False, border=1, align="C")
                        pdf.cell(15, 5, txt=str(transaction.rate), ln=False, border=1, align="C")
                        pdf.cell(15, 5, txt=str(transaction.amount), ln=True, border=1, align="C")
                        ln += 1

                    if Meter.objects.filter(issue=issue).exists():
                        pdf.multi_cell(190, 10, txt="Meters: [" + ', '.join([meter.meter_no for meter in Meter.objects.filter(issue=issue)]) + "]",border=True, align="L")
                    pdf.ln(10)
                    
                    pdf.set_font("Arial", size=10)
                    pdf.cell(170, 5, txt="Total Amount: ", ln=False, align="R",border=0)
                    pdf.cell(20, 5, txt=str(issue.obj().get('total_amount')), ln=True, align="R",border=0)

                    pdf.set_font("Arial", size=10)
                    pdf.cell(170, 5, txt="Total Qty: ", ln=False, align="R",border=0)
                    pdf.cell(20, 5, txt=str(issue.obj().get('total_qty')), ln=True, align="R",border=0)

                    pdf.ln(10)
                    pdf.set_font("Arial", size=10)
                    pdf.cell(190 / 6, 10, txt=f"{issue.created_by.first_name} {issue.created_by.last_name}", ln=False, align="C",border=1)
                    pdf.cell(190 / 6, 10, txt="", ln=False, align="R",border=0)
                    pdf.cell(190 / 12, 10, txt="", ln=False, align="R",border=0)
                    pdf.cell(190 / 6, 10, txt=f"{issue.obj().get('posted_by')}", ln=False, align="C",border=1)
                    pdf.cell(190 / 12, 10, txt="", ln=False, align="R",border=0)
                    pdf.cell(190 / 6, 10, txt="", ln=False, align="R",border=0)
                    pdf.cell(190 / 6, 10, txt="", ln=True, align="R",border=1)
                    
                    pdf.cell(190 / 6, 10, txt="Created By", ln=False, align="C",border=0)
                    pdf.cell(190 / 6, 10, txt="", ln=False, align="R",border=0)
                    pdf.cell(190 / 12, 10, txt="", ln=False, align="R",border=0)
                    pdf.cell(190 / 6, 10, txt="Posted By", ln=False, align="C",border=0)
                    pdf.cell(190 / 12, 10, txt="", ln=False, align="R",border=0)
                    pdf.cell(190 / 6, 10, txt="", ln=False, align="R",border=0)
                    pdf.cell(190 / 6, 10, txt="Confirmed By", ln=True, align="C",border=0)
                    
                    file_name = f"ISSUE-{issue.issue_no}.pdf"
                    file = f"static/general/tmp/{file_name}"
                    pdf.output(file)
                    success_response['message'] = file

                elif module == 'print_services':
                    from fpdf import FPDF
                    pdf = FPDF()
                    pdf.add_page()
                    
                    pdf.set_font("Arial","B", size=12)
                    pdf.cell(190, 5, txt="Services", ln=True, border=0, align="C")
                    pdf.ln(10)
                    
                    
                    pdf.cell(20, 5, txt="Code", ln=False, align="L",border=0)
                    pdf.cell(50, 5, txt="Name", ln=False, align="L",border=0)
                    pdf.cell(50, 5, txt="Description", ln=False, align="L",border=0)
                    pdf.cell(20, 5, txt="Rate", ln=True, align="L",border=0)
                    pdf.ln(10)
                    
                    services = Service.objects.all()
                    for service in services:
                        pdf.set_font("Arial","B", size=10)
                        pdf.cell(20, 5, txt=f"{service.code}", ln=False, align="L",border=0)
                        pdf.cell(50, 5, txt=f"{service.name}", ln=False, align="L",border=0)
                        pdf.cell(50, 5, txt=f"{service.description}", ln=False, align="L",border=0)
                        pdf.cell(20, 5, txt=f"{service.rate}", ln=True, align="L",border=0)
                        pdf.ln(5)

                    file_name = f"Services.pdf"
                    file = f"static/general/tmp/{file_name}"
                    pdf.output(file)
                    success_response['message'] = file
                    
                    

                elif module == 'issue':
                    id = data.get('id','*')
                    find_by = data.get('find_by','id')
                    if id == '*':
                        issues = Issue.objects.all()
                        success_response['message'] = [issue.obj() for issue in issues]
                    else:
                        if find_by == 'id':
                            issue = Issue.objects.get(id=id)
                        else:
                            issue = Issue.objects.get(issue_no=id)
                        success_response['message'] = issue.obj()

                elif module == 'issue_meters':
                    issue_id = data.get('id')
                    issue = Issue.objects.get(id=issue_id)
                    meters = Meter.objects.filter(issue=issue)
                    success_response['message'] = [meter.obj() for meter in meters]

                elif module == 'service_order':
                    id = data.get('id','*')
                    filter = data.get('filter','*')
                    status = data.get('status','pending')
                    print(data)
                    if id == '*':
                        service_orders = ServiceOrder.objects.all()[:10]
                        if filter == 'contractor':
                            limit = data.get('filter',1000)
                            service_orders = ServiceOrder.objects.filter(contractor=data.get('contractor'),status=status)
                            status = data.get('status')
                            
                        elif filter == 'plot':
                            service_orders = ServiceOrder.objects.filter(plot=data.get('plot'))[:10]
                        elif filter == 'service_type':
                            service_orders = ServiceOrder.objects.filter(service_type=data.get('service_type'))[:10]
                            
                        success_response['message'] = [service_order.obj() for service_order in service_orders]
                    else:
                        service_order = ServiceOrder.objects.filter(Q(id=id) | Q(new_meter=id)).first()
                        success_response['message'] = [service_order.obj()]
                
                elif module == 'service_order_item':
                    print(data)
                    service_order_id = data.get('service_order_id')
                    if service_order_id == '*':
                        service_order_items = ServiceOrderItem.objects.all()
                        success_response['message'] = [service_order_item.obj() for service_order_item in service_order_items]
                    else:
                        service_order_item = ServiceOrderItem.objects.get(id=service_order_id)
                        success_response['message'] = service_order_item.obj()
                
                elif module == 'service_order_return':
                    id = data.get('id')
                    if id == '*':
                        return_items = ServiceOrderReturns.objects.all()
                        success_response['message'] = [return_item.obj() for return_item in return_items]
                    else:
                        return_item = ServiceOrderReturns.objects.get(id=id)
                        success_response['message'] = return_item.obj()
                
                elif module == 'reedem':
                    pk = data.get('pk')
                    entry = Reedem.objects.get(pk=pk)

                    if pk == '*':
                        redeems = Reedem.objects.all()
                        success_response['message'] = [redeem.obj() for redeem in redeems]
                    else:
                        redeem = Reedem.objects.get(pk=pk)
                        success_response['message'] = redeem.obj()

                elif module == 'daily_report':
                    from django.utils import timezone
                    from datetime import datetime
                    from .models import ServiceOrder
                    print(data)
                    day_date = data.get('date',datetime.now().date())
                    print(day_date)
                    from fpdf import FPDF
                    pdf = FPDF('P','mm','A4')
                    pdf.add_page()
                    pdf.set_font("Arial","B", size=12)
                    pdf.cell(190, 5, txt=f"Daily Report {day_date}", ln=True, border=0, align="C")
                    pdf.ln(5)
                    
                    pdf.set_font("Arial","B", size=10)
                    pdf.cell(100, 5, txt="Service Type", ln=False, align="L",border=1)
                    pdf.cell(45, 5, txt="Today Jobs", ln=False, align="L",border=1)
                    pdf.cell(45, 5, txt="Total Jobs", ln=True, align="L",border=1)

                    total_today_jobs = 0
                    total_total_jobs = 0
                    
                    for st in ServiceType.objects.all():
                        
                        today_jobs = st.today_jobs(day_date)
                        total_jobs = st.total_installations(day_date)

                        print("TOTAL JOBS",total_jobs)
                        
                        pdf.set_font("Arial","", size=10)
                        pdf.cell(100, 5, txt=f"{st.name}", ln=False, align="L",border=1)
                        pdf.cell(45, 5, txt=f"{today_jobs}", ln=False, align="L",border=1)
                        pdf.cell(45, 5, txt=f"{total_jobs}", ln=True, align="L",border=1)
                        
                        total_today_jobs += today_jobs
                        total_total_jobs += total_jobs

                    pdf.set_font("Arial","B", size=10)
                    pdf.cell(100, 5, txt="Total", ln=False, align="L",border=1)
                    pdf.cell(45, 5, txt=f"{total_today_jobs}", ln=False, align="L",border=1)
                    pdf.cell(45, 5, txt=f"{total_total_jobs}", ln=True, align="L",border=1)

                    # contractor wise
                    pdf.ln(5)
                    pdf.set_font("Arial","B", size=10)
                    pdf.cell(190, 5, txt="Contractor Wise", ln=True, border=0, align="C")

                    pdf.set_font("Arial","B", size=10)
                    pdf.cell(100, 5, txt="Contractor", ln=False, align="L",border=1)
                    pdf.cell(45, 5, txt="Today Jobs", ln=False, align="L",border=1)
                    pdf.cell(45, 5, txt="Total Jobs", ln=True, align="L",border=1)

                    total_today_jobs = 0
                    total_total_jobs = 0
                    print("Cash")
                    for contractor in Contractor.objects.filter(~Q(company="A TRADITION")):
                        today_jobs = ServiceOrder.objects.filter(contractor=contractor,service_date=day_date).count()
                        total_jobs = ServiceOrder.objects.filter(contractor=contractor, service_date__lte=day_date).count()

                        pdf.set_font("Arial","", size=10)   
                        pdf.cell(100, 5, txt=f"{contractor.company[:30]}", ln=False, align="L",border=1)
                        pdf.cell(45, 5, txt=f"{today_jobs}", ln=False, align="L",border=1)
                        pdf.cell(45, 5, txt=f"{total_jobs}", ln=True, align="L",border=1)

                        total_today_jobs += today_jobs
                        total_total_jobs += total_jobs

                    pdf.set_font("Arial","B", size=10)
                    pdf.cell(100, 5, txt="Total", ln=False, align="L",border=1)
                    pdf.cell(45, 5, txt=f"{total_today_jobs}", ln=False, align="L",border=1)
                    pdf.cell(45, 5, txt=f"{total_total_jobs}", ln=True, align="L",border=1)


                    # contractor errors
                    pdf.ln(5)
                    pdf.set_font("Arial","B", size=10)
                    pdf.cell(190, 5, txt="Contractor Errors", ln=True, border=0, align="C")

                    pdf.set_font("Arial","B", size=10)  
                    pdf.cell(100, 5, txt="Contractor", ln=False, align="L",border=1)
                    pdf.cell(45, 5, txt="Errors", ln=False, align="L",border=1)
                    pdf.cell(45, 5, txt="Weekly Margin Reach", ln=True, align="L",border=1)
                    

                    total_errors = 0
                    
                    for contractor in Contractor.objects.filter(~Q(company="A TRADITION")):
                        errors = ContractorError.objects.filter(contractor=contractor)
                        total_errors += errors.count()

                        pdf.set_font("Arial","", size=10)
                        pdf.cell(100, 5, txt=f"{contractor.company[:30]}", ln=False, align="L",border=1)
                        pdf.cell(45, 5, txt=f"{errors.count()}", ln=False, align="L",border=1)

                        week_limit = contractor.error_limit()['week']
                        error_margin = f"{(errors.count() / week_limit * 100):.2f}%" if week_limit > 0 else "0.00%"

                        pdf.cell(45, 5, txt=f"{error_margin}", ln=True, align="L",border=1)

                    pdf.set_font("Arial","B", size=10)
                    pdf.cell(100, 5, txt="Total", ln=False, align="L",border=1)
                    pdf.cell(90, 5, txt=f"{total_errors}", ln=True, align="L",border=1)
                    
                    
                    
                    
                    

                    
                    
                    

                    # issued materials
                    pdf.ln(5)
                    pdf.set_font("Arial","B", size=10)
                    pdf.cell(190, 5, txt="Issued Materials", ln=True, border=0, align="C")
                   

                    pdf.set_font("Arial","B", size=10)
                    pdf.cell(100, 5, txt="Material", ln=False, align="L",border=1)
                    pdf.cell(25, 5, txt="Total Qty", ln=False, align="L",border=1)
                    pdf.cell(30, 5, txt="Estmated Value", ln=False, align="L",border=1)
                    pdf.cell(30, 5, txt="Total Amount", ln=True, align="L",border=1)

                    pdf.set_font("Arial","", size=10)

                    total = 0
                    tq = 0
                    total_value = 0

                    materials = InventoryMaterial.objects.filter(barcode__in=IssueTransaction.objects.filter(issue__issue_date=timezone.now().date(),issue__issue_type='ISS').values_list('barcode', flat=True))
                    from django.db.models import Sum
                    for material in materials:
                        total_qty = IssueTransaction.objects.filter(material=material,issue__issue_date=timezone.now().date()).aggregate(total_qty=Sum('total_qty'))['total_qty'] or 0
                        total_amount = total_qty * material.value
                        pdf.cell(100, 5, txt=f"{material.name}", ln=False, align="L",border=1)
                        pdf.cell(25, 5, txt=f"{total_qty}", ln=False, align="L",border=1)
                        pdf.cell(30, 5, txt=f"{'{:,.2f}'.format(material.value)}", ln=False, align="L",border=1)
                        pdf.cell(30, 5, txt=f"{'{:,.2f}'.format(total_amount)}", ln=True, align="L",border=1)

                        total += total_amount
                        tq += total_qty
                        total_value += material.value

                    pdf.set_font("Arial","B", size=10)
                    pdf.cell(100, 5, txt="Total", ln=False, align="L",border=1)
                    pdf.cell(25, 5, txt=f"{tq}", ln=False, align="L",border=1)
                    pdf.cell(30, 5, txt=f"{'{:,.2f}'.format(total_value)}", ln=False, align="L",border=1)
                    pdf.cell(30, 5, txt=f"{'{:,.2f}'.format(total)}", ln=True, align="L",border=1)


                    # returned materials
                    pdf.ln(5)
                    pdf.set_font("Arial","B", size=10)
                    pdf.cell(190, 5, txt="Returned Materials", ln=True, border=0, align="C")

                    pdf.set_font("Arial","B", size=10)
                    pdf.cell(100, 5, txt="Material", ln=False, align="L",border=1)
                    pdf.cell(25, 5, txt="Total Qty", ln=False, align="L",border=1)
                    pdf.cell(30, 5, txt="Estmated Value", ln=False, align="L",border=1)
                    pdf.cell(30, 5, txt="Total Amount", ln=True, align="L",border=1)

                    total = 0
                    tq = 0
                    total_value = 0

                    pdf.set_font("Arial","", size=10)

                    materials = InventoryMaterial.objects.filter(barcode__in=IssueTransaction.objects.filter(issue__issue_date=timezone.now().date(),issue__issue_type='RET').values_list('barcode', flat=True))
                    for material in materials:
                        total_qty = IssueTransaction.objects.filter(material=material,issue__issue_date=timezone.now().date()).aggregate(total_qty=Sum('total_qty'))['total_qty'] or 0
                        total_amount = total_qty * material.value
                        pdf.cell(100, 5, txt=f"{material.name}", ln=False, align="L",border=1)
                        pdf.cell(25, 5, txt=f"{total_qty}", ln=False, align="L",border=1)
                        pdf.cell(30, 5, txt=f"{'{:,.2f}'.format(material.value)}", ln=False, align="L",border=1)
                        pdf.cell(30, 5, txt=f"{'{:,.2f}'.format(total_amount)}", ln=True, align="L",border=1)

                        total += total_amount
                        tq += total_qty
                        total_value += material.value

                    pdf.set_font("Arial","B", size=10)
                    pdf.cell(100, 5, txt="Total", ln=False, align="L",border=1)
                    pdf.cell(25, 5, txt=f"{tq}", ln=False, align="L",border=1)
                    pdf.cell(30, 5, txt=f"{'{:,.2f}'.format(total_value)}", ln=False, align="L",border=1)
                    pdf.cell(30, 5, txt=f"{'{:,.2f}'.format(total)}", ln=True, align="L",border=1)

                    # list all meters installed today
                    pdf.ln(5)
                    pdf.set_font("Arial","B", size=8)
                    pdf.cell(190, 5, txt="Meters Installed Today", ln=True, border=0, align="C")

                    pdf.set_font("Arial","B", size=10)
                    pdf.cell(5, 5, txt="LN.", ln=False, align="L",border=1)
                    pdf.cell(20, 5, txt="Meter No", ln=False, align="L",border=1)
                    pdf.cell(65, 5, txt="Customer", ln=False, align="L",border=1)
                    pdf.cell(30, 5, txt="Mob. ", ln=False, align="L",border=1)
                    pdf.cell(65, 5, txt="Contractor", ln=True, align="L",border=1)

                    pdf.set_font("Arial","", size=8)

                    meters = ServiceOrder.objects.filter(service_date=day_date)
                    ln = 1
                    for meter in meters:
                        pdf.cell(5, 5, txt=f"{ln}", ln=False, align="L",border=1)
                        pdf.cell(20, 5, txt=f"{meter.new_meter}", ln=False, align="L",border=1)
                        pdf.cell(65, 5, txt=f"{meter.customer[:40]}", ln=False, align="L",border=1)
                        pdf.cell(30, 5, txt=f"{meter.customer_no}", ln=False, align="L",border=1)
                        pdf.cell(65, 5, txt=f"{meter.contractor.company[:35]}", ln=True, align="L",border=1)
                        ln += 1

                    

                    
                    




                    

                    
                    
                    
                    file_name = f"Daily Report {day_date}.pdf"
                    file = f"static/general/tmp/{file_name}"
                    pdf.output(file)
                    success_response['message'] = file
                        
                        
                elif module == 'item_availability':
                    from django.utils import timezone
                    
                    as_of = data.get('as_of_field',timezone.now().date())
                    loc_id = data.get('loc_id','*')
                    stock_only = data.get('stock_only','no')
                    print(data)
                    if stock_only == 'yes':
                        stock_only = True
                    else:
                        stock_only = False

                    print(data)
                    print(as_of)
                    arr = []

                    from fpdf import FPDF
                    pdf = FPDF('P','mm','A4')
                    pdf.add_page()
                    pdf.set_font("Arial","B", size=12)
                    pdf.cell(190, 5, txt=f"Item Availability {as_of}", ln=True, border=0, align="C")
                    pdf.ln(10)

                    pdf.set_font("Arial","B", size=10)
                    pdf.cell(100, 5, txt="Material", ln=False, align="L",border=1)
                    pdf.cell(25, 5, txt="Stock Qty", ln=False, align="L",border=1)
                    pdf.cell(30, 5, txt="Estmated Value", ln=False, align="L",border=1)
                    pdf.cell(30, 5, txt="Total Amount", ln=True, align="L",border=1)

                    pdf.set_font("Arial","", size=10)

                    total = 0
                    tq = 0
                    total_value = 0

                    materials = InventoryMaterial.objects.all().order_by('name')
                    for material in materials:
                        total_qty = material.stock(as_of,loc_id)
                        print(total_qty,stock_only)
                        if stock_only == True and total_qty <= 0:
                            pass
                        else:
                            total_amount = total_qty * material.value
                            pdf.cell(100, 5, txt=f"{material.name[:30]}", ln=False, align="L",border=1)
                            pdf.cell(25, 5, txt=f"{total_qty}", ln=False, align="L",border=1)
                            pdf.cell(30, 5, txt=f"{'{:,.2f}'.format(material.value)}", ln=False, align="L",border=1)
                            pdf.cell(30, 5, txt=f"{'{:,.2f}'.format(total_amount)}", ln=True, align="L",border=1)

                            total += total_amount
                            tq += total_qty
                            total_value += material.value

                            arr.append({
                                'name':material.name,
                                'stock_qty':total_qty,
                                'value':material.value,
                                'total_amount':total_amount
                            })
                        

                    file_name = f"Item Availability As Of {as_of}.pdf"
                    file = f"static/general/tmp/{file_name}"
                    pdf.output(file)
                    success_response['message'] = {
                        'file':file,
                        'data':arr
                    }
                        
                

                elif module == 'contractor_wise':
                    from django.utils import timezone
                    from datetime import timedelta  
                    arr = []
                    pass
                        
                    success_response['message'] = arr


                else:
                    success_response['status_code'] = 400
                    success_response['message'] = f"Module Not Found {module}"



                response = success_response
        elif method == 'DELETE':
            if module == 'service':
                id = data.get('id')
                service = Service.objects.get(id=id)
                service.delete()
                success_response['message'] = "Service Deleted Successfully"

            elif module == 'delete_transfer':
                entry_no = data.get('entry_no')
                transfer = TransferHd.objects.get(entry_no=entry_no)
                actor = User.objects.get(pk=data.get('mypk'))


                
                TransferTr.objects.filter(entry=transfer).delete()
                Cardex.objects.filter(
                        doc_type='TRF',
                        doc_no=entry_no
                    ).delete()
                rm = transfer.remarks
                transfer.remarks = f"Deleted By {actor.username} | {rm}"
                transfer.is_valid = False
                transfer.is_sent=False
                transfer.is_posted=False
                transfer.sent_by = None
                transfer.approved_by = None
                transfer.car_no = ""
                transfer.transporter = ""

                transfer.save()
                success_response['message'] = "Transfer Deleted"
                response = success_response


            elif module == 'meter':
                id = data.get('id')
                meter = Meter.objects.get(id=id)
                meter.delete()
                success_response['message'] = "Meter Deleted Successfully"

            elif module == 'order_service':
                id = data.get('id')
                service_order_item = ServiceOrderItem.objects.get(id=id)
                service_order_item.delete()
                success_response['message'] = "Service Order Item Deleted Successfully"

            elif module == 'service_order':
                id = data.get('id')
                service_order = ServiceOrder.objects.get(id=id)
                meter_no = service_order.new_meter
                meter = Meter.objects.get(meter_no=meter_no)
                meter.is_issued = False
                meter.contractor = Contractor.objects.get(company='A TRADITION')
                meter.save()
                service_order.delete()
                success_response['message'] = "Service Order Deleted Successfully"

            elif module == 'issue':
                id = data.get('id')
                issue = Issue.objects.get(id=id)
                # delete transactions
                IssueTransaction.objects.filter(issue=issue).delete()
                issue.is_deleted = True
                issue.save()
                success_response['message'] = "Issue Deleted Successfully"

            elif module == 'grn':
                id = data.get('id')
                grn = Grn.objects.get(id=id)
                # delete transactions
                GrnTransaction.objects.filter(grn=grn).delete()
                Cardex.objects.filter(doc_no=grn.grn_no,doc_type='GR').delete()
                grn.is_deleted = True
                grn.save()
                success_response['message'] = "GRN Deleted Successfully"

            elif module == 'service_material':
                id = data.get('id')
                MaterialOrderItem.objects.get(id=id).delete()
                success_response['message'] = "Material Deleted"

            elif module == 'service_order_return':
                id = data.get('id')
                ServiceOrderReturns.objects.get(id=id).delete()
                success_response['message'] = "Return Deleted"

            elif module == 'reedem':
                id = data.get('id')
                redeem = Reedem.objects.get(id=id)
                RedeemTransactions.objects.filter(redeem=redeem).delete()
                redeem.is_valid = False
                redeem.save()

                success_response['message'] = "Reedem Deleted"

            elif module == 'material':
                id = data.get('id')
                material = InventoryMaterial.objects.get(id=id)

                # check cardex
                cardex = Cardex.objects.filter(material=material).count()
                if cardex > 0:
                    success_response['status_code'] = 400
                    success_response['message'] = "Material Cannot Be Deleted"
                else:
                    material.delete()
                    success_response['message'] = "Material Deleted"

            else:
                success_response['status_code'] = 400
                success_response['message'] = f"Module Not Found {module}"
        
        elif method == 'PATCH':
            if module == 'material':
                
                barcode = data.get('barcode')
                name = data.get('name')
                group = InventoryGroup.objects.get(id=data.get('group'))
                reorder_qty = data.get('reorder_qty')
                value = data.get('value')
                is_return = data.get('is_return')
                is_issue = data.get('is_issue')
                auto_issue = data.get('auto_issue')
                item_type = data.get('item_type')
                
                print(data)
                material = InventoryMaterial.objects.get(barcode=barcode)
                material.name = name
                material.group = group
                material.reorder_qty = reorder_qty
                material.value = value
                material.item_type = item_type
                material.is_return = is_return
                material.is_issue = is_issue
                material.issue_qty = data.get('issue_qty')
                material.auto_issue = auto_issue


                material.save()
                success_response['message'] = "Material Updated Successfully"

            elif module == 'location_default':
                print(data)
                Location.objects.all().update(is_default=False)
                lc = Location.objects.get(id=data.get('id'))
                lc.is_default = True
                lc.save()

                response = success_response

            elif module == 'stock_transfer':
                print(data)
                source = data.get('source')
                desination = data.get('destination')

                sr_prod = InventoryMaterial.objects.get(barcode=source)
                des_prod = InventoryMaterial.objects.get(barcode=desination)

                # Use try/except for each update to ensure one failure doesn't stop the rest
                try:
                    GrnTransaction.objects.filter(material=sr_prod).update(material=des_prod)
                except Exception as e:
                    print(f"Error updating GrnTransaction: {e}")
                try:
                    IssueTransaction.objects.filter(material=sr_prod).update(material=des_prod)
                except Exception as e:
                    print(f"Error updating IssueTransaction: {e}")
                try:
                    TransferTr.objects.filter(material=sr_prod).update(material=des_prod)
                except Exception as e:
                    print(f"Error updating TransferTr: {e}")
                try:
                    RequiredReturn.objects.filter(material=sr_prod).update(material=des_prod)
                except Exception as e:
                    print(f"Error updating RequiredReturn: {e}")
                try:
                    RedeemTransactions.objects.filter(material=sr_prod).update(material=des_prod)
                except Exception as e:
                    print(f"Error updating RedeemTransactions: {e}")
                try:
                    ServiceOrderReturns.objects.filter(material=sr_prod).update(material=des_prod)
                except Exception as e:
                    print(f"Error updating ServiceOrderReturns: {e}")
                try:
                    MaterialOrderItem.objects.filter(material=sr_prod).update(material=des_prod)
                except Exception as e:
                    print(f"Error updating MaterialOrderItem: {e}")
                try:
                    ServiceMaterials.objects.filter(material=sr_prod).update(material=des_prod)
                except Exception as e:
                    print(f"Error updating ServiceMaterials: {e}")
                try:
                    IssueTransaction.objects.filter(material=sr_prod).update(material=des_prod)
                except Exception as e:
                    print(f"Error updating IssueTransaction (2nd): {e}")
                try:
                    Cardex.objects.filter(material=sr_prod).update(material=des_prod)
                except Exception as e:
                    print(f"Error updating Cardex: {e}")

                if data.get('delete') == 'YES':
                    sr_prod.delete()
                

                # for tr in Cardex.objects.filter(material=sr_prod):
                #     dis = tr.obj()
                #     doc_type = dis.get('doc_type')

                #     doc_type_choices = [
                #         ('GR', 'GRN'), # IN
                #         ('ISS', 'ISSUE'), # OUT
                #         ('CIS', 'Contractor Issue Return'), # IN
                #         ('RET', 'RETURN'), # IN
                #         ('TRF',"Transfer")# IN OUT
                #     ]
                #     if doc_type == 'GR':
                #         GrnTransaction.objects.filter(entry_no=dis.get('doc_no'),material=sr_prod)
                #     print(dis)

            elif module == 'approve_expense':
                pk = data.get('pk','*')
                print(data)
                if pk == '*':
                    expenses = Expense.objects.filter(is_approved=False)
                else:
                    expenses = Expense.objects.filter(id=pk)
                for expense in expenses:

                    
                    owner = expense.created_by
                    owner_email = owner.email

                    # Compose email subject and body for approving expense
                    subject = f"Expense Approved: {expense.category} - {expense.amount}"
                    body = (
                        f"Dear {owner.get_full_name()},\n\n"
                        f"Your expense request has been approved.\n\n"
                        f"Details:\n"
                        f"Category: {expense.category}\n"
                        f"Amount: {expense.amount}\n"
                        f"Reference: {expense.reference}\n"
                        f"Description: {expense.description}\n"
                        f"Date: {expense.date}\n\n"
                        f"Thank you."
                    )

                    expense.is_approved = True
                    expense.save()

                    if MailSenders.objects.filter(address='solomon@snedaghana.com').exists():
                        sender = MailSenders.objects.get(address='solomon@snedaghana.com')
                    else:
                        sender = MailSenders.objects.get(is_default=True)

                    mail = MailQueues(sender=sender, subject=subject, body=body,
                                                recipient='solomon@snedaghana.com',
                                                cc='',mail_key=make_md5_hash(f"{subject}{body}"))
                    mail.save()

            elif module == 'update_cardex_tally':
                from django.db.models import Sum
                for material in InventoryMaterial.objects.all():
                    cardex_items = Cardex.objects.filter(material=material)
                    for ci in cardex_items:
                        created_at = ci.created_at
                        # get sum of material including this
                        # Get sum of material in cardex before this transaction including this
                        sum_qty = Cardex.objects.filter(
                            material=material,
                            created_at__lte=created_at
                        ).order_by('created_at').aggregate(total_qty=Sum('qty'))['total_qty'] or 0
                        ci.balance = sum_qty

                        print(material.name,"SET TO",sum_qty)
                        ci.save()

                response = success_response

            elif module == 'send_transfer':
                from django.utils import timezone
                entry_no = data.get('entry_no')
                car_no = data.get('car_no')
                driver = data.get('driver')
                mypk = data.get('mypk')

                sender = User.objects.get(pk=mypk)
                entry = TransferHd.objects.get(entry_no=entry_no)
                entry.is_sent = True
                entry.sent_by = sender
                entry.sent_date = timezone.now()
                entry.transporter = driver
                entry.car_no = car_no
                entry.save()

                success_response['message'] = "Transfer Sent"
                response = success_response


            elif module == 'update_return_as_of':
                tot = 0
                for service in ServiceOrder.objects.all():
                    
                    if RequiredReturn.objects.filter(service_type=service.service_type).exists():
                        
                        for record in RequiredReturn.objects.filter(service_type=service.service_type):
                            print(service.new_meter,record.service_type.name,"QUALIFIED")
                            try:
                                ServiceOrderReturns.objects.create(
                                    service_order=service,
                                    material=record.material,
                                    quantity=record.quantity,
                                    created_at=service.created_at,
                                    created_by=service.created_by

                                )
                                tot += 1
                            except Exception as e:
                                print(e)
                                pass
                    else:
                        print(service.new_meter,service.service_type.name,"NOT QUALIFIED")

                success_response['message'] = f"{tot} jobs updated"

            elif module == 'patch_locations':
                raise Exception("Disabled")
                #  Add field location to cardex
                location=Location.objects.get(loc_id='002')
                Cardex.objects.all().update(
                    location=location
                )
                # + Add field location to grn
                Grn.objects.all().update(
                    location=location
                )
                # + Add field location to issue
                Issue.objects.all().update(
                    location=location
                )
                # + Add field location to meter
                Meter.objects.all().update(
                    location=location
                )
                # + Add field location to reedem
                Reedem.objects.all().update(
                    location=location
                )
                # + Add field location to serviceorder
                ServiceOrder.objects.all().update(
                    location=location
                )

                success_response['message'] = "Default Locations Updated"


            elif module == 'meter_remove_issue':
                id = data.get('id')
                meter = Meter.objects.get(id=id)
                meter.is_issued = False
                meter.service_order = None
                meter.save()
                success_response['message'] = "Meter Updated Successfully"

            elif module == 'issue_meter':
                id = data.get('id')
                meter = Meter.objects.get(id=id)
                
                if data.get('status') == 'unassigned':
                    meter.is_issued = False
                    
                else:   
                    meter.is_issued = True
                    meter.contractor = Contractor.objects.get(id=data.get('contractor'))

                meter.save()
                success_response['message'] = "Meter Updated Successfully"

            elif module == 'issue_def_qty':
                updated = 0
                created = 0

                # for meterial in InventoryMaterial.objects.filter(is_issue=True,auto_issue=True):
                    
                #     issue_qty = meterial.issue_qty
                #     MaterialOrderItem.objects.filter(material=meterial).update(
                #         quantity=issue_qty,
                #         rate=meterial.value,
                #         amount=issue_qty*meterial.value
                #     )
                #     updated += 1

                #     mt = meterial

                #     print (mt.name, mt.issue_qty, '\n')

                if True:
                    for service in ServiceOrder.objects.all():
                        service_type = service.service_type
                        # get linked materials
                        
                        for mt in ServiceMaterials.objects.filter(service=service_type):
                            try:
                            # create
                                MaterialOrderItem.objects.create(
                                    material_type='is',
                                    material=mt.material,
                                    quantity=mt.qty,
                                    rate=mt.material.value,
                                    amount = mt.qty * mt.material.value ,
                                    service_order=service
                                )
                                print(mt.qty,'of',mt.material.name,'added to',service.new_meter)
                                created += 1
                            except Exception as e:
                                print(e)

                success_response['message'] = f"UPDATED {updated}, CREATED:{created}"

            elif module == 'post_invoice':
                entry_no = data.get('entry_no')
                invoice = InvoiceHD.objects.get(entry_no=entry_no)
                regs = 0
                reg_amt = 0

                reps = 0
                rep_amt = 0

                news = 0
                new_amt = 0

                

                for tran in invoice.transactions():
                    meter_no = tran.get('asset')
                    remarks = tran.get('remarks')

                    srv = ServiceOrder.objects.get(new_meter=meter_no)

                    tm = srv.total_amount()
                    

                    if remarks == 'Installation':
                        news += 1
                        new_amt += tm
                    
                    if remarks == 'Replacement':
                        reps += 1
                        rep_amt += tm

                    if remarks == 'Regularization':
                        regs += 1
                        reg_amt += tm

                ps_entry_no = f"SSML-PS-{PaySlipHD.objects.all().conut()}"
                ref = entry_no
                entry_date = data.get('entry_date',timezone.now())
                ps_type = 'CI'
                remarks = f"Payslip for {regs + reps + news} jobs done ({regs} Regularization, {reps} Replacements, {news} New Installations)"
                total_amount = reg_amt + rep_amt + new_amt


                from datetime import datetime, timedelta

                # Ensure entry_date is a date object
                if isinstance(entry_date, str):
                    entry_date_obj = datetime.strptime(entry_date, "%Y-%m-%d").date()
                else:
                    entry_date_obj = entry_date

                due_date = entry_date_obj + timedelta(days=60)
                created_by = User.objects.get(pk=data.get('mypk'))

                print("Regularization",regs,reg_amt)
                print("Replacement",reps,rep_amt)
                print("Installations",news,new_amt)

                # create header
                PaySlipHD.objects.create(
                    entry_no=ps_entry_no,
                    etry_date=entry_date,
                    ps_type=ps_type,
                    ref=entry_no,
                    remarks=remarks,
                    
                )

                print(reg_amt + rep_amt + new_amt)

                
                
            
            elif module == 'post_grn':
                grn_id = data.get('id')
                grn = Grn.objects.get(id=grn_id)
                mypk = data.get('mypk')
                user = User.objects.get(id=mypk)
                try:
                    transactions = GrnTransaction.objects.filter(grn=grn)
                    for transaction in transactions:
                        
                        Cardex.objects.create(
                            doc_type='GR',
                            doc_no=grn.grn_no,
                            ref_no=grn.id,
                            material=transaction.material,
                            qty=transaction.total_qty,
                            location=grn.location,
                            created_at=grn.grn_date,
                        )
                    
                    grn.is_posted = True
                    grn.posted_by = user
                    success_response['message'] = "GRN Posted Successfully"
                except Exception as e:
                    Cardex.objects.filter(doc_no=grn.grn_no,doc_type='GR').delete()

                    grn.is_posted = False
                    success_response['status_code'] = 400
                    success_response['message'] = f"Error on line {sys.exc_info()[2].tb_lineno}: {e}"
                finally:
                    grn.save()
            elif module == 'grn':
                grn_id = data.get('id')
                grn = Grn.objects.get(id=grn_id)
                grn.supplier = Supplier.objects.get(id=data.get('supplier'))
                grn.grn_no = data.get('grn_no')
                grn.grn_date = data.get('grn_date')
                grn.remarks = data.get('remarks')
                grn.grn_type = data.get('grn_type')

                transactions = GrnTransaction.objects.filter(grn=grn)
                transactions.delete()
                for transaction in data.get('transactions'):
                    material = InventoryMaterial.objects.get(barcode=transaction.get('barcode'))
                    GrnTransaction.objects.create(
                            grn=grn, 
                            material=material, 
                            barcode=transaction.get('barcode'),
                            name=transaction.get('name'),
                            amount=transaction.get('amount'),
                            pack_qty=transaction.get('pack_qty'),
                            uom=transaction.get('uom'),
                            rate=transaction.get('rate'),
                            qty=transaction.get('qty'),
                            total_qty=Decimal(transaction.get('qty')) * Decimal(transaction.get('pack_qty'))    
                        )
                
                success_response['message'] = "GRN Updated Successfully"

            elif module == 'post_transfer':
                entry_no = data.get('entry_no','*')
                if entry_no == '*':
                    transfers = TransferHd.objects.filter(is_valid=True,is_posted=False,is_sent=True)
                else:
                    transfers = TransferHd.objects.filter(is_valid=True,is_posted=False,is_sent=True,entry_no=entry_no)

                for tr in transfers:
                    loc_from = tr.loc_fr
                    loc_to = tr.loc_to
                    entry_date = tr.entry_date

                    for tran in tr.transactions():
                        print(tran)
                        material_id = tran.get('material_id')
                        print(material_id)
                        material = InventoryMaterial.objects.get(id=material_id)
                        total = Decimal(tran.get('sent_qty') ) * Decimal(tran.get('pack_qty'))

                        Cardex.objects.create(
                            doc_type='TRF',
                            doc_no = tr.entry_no,
                            ref_no = tr.entry_no,
                            material=material,
                            qty=total,
                            created_at=entry_date,
                            location=loc_to,
                        )

                        Cardex.objects.create(
                            doc_type='TRF',
                            doc_no = tr.entry_no,
                            ref_no = tr.entry_no,
                            material=material,
                            qty=total * Decimal(-1),
                            created_at=entry_date,
                            location=loc_from,
                        )
                    tr.is_posted = True
                    tr.save()

            elif module == 'unpost_transfer':
                entry_no = data.get('entry_no','*')
                if entry_no == '*':
                    transfers = TransferHd.objects.filter(is_valid=True,is_posted=True,is_sent=True)
                else:
                    transfers = TransferHd.objects.filter(is_valid=True,is_posted=True,is_sent=True,entry_no=entry_no)

                for tr in transfers:
                    entry = tr.entry_no

                    # delete from cardex
                    Cardex.objects.filter(
                        doc_type='TRF',
                        doc_no=entry
                    ).delete()

                    tr.is_posted = False
                    tr.save()

                response = success_response

            elif module == 'issue':
                issue_id = data.get('id')
                issue = Issue.objects.get(id=issue_id)
                issue.contractor = Contractor.objects.get(id=data.get('contractor'))
                issue.issue_date = data.get('issue_date')
                issue.remarks = data.get('remarks')
                issue.issue_type = data.get('issue_type')
                issue.car_no = data.get('car_no')
                issue.modified_by = User.objects.get(id=data.get('created_by'))

                transactions = IssueTransaction.objects.filter(issue=issue)
                transactions.delete()
                for transaction in data.get('transactions'):
                    material = InventoryMaterial.objects.get(barcode=transaction.get('barcode'))
                    IssueTransaction.objects.create(
                        issue=issue,
                        material=material,
                        qty=transaction.get('qty'),
                        pack_qty=transaction.get('pack_qty'),
                        uom=transaction.get('uom'),
                        rate=transaction.get('rate'),
                        barcode=transaction.get('barcode'),
                        name=transaction.get('name'),
                        total_qty=Decimal(transaction.get('qty')) * Decimal(transaction.get('pack_qty')),
                        amount=Decimal(transaction.get('qty')) * Decimal(transaction.get('rate'))
                    )
                issue.save()
                success_response['message'] = "Issue Updated Successfully"

            elif module == 'post_issue':
                
                try:
                    print(data)
                    issue_id = data.get('id','*')
                    if issue_id == '*':
                        issues = Issue.objects.filter(is_posted=False)
                    else:
                        issues = Issue.objects.filter(id=issue_id)
                    for issue in issues:
                        tp = issue.issue_type
                        mypk = data.get('mypk')
                        ct = 0
                        user = User.objects.get(id=mypk)
                        transactions = IssueTransaction.objects.filter(issue=issue)
                        
                        for transaction in transactions:
                            location = transaction.issue.location
                            ct += 1
                            qty = transaction.total_qty * -1
                            if tp == 'RET':
                                qty = transaction.total_qty
                            
                            if tp == 'ISS':
                                qty = qty

                            if tp == 'CIS':
                                qty = qty

                            Cardex.objects.create(
                                doc_type=tp,
                                doc_no=issue.issue_no,
                                ref_no=issue.id,
                                material=transaction.material,
                                qty=qty,
                                created_at=issue.issue_date,
                                location=location
                            )
                            issue.is_posted = True
                            issue.posted_by = user
                            issue.save()
                    success_response['message'] = f"Issue Posted Successfully {ct}"
                except Exception as e:
                    Cardex.objects.filter(doc_no=issue.issue_no,doc_type='IS').delete()
                    raise Exception(e)

                issue.save()
            elif module == 'unpost_issue':
                ct = 0
                issue_id = data.get('id','*')
                if issue_id == '*':
                    issues = Issue.objects.filter(is_posted=True)
                else:
                    issues = Issue.objects.filter(id=issue_id)
                
                for issue in issues:
                    ct += 1
                    issue.is_posted = False
                    issue.posted_by = None
                    Cardex.objects.filter(doc_no=issue.issue_no, doc_type__in=['ISS', 'RET']).delete()
                    issue.save()
                success_response['message'] = f"Issue Unposted Successfully {ct}"


            elif module == 'unpost_grn':
                grn_id = data.get('id')
                grn = Grn.objects.get(id=grn_id)
                grn.is_posted = False
                grn.posted_by = None
                grn.save()
                Cardex.objects.filter(doc_no=grn.grn_no,doc_type='GR').delete()
                success_response['message'] = "GRN Unposted Successfully"

                

            elif module == 'service':
                id = data.get('id')
                name = data.get('service_name')
                description = data.get('service_description')
                uom = data.get('service_uom')
                rate = data.get('service_rate')

                service = Service.objects.get(id=id)
                service.name = name
                service.description = description
                service.uom = uom
                service.rate = rate
                service.save()
                success_response['message'] = "Service Updated Successfully"
            elif module == 'service_order_item':
                
                service_order_item = ServiceOrderItem.objects.get(id=data.get('service_item_id'))
                service_order_item.quantity = data.get('serv_qty')
                user = User.objects.get(id=data.get('mypk'))
                service_order_item.modified_by = user
                service_order_item.amount = Decimal(data.get('serv_qty')) * Decimal(service_order_item.rate)
                service_order_item.save()
                success_response['message'] = "Service Order Item Updated Successfully"

            elif module == 'contractor':
                id = data.get('id')
                contractor = Contractor.objects.get(id=id)
                contractor.link = data.get('link')
                contractor.company = data.get('company')
                contractor.owner = data.get('owner')
                contractor.phone = data.get('phone')
                contractor.email = data.get('email')
                contractor.country = data.get('country')
                contractor.city = data.get('city')
                contractor.postal_code = data.get('postal_code')
                contractor.gh_post_code = data.get('gh_post_code')
                contractor.gh_card_no = data.get('gh_card_no')
                contractor.code = data.get('code')
                contractor.save()
        
            elif module == 'service_order_return':
                id = data.get('ret_id')
                return_item = ServiceOrderReturns.objects.get(id=id)
                return_item.quantity = data.get('ret_qty')
                return_item.modified_by = User.objects.get(id=data.get('mypk'))
                return_item.save()
                success_response['message'] = "Service Order Return Updated Successfully"
              
            elif module == 'close_service_order':
                id = data.get('id')
                mypk = data.get('mypk')
                user = User.objects.get(id=mypk)
                service_order = ServiceOrder.objects.get(id=id)
                total_amount = service_order.total_amount()
                print(total_amount)

                try:


                    Ledger.objects.create(
                        contractor=service_order.contractor,
                        amount=total_amount,
                        reference_no=service_order.new_meter,
                        transaction_type='credit',
                        remarks=f"Service Order {service_order.id} Closed",
                        created_by=user
                    )

                except Exception as e:
                    pass

                service_order.status = 'completed'
                service_order.closed_by = user
                service_order.total_amount = total_amount
                contractor_id = service_order.contractor.id
                service_order.save()
                success_response['message'] = contractor_id

            elif module == 'ret_def_rec':
                try:
                    service_id = data.get('service_id')
                    service = ServiceOrder.objects.get(id=service_id)
                    user = User.objects.get(pk=data.get('mypk'))
                    for service in ServiceOrder.objects.all():
                        for material in InventoryMaterial.objects.filter(is_return=True):
                            try:
                                ServiceOrderReturns.objects.create(
                                service_order=service,
                                material=material,
                                quantity=1,
                                created_by=user
                            )
                            except Exception as e:
                                print("Error")
                                print(e)
                                pass
                    success_response['status_code'] = 200
                    success_response['message'] = 'done'
                except Exception as e:
                    success_response['status_code'] = 505
                    success_response['message'] = str(e)

                print(success_response)

            elif module == "service_type":
                debit_account = data.get('debit_account')
                credit_account = data.get('credit_account')
                service_id = data.get('service_id')

                try:
                    service_type = ServiceType.objects.get(pk=service_id)
                    service_type.debit_account_id = debit_account
                    service_type.credit_account_id = credit_account
                    service_type.save()
                    success_response['status_code'] = 200
                    success_response['message'] = "Service type updated successfully"
                except ServiceType.DoesNotExist:
                    success_response['status_code'] = 404
                    success_response['message'] = "Service type not found"
                except Exception as e:
                    success_response['status_code'] = 500
                    success_response['message'] = str(e)
                

            else:
                success_response['status_code'] = 505
                success_response['message'] = "Invalid Module"
        else:
            success_response['status_code'] = 400
            success_response['message'] = f"Method Not Found"

        response = success_response

    except Exception as e:
        error_type, error_instance, traceback = sys.exc_info()
        tb_path = traceback.tb_frame.f_code.co_filename
        line_number = traceback.tb_lineno
        response["status_code"] = 500
        response[
            "message"] = f"An error occurred on line {line_number} in file {tb_path}. Details: {e}"

    return JsonResponse(response, safe=False)










