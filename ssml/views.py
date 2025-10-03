
from datetime import datetime
from decimal import Decimal
from django.http import HttpResponse, HttpResponseBadRequest, JsonResponse
from django.shortcuts import redirect, render
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from admin_panel.models import Emails, Locations, BusinessEntityTypes,MailSenders, MailQueues,MailAttachments
from admin_panel.anton import make_md5_hash


from ssml.form import ContractorErrorForm
from ssml.models import Contractor, ContractorError, Grn, GrnTransaction, InventoryMaterial, InvoiceHD, Issue, \
    IssueTransaction, Location, MaterialOrderItem, Meter, Plot, Reedem, RequiredReturn, Service, ServiceMaterialRates, ServiceMaterials, ServiceOrder, \
    ServiceOrderItem, ServiceType, ServiceTypeServices, Supplier, TransferHd,Documents


# Create your views here.
@login_required
def index(request):
    page = {
        'title': 'SSML',
        'page': 'ssml',
        'page_title': 'SSML',
        'page_description': 'SSML',
        'page_icon': 'bi bi-app',
        'page_url': 'ssml',
        'page_url_name': 'SSML',
        'nav':True
    }

    context = {
        'page': page
    }
    context['installations'] = ServiceType.objects.all()
    context['total_installs'] = ServiceOrder.objects.all().count()
    context['today_installs'] = ServiceOrder.objects.filter(service_date=datetime.now().date()).count()
    context['materials'] = InventoryMaterial.objects.all()
    # context['low_stock'] = InventoryMaterial.objects.filter()
    return render(request, 'ssml/index.html', context)

@login_required
def materials(request):
    page = {
        'title': 'Materials',
        'page': 'materials',
        'page_title': 'Materials',
        'page_description': 'Materials',
        'nav':True,
        'last_material': InventoryMaterial.objects.last()
    }

    sr_opt = ""
    for rt in Service.objects.all().order_by('name'):
        sr_opt += f"<option value='{rt.id}'>{rt.name} - {rt.rate}</option>"
    context = {
        'page': page,
        'next_barcode': f"MT0{InventoryMaterial.objects.filter(is_issue=False).last().id if InventoryMaterial.objects.filter(is_issue=False).exists() else 0 + 1}",
        'service_rates':sr_opt,
        'services':ServiceType.objects.all().order_by('name')
        
    }
    return render(request, 'ssml/materials.html', context)

@login_required
def grn(request):
    if Grn.objects.count() == 0:
        return redirect('grn_add')
    else:
        last_grn = Grn.objects.last()
        last_grn_no = last_grn.grn_no
        last_grn_date = last_grn.grn_date
        last_grn_remarks = last_grn.remarks
        last_grn_created_by = last_grn.created_by
        last_grn_total_amount = last_grn.total_amount
        last_grn_total_qty = last_grn.total_qty
    page = {
        'title': 'GRN',
        'page': 'grn',
        'page_title': 'GRN',
        'page_description': 'GRN',
        'nav':True
    }
    context = {
        'page': page,
        'last_grn': last_grn,
        'last_grn_no': last_grn_no,
        'last_grn_date': last_grn_date,
        'last_grn_remarks': last_grn_remarks,
        'last_grn_created_by': last_grn_created_by,
        'last_grn_total_amount': last_grn_total_amount,
        'last_grn_total_qty': last_grn_total_qty,
        'last_grn_transactions': GrnTransaction.objects.filter(grn_id=last_grn.pk),
        'last_grn_next_row': last_grn.next_row(),
        'last_grn_prev_row': last_grn.prev_row(),
        'last_grn_status': last_grn.status()
    }
    return render(request, 'ssml/grn.html', context)

@login_required
def issue(request):
    if Issue.objects.count() == 0:
        return redirect('issue_add')
    else:
        last_issue = Issue.objects.last()
        last_issue_no = last_issue.issue_no
        last_issue_date = last_issue.issue_date
        last_issue_remarks = last_issue.remarks
    page = {
        'title': 'Issue',
        'page': 'issue',
        'page_title': 'Issue',
        'page_description': 'Issue',
        'nav':True
    }
    context = {
        'page': page,
        'last_issue': last_issue,
        'last_issue_no': last_issue_no,
        'last_issue_date': last_issue_date,
        'last_issue_remarks': last_issue_remarks,
        'last_issue_transactions': IssueTransaction.objects.filter(issue_id=last_issue.pk),
        'last_issue_next_row': last_issue.next_row(),
        'last_issue_prev_row': last_issue.prev_row(),
        'last_issue_status': last_issue.status(),
        'searchButton':'search_issue_by_issue_no'
    }
    return render(request, 'ssml/issue.html', context)

@login_required
def issue_add(request):
    page = {
        'title': 'Add Issue',
        'page': 'issue_add',
        'page_title': 'Add Issue',
        'page_description': 'Add Issue',
        'nav':True
    }
    context = {
        'page': page,
        'issue_types': Issue.issue_type_choices
    }
    return render(request, 'ssml/issue_add.html', context)

@login_required
def grn_add(request):
    page = {
        'title': 'Add GRN',
        'page': 'grn_add',
        'page_title': 'Add GRN',
        'page_description': 'Add GRN',
        'nav':True,

    }
    context = {
        'page': page,
        'locations': Location.objects.all().order_by('loc_name')
        
    }   
    return render(request, 'ssml/grn_add.html', context)

@login_required
def grn_edit(request, id):
    page = {
        'title': 'Edit GRN',
        'page': 'grn_edit',
        'page_title': 'Edit GRN',
        'page_description': 'Edit GRN',
        'nav':True
    }
    context = {
        'page': page,
        'grn': Grn.objects.get(id=id),
        'grn_transactions': GrnTransaction.objects.filter(grn_id=id),
        'nav':True,
        'suppliers': Supplier.objects.all(),
        'grn_date_to_html': Grn.objects.get(id=id).grn_date.strftime('%Y-%m-%d')
    }
    return render(request, 'ssml/grn_edit.html', context)

@login_required
def upload_image(request):
    if request.method == 'POST':
        id = request.POST.get('id')
        image = request.FILES.get('image')
        material = InventoryMaterial.objects.get(id=id)
        material.image = image
        material.save()
        return redirect('materials')
    return redirect('materials')

@login_required
def issue_edit(request, id):
    page = {
        'title': 'Edit Issue',
        'page': 'issue_edit',
        'page_title': 'Edit Issue',
        'page_description': 'Edit Issue',
        'nav':True
    }
    context = {
        'page': page,
        'issue': Issue.objects.get(id=id),
        'issue_transactions': IssueTransaction.objects.filter(issue_id=id),
        'contractors': Contractor.objects.all(),
        'issue_date_to_html': Issue.objects.get(id=id).issue_date.strftime('%Y-%m-%d'),
        'issue_types': Issue.issue_type_choices
    }
    return render(request, 'ssml/issue_edit.html', context)

@login_required
def plots(request):
    page = {
        'title': 'Plots',
        'page': 'plots',
        'page_title': 'Plots',
        'page_description': 'Plots',
        'nav':True
    }
    context = {
        'page': page
    }
    return render(request, 'ssml/plots.html', context)


@login_required
def service_orders(request):
    # ServiceOrder.objects.all()
    if(ServiceOrder.objects.all().count() < 1):
        return redirect('service_order_new')
    
    opts = ""
    for st in ServiceType.objects.all():
        opts += f"<option value='{st.id}'>{st.name}</option>"
    page = {
        'title': 'Service Order',
        'page': 'service_order',
        'page_title': 'Service Order',
        'page_description': 'Service Order',
        'nav':True,
        
    }
    
    context = {
        
        'page': page,
        'contractors': Contractor.objects.all(),
        'plots': Plot.objects.all(),
        'service_types': {
            'water': 'Water',
            'electricity': 'Electricity',
            'sewerage': 'Sewerage',
            'gas': 'Gas',
            'telecom': 'Telecom'
        },
        'last_service_order': ServiceOrder.objects.last(),
        'searchButton':'search_service_by_meter',
        'service_ops':opts
    }
    return render(request, 'ssml/service_order.html', context)


@login_required
def services(request):
    page = {
        'title': 'Services',
        'page': 'services',
        'page_title': 'Services',
        'page_description': 'Services',
        'nav':True
    }
    context = {
        'page': page
    }
    return render(request, 'ssml/services.html', context)

@login_required
def contractors(request):
    page = {
        'title': 'Contractors',
        'page': 'contractors',
        'page_title': 'Contractors',
        'page_description': 'Contractors',
        'nav':True
    }
    context = {
        'page': page,
        'contractors': Contractor.objects.all(),
        'last_contractor': Contractor.objects.last()
    }
    return render(request, 'ssml/contractors.html', context)


@login_required
def service_order_new(request):
    
    page = {
        'title': 'New Service Order',
        'page': 'service_order_new',
        'page_title': 'New Service Order',
        'page_description': 'New Service Order',
        'nav':True,
        
    }
    context = {
        'page': page,
        'contractors': Contractor.objects.all(),
        'service_types': ServiceType.objects.all(),
        'issue_materials':InventoryMaterial.objects.filter(is_issue=True).order_by('name')
    }
    return render(request, 'ssml/service_order_new.html', context)


@login_required
def redeem(request):
    if Reedem.objects.all().count() < 1:
        return redirect('new_redeem')
    
    last_entry = Reedem.objects.all().last()
    page = {
        'title': 'Redeem',
        'page': 'service_order_new',
        'page_title': 'New Service Order',
        'page_description': 'New Service Order',
        'nav':True,
        
    }
     
    context = {
         'page':page,
         'last_entry':last_entry
     }
    
    return render(request, 'ssml/redeem.html', context)

@login_required
def new_redemption(request):
    page = {
        'title': 'New Redeem',
        'page': 'service_order_new',
        'page_title': 'New Service Order',
        'page_description': 'New Service Order',
        'nav':True,
        
    }
     
    context = {
         'page':page,
         'contractors':Contractor.objects.all().order_by('company')
     }
    
    return render(request, 'ssml/new_redeem.html', context)

@login_required
def meters(request):
    page = {
        'title':"Meters",
        'nav':True
    }

    # delete all return materials in issued
    #MaterialOrderItem.objects.filter(material__is_return=True,material_type='is').delete()

    context = {
         'page':page,
         'meters':Meter.objects.all()
     }
    
    return render(request, 'ssml/meters.html', context)


@login_required
def contractor_errors(request):
    page = {
        'title': 'Contractor Errors',
        'page': 'contractor_errors',
        'page_title': 'Contractor Errors',
        'page_description': 'Contractor Errors',
        'nav':True
    }
    context = {
        'page': page,
        'contractor_errors': ContractorError.objects.all(),
        'contractors': Contractor.objects.all().order_by('company'),
        'locations':Location.objects.all().order_by('loc_id')
    }
    return render(request, 'ssml/contractor_errors.html', context)

@login_required
def save_contractor_error(request):
    if request.method == 'POST':
        meter_no = request.POST.get('meter')
        location = Location.objects.get(pk=request.POST.get('mt_location'))

        print(request.POST)

    # add or get meter
        if not Meter.objects.filter(meter_no=meter_no).exists():
            Meter.objects.create(
                meter_no=meter_no,
                contractor=Contractor.objects.get(pk=request.POST.get('contractor')),
                created_by=request.user,
                location=location
            )


        form = ContractorErrorForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('contractor_errors')
        else:
            return HttpResponse(form.errors)
    return redirect('contractor_errors')


@login_required
def accounts(request):
    page = {
        'title': 'Accounts',
        'page': 'accounts',
        'page_title': 'Accounts',
        'page_description': 'Accounts',
        'nav':True
    }
    context = {
        'page': page
    }
    return render(request, 'ssml/accounts.html', context)



@login_required
def servive_type(request,id):
    service = ServiceType.objects.get(id=id)
    page = {
        'title': 'Service Type',
        'page': 'service_type',
        'page_title': 'Service Type',
        'page_description': 'Service Type',
        'nav':True
    }
    context = {
        'page': page,
        'service':service,
        'service_rates':Service.objects.all().order_by('name'),
        'my_rates':ServiceTypeServices.objects.filter(service_type=service),
        'id':id,
        'required_returns':RequiredReturn.objects.filter(service_type=service)
    }

    return render(request,'ssml/service-type.html',context=context)

@login_required
def add_service_to_type(request):
    if request.method == 'POST':
        selected_services = request.POST.getlist('service[]')
        s_type = request.POST['s_type']
        for service in selected_services:
            service_rate = Service.objects.get(id=service)
            try:
                ServiceTypeServices(
                    service_type_id=s_type,
                    service = service_rate
                ).save()
            except Exception as e:
                pass
            print(service_rate.name,service_rate.rate)
        # You can now use this list however you want, e.g., save to DB

    

    return redirect(request.META.get('HTTP_REFERER', '/'))

@csrf_exempt
@login_required
def upload_service_order(request):
    if request.method == 'POST':
        if 'excel_file' not in request.FILES:
            return HttpResponseBadRequest("No file uploaded.")
        
        import openpyxl
        excel_file = request.FILES['excel_file']
        wb = openpyxl.load_workbook(excel_file)
        sheet = wb.active  # or wb['SheetName'] if you know it
        service_id = request.POST['service']
        service = ServiceType.objects.get(pk=service_id)
        service_name = service.name

        headers = [cell.value for cell in sheet[1]] 
        sample_data = []

        service_rates = ServiceTypeServices.objects.filter(service_type=service)

        cts = 0

        for row in sheet.iter_rows(min_row=2, values_only=True):
            cts += 1
            row_dict = dict(zip(headers, row))
            customer = row_dict['CUSTOMER NAME']
            customer_phone = row_dict['CUSTOMER PHONE']
            contractor_code = row_dict['CONTRACTOR CODE']
            print(contractor_code)
            contractor = Contractor.objects.get(code=contractor_code)
            dt = row_dict['DATE APPROVED'] if row_dict.get('DATE APPROVED')  else row_dict['DATE RESOLVED']
            from datetime import datetime
            resolved_on = datetime.strptime(dt, "%d/%m/%Y %H:%M")
            print(resolved_on)
            meter_no = row_dict['MFG Serial Number']
            ip_address = row_dict['IP Address']

            order_status = row_dict['PREPAID SYNC STATUS']
            address = row_dict['DIGITAL ADDRESS']
            old_meter = row_dict['OLD METER NUMBER'] if row_dict.get('OLD METER NUMBER') else row_dict['SERVICE POINT NUMBER']


            

            # get materials
            service_charge = 0

            #if order_status == 'Completed': 
            if True:
                hd = [service_name,customer,customer_phone,contractor.company,resolved_on,meter_no,ip_address]

                # check if meter job created
                ServiceOrder.objects.filter(new_meter=meter_no).delete()
                                  
                    # create
                try:
                    ServiceOrder(
                            service_type= service,
                            service_date = resolved_on,
                            contractor=contractor,
                            plot = "",
                            geo_code = "",
                            customer = customer,
                            customer_no = customer_phone,
                            old_meter_no = old_meter,
                            new_meter=meter_no,
                            status=order_status,
                            created_by=request.user,
                            location=Location.objects.get(loc_id='002')
                        ).save()
                    print("Service Created")
                except Exception as e:
                    # update type
                    ServiceOrder.objects.filter(new_meter=meter_no).update(service_type=service,service_date=resolved_on)
                    print(e)
                print(meter_no)
                met_service = ServiceOrder.objects.get(new_meter=meter_no)
            

                # print(hd)
                # print('RATES')
                rate_message = ""
                total_rate = 0
                for rate in service_rates:
                    rx = rate.service
                    name = rate.service.name
                    rt = rate.service.rate
                    total_rate += rt
                    # rate_message += f"{name},1,{rt},{rt*1}\n"

                    if ServiceOrderItem.objects.filter(service_order=met_service,service=rate.service).exists:
                        ServiceOrderItem.objects.filter(service_order=met_service,service=rate.service).delete()
                    
                        # insert
                    ServiceOrderItem(
                            service_order=met_service,service=rx,
                            quantity=1,rate=rt,amount=Decimal(rt)*Decimal(1)

                        ).save()
                    # print(name,rt)    
                # print("Materials")       
                materials_message = ''
                for material in InventoryMaterial.objects.all().order_by('name'):
                    qty = row_dict.get(material.name,0)
                    # add material or update
                    MaterialOrderItem.objects.filter(service_order=met_service,material=material).delete()
                    

                    if qty > 0:
                        # print("FISH",material.name,qty)
                        materials_message += f"{material.name} | {qty}\n"

                        
                        MaterialOrderItem(
                                service_order=met_service,material=material,quantity=qty,rate=0,amount=0
                            ).save()
                        
                        # get service rates
                        other_rates = ServiceMaterialRates.objects.filter(material__barcode=material.barcode)
                        # print(other_rates.count(),"COUNTED")
                        if other_rates.count() > 0:
                            extra_rate = ServiceMaterialRates.objects.get(material__barcode=material.barcode)
                            if qty > extra_rate.break_point :
                                # print("GREATER")
                                rate = extra_rate.service.rate
                                name = extra_rate.service.name
                                total_rate += extra_rate.service.rate
                                # rate_message += f"{name},1,{rt},{rt*1}\n"
                            else:
                                pass
                                # print("NOT GREATER")
                        
                        # print(meter_no, material.name,qty)

                # default issues    
                MaterialOrderItem.objects.filter(service_order=met_service)    
                for mat in ServiceMaterials.objects.filter(service=service):
                    print(mat.material.name,mat.qty)
                    try:
                        MaterialOrderItem(
                                    service_order=met_service,material=mat.material,quantity=mat.qty,rate=0,amount=0
                                ).save()
                    except Exception as e:
                        print(e)
                        
                    
                # update total amount
                met_service.total_amount = ServiceOrder.objects.get(new_meter=meter_no).total_amount()



            # print(f'###### {contractor.company} ##### {meter_no}')
            # print("MATERIALS")
            # print(materials_message)
            # print("SERVICE RATES")
            # print(rate_message)

            # print(f'###### total ##### {total_rate}')
            # print('\n')
            sample_data.append(row_dict)

        print(cts, service_name, "Added")

        return HttpResponse(sample_data[:5])

    else:
        return HttpResponse("Invalid Method")

@csrf_exempt
@login_required
def sync_retired_meter(request):
    if request.method == 'POST':
        if 'excel_file' not in request.FILES:
            return HttpResponseBadRequest("No file uploaded.")
        
        import openpyxl
        excel_file = request.FILES['excel_file']
        wb = openpyxl.load_workbook(excel_file)
        no_service_file_name = "sync_ret_meter"
        book = openpyxl.Workbook()
        sheet = book.active
        
        foc_list = {}
        # Loop through each sheet in the workbook
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            # print(f"Sheet: {sheet_name}")
            # Loop through each row in the sheet
            for row in sheet.iter_rows(values_only=True):
                box_no = row[2]
                box_litral = box_no.split('-')[0] if box_no else 'not a valid row'
                
                if box_litral== 'SNE':
                    old_meter_no = str(row[4]).split('.')[0]
                    
                    foc_name = row[21] or None
                    is_service = False
                    if old_meter_no:
                        # check if old meter exisit in done job
                        if ServiceOrder.objects.filter(old_meter_no=old_meter_no).exists():
                            is_service = True
                        else:
                            # If old_meter_no starts with '0', remove the leading '0'
                            if old_meter_no and str(old_meter_no).startswith('0'):
                                old_meter_no = str(old_meter_no)[1:]
                                if ServiceOrder.objects.filter(old_meter_no=old_meter_no).count() == 0:
                                    old_meter_no = f"0{old_meter_no}"
                                    if ServiceOrder.objects.filter(old_meter_no=old_meter_no).count() == 0:
                                        # try adding 0 if no 0
                                        
                                        print(old_meter_no,"HAS NO SERVICE")
                                        pass
                                    else:
                                        
                                        is_service = True
                                else:
                                    # print(old_meter_no)
                                    is_service = True
                                    # update return
                    
                    


                    if is_service:
                        phase = row[13]
                        service = ServiceOrder.objects.get(old_meter_no=old_meter_no)
                        if '3' in str(service.service_type.name):
                            return_barcode = "MT067"
                        else:
                            return_barcode = 'MT026'
                        
                        
                        product = InventoryMaterial.objects.get(barcode=return_barcode)
                        
                        foc_name = service.contractor.company

                        if foc_name not in foc_list:
                            foc_list[foc_name] = {
                                "MT067":0,
                                "MT026":0
                            }
                            
                        
                        foc_list[foc_name][return_barcode] = foc_list[foc_name][return_barcode] + 1
                            
                        
                        # if foc_name == 'KOSMEN ELECTRICAL WORKS':
                        if True:
                            print(foc_name,box_no,old_meter_no,service.new_meter,phase,service.service_type.name,product.name)
                            sheet.append([foc_name,box_no,old_meter_no,service.new_meter,phase,service.service_type.name,product.name])
                        # add to return
                        try:
                            MaterialOrderItem.objects.create(
                                material=product,
                                quantity=1,
                                service_order=service,
                                material_type='rt',
                                rate=0,
                                amount=0
                            )
                            # print(product.name,'ADDED')
                        except Exception as e:
                            pass
                            # print(e)
                        # print(old_meter_no,service.new_meter,phase,service.service_type.name,product.name)
                        

        print(foc_list)

        book.save(f'static/general/tmp/{no_service_file_name}.xlsx')
                

    return JsonResponse({"status_code":200,"message":no_service_file_name})

@csrf_exempt
@login_required
def add_material_rate(request):
    if request.method == 'POST':
        form = request.POST
        mat_id = form.get('mat_id')
        break_point = form.get('break_point')
        service_rate_id = form.get('service_rate')
        material = InventoryMaterial.objects.get(id=mat_id)
        rate = Service.objects.get(id=service_rate_id)
        try:
            ServiceMaterialRates(
                service=rate,material=material,break_point=break_point
            ).save()
        except Exception as e:
            pass

        ## add to model

        print(form)
    return redirect(request.META.get('HTTP_REFERER', '/'))


def delete_service_rate(request,id):
    ServiceMaterialRates.objects.filter(id=id).delete()

    return redirect(request.META.get('HTTP_REFERER', '/'))
def delete_type_service_rate(request,id):
    ServiceTypeServices.objects.filter(id=id).delete()

    return redirect(request.META.get('HTTP_REFERER', '/'))

@login_required
def invoice(request):
    if InvoiceHD.objects.all().count() == 0:
        return redirect('new-cont-inv')
    
    last_entry = InvoiceHD.objects.all().last().id
    page = {
        'title': 'Contractor Invoice',
        'page': 'contractor_invoice',
        'page_title': 'Contractor Invoice',
        'page_description': 'Service Type',
        'nav':True
    }
    context = {
        'page': page,
        'last_entry':last_entry
    }

    return render(request,'ssml/contractor-invoice.html',context=context)
@login_required
def invoice_new(request):
    page = {
        'title': 'NEW Contractor Invoice',
        'page': 'contractor_invoice',
        'page_title': 'Contractor Invoice',
        'page_description': 'Service Type',
        'nav':True
    }
    context = {
        'page': page,
        'contractors':Contractor.objects.all().order_by('company')
    }

    return render(request,'ssml/contractor-invoice-new.html',context=context)

@login_required
def location_master(request):
    page = {
        'title': 'Location Master',
        'page': 'ssml_location_master',
        'page_title': 'Location Master',
        'page_description': 'Location Master',
        'nav':True
    }
    context = {
        'page': page
    }

    return render(request,'ssml/location-master.html',context=context)

@login_required()
def transfer(request):
    # TransferHd.objects.all().delete()
    if TransferHd.objects.all().count() == 0:
        return redirect('ssml_add_transfer')
    page = {
        'title': 'Transfer',
        'page': 'transfer',
        'page_title': 'Transfer',
        'page_description': 'Transfer',
        'nav':True,
        
    }
    context = {
        'page': page,
        'last_entry':TransferHd.objects.all().last().id,
        'nav':True
    }

    return render(request,'ssml/inventory/transfer.html',context=context)

@login_required()
def add_transfer(request):
    page = {
        'title': 'Add Transfer',
        'nav':True
    }

    context = {
        'page': page,
        'locations': Location.objects.all().order_by('loc_name'),
    }

    return render(request,'ssml/inventory/transfer_add.html',context=context)


    # INSERT_YOUR_CODE
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from .models import FinancialEvidence ,Expense # Assuming these models exist

@login_required()
@csrf_exempt
def financial_evidence(request):
    if request.method == 'POST':
        transaction_id = request.POST.get('transaction_id')
        title = request.POST.get('title')
        file = request.FILES.get('file')

        if not (transaction_id and title and file):
            return HttpResponse(title)
            return JsonResponse({'status': False, 'message': 'Missing required fields.'}, status=400)

        try:
            transaction = Expense.objects.get(id=transaction_id)
        except Expense.DoesNotExist:
            return JsonResponse({'status': False, 'message': 'Transaction not found.'}, status=404)

        evidence = FinancialEvidence.objects.create(
            transaction=transaction,
            title=title,
            file=file,
            uploaded_by=request.user
        )
        from django.shortcuts import redirect
        return redirect(request.META.get('HTTP_REFERER', '/'))

    else:
        return JsonResponse({'status': True, 'message': 'Invalid Method',"":""})
    


        # INSERT_YOUR_CODE
from django.contrib.auth.decorators import login_required
from django.shortcuts import render

@login_required()
def request_expense(request):
    page = {
        'title': 'Request Expense',
        'nav': False
    }
    context = {
        'page': page,
    }
    # INSERT_YOUR_CODE
    # Pass category_choices in context
    from .models import Expense  # Assuming Expense model has category choices

    if hasattr(Expense, 'category_choices'):
        category_choices = Expense.category_choices
    else:
        category_choices = []

    context['category_choices'] = category_choices
    return render(request, 'ssml/accounts-req-expense.html', context=context)




from django.http import JsonResponse, HttpResponseRedirect
from django.urls import reverse
from django.utils import timezone


@login_required()
def save_expense_request(request):
    if request.method == "POST":
        from decimal import Decimal, InvalidOperation

        # Get data from POST
        category = request.POST.get('category')
        amount = request.POST.get('amount')
        date = request.POST.get('date')
        reference = request.POST.get('reference', '')
        description = request.POST.get('description', '')
        evidence_file = request.FILES.get('evidence')

        # Validate required fields
        if not (category and amount and date):
            return JsonResponse({'status': False, 'message': 'Missing required fields.'}, status=400)

        # Validate amount
        try:
            amount = Decimal(amount) * -1
        except (InvalidOperation, TypeError):
            return JsonResponse({'status': False, 'message': 'Invalid amount.'}, status=400)

        # Save Expense (direction is always 'out' for requests, transaction_type is 'cash' by default)
        expense = Expense.objects.create(
            direction='out',
            date=date,
            category=category,
            transaction_type='cash',
            amount=amount,
            reference=reference,
            description=description,
            created_by=request.user
        )

        # Save evidence if provided
        file_name = ""
        if evidence_file:
            from .models import FinancialEvidence
            FinancialEvidence.objects.create(
                transaction=expense,
                title="Evidence",
                file=evidence_file,
                uploaded_by=request.user
            )
        
            file_name = FinancialEvidence.objects.last().file.url
        # INSERT_YOUR_CODE
        # Prepare email subject and body for notification (not sending, just variables)
        email_subject = f"Expense Request Submitted by {request.user.get_full_name() or request.user.username}"
        email_body = f"""
        <div style="padding:0;margin:0;">
            <div style="max-width:420px;margin:0 auto;background:#fff;border-radius:6px;box-shadow:0 2px 8px #0001;padding:40px 30px 30px 30px;position:relative;top:40px;">
                <h1 style="text-align:center;font-size:2.2em;letter-spacing:2px;margin-bottom:18px;font-family:sans-serif;">Expense Request</h1>
                <p style="font-size:1.1em;color:#222;margin-bottom:24px;font-family:sans-serif;">
                    An expense request has been submitted. Please review the details below.
                </p>
                <div style="background:#f4f8fb;border-radius:5px;padding:18px 20px 10px 20px;margin-bottom:24px;alight-text:left">
                    <p style="margin:0 0 10px 0;font-family:sans-serif;"><b>Submitted by:</b> {request.user.get_full_name() or request.user.username}</p>
                    <p style="margin:0 0 10px 0;font-family:sans-serif;"><b>Category:</b> {category}</p>
                    <p style="margin:0 0 10px 0;font-family:sans-serif;"><b>Amount:</b> {amount}</p>
                    <p style="margin:0 0 10px 0;font-family:sans-serif;"><b>Date:</b> {date}</p>
                    <p style="margin:0 0 10px 0;font-family:sans-serif;"><b>Reference:</b> {reference}</p>
                    <p style="margin:0 0 10px 0;font-family:sans-serif;"><b>Description:</b> {description}</p>
                    <p style="margin:0 0 0 0;font-family:sans-serif;"><b>Submitted at:</b> {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
                </div>
                
            </div>
            
        </div>
        """

        # que mail
        
        if MailSenders.objects.filter(address='solomon@snedaghana.com').exists():
            sender = MailSenders.objects.get(address='solomon@snedaghana.com')
        else:
            sender = MailSenders.objects.get(is_default=True)

        subject = email_subject
        mail = MailQueues(sender=sender, subject=subject, body=email_body,
                                      recipient='solomon@snedaghana.com',
                                      cc='uyinsolomon2@gmail.com,ajay@snedaghana.com,bharat@snedaghana.com',mail_key=make_md5_hash(f"{subject}{email_body}"))
        mail.save()
        
        if evidence_file:
            # INSERT_YOUR_CODE
            
            MailAttachments(mail=mail, attachment=file_name).save()

        
        # Redirect to accounts page or show success
        return HttpResponseRedirect(reverse('accounts'))

    # If not POST, redirect to request page
    return HttpResponseRedirect(reverse('request_expense'))

@csrf_exempt
@login_required
def approve_document(request):
    if request.method == 'POST':
        try:
            form = request.POST
            entry_no = form.get('entry_no')
            doc = form.get('doc')
            title = form.get('title')
            description = form.get('title')
            entry_no = form.get('entry_no')

            if doc == 'ssml_tr':
                document = TransferHd.objects.get(entry_no=entry_no)
                document.post()
                

            # upload file
            Documents.objects.create(
                title=title,
                file=request.FILES.get('file') if request.FILES else None,
                description=description,
                doc_type=doc,
                ref=entry_no
            )

            # INSERT_YOUR_CODE
            from django.http import JsonResponse
            return JsonResponse({'status_code': '200','message':"Document Approved"})
        
        except Exception as e:
            # INSERT_YOUR_CODE
            from django.http import JsonResponse
            return JsonResponse({'status_code': '500', 'message': str(e)})

@login_required()
def foc_portal(request):
    # Render the template into a response object
    response = render(request, 'ssml/foc-potal.html')

    # Check if 'stage' cookie exists
    if not request.COOKIES.get('stage'):
        # Set the cookie if it doesn't exist
        response.set_cookie('stage', 'auth', max_age=3600)  # expires in 1 hour

    return response