import json
import sys
from datetime import date
from decimal import Decimal

import pyodbc
import requests
from django.contrib.auth.models import User
from django.db.models import Q, Sum
from django.http import JsonResponse
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from fpdf import FPDF
from openpyxl.chart import PieChart, Reference
from openpyxl.pivot.table import Location
from openpyxl.styles.builtins import total
from scipy.ndimage import sobel
from sympy import Product

from admin_panel.anton import format_currency
from admin_panel.models import Emails, Locations, BusinessEntityTypes
from inventory.views import transfer
from ocean.settings import RET_DB_HOST, RET_DB_USER, RET_DB_PASS, RET_DB_NAME, BOLT_PROVIDER_ID, BOLT_MARGIN
from retail.db import ret_cursor, get_stock, updateStock, percentage_difference, stock_by_moved, stock_by_prod
from retail.models import BoltItems, BoltGroups, ProductSupplier, ProductGroup, ProductSubGroup, Products, Stock, \
    RecipeGroup, RecipeProduct, Recipe, StockHd, StockMonitor, RawStock, ButcheryLiveTransactions, ButchSales, \
    StockToSend, TranHd, TranTr, RetailSales, SampleHd, SampleTran, BoltSubGroups, BillHeader, BillTrans, ProductMoves, \
    Barcode, MoveStock
from retail.prodMast import ProdMaster
from retail.retail_tools import create_recipe_card


@csrf_exempt
def interface(request):
    response = {
        'status_code': 0,
        'message': ""
    }

    success_response = {
        'status_code': 200,
        'message': "Procedure Completed Successfully"
    }

    # get method
    method = request.method

    try:

        body = json.loads(request.body)
        module = body.get('module')
        data = body.get('data')

        if method == 'PUT':  # add bolt product
            if module == 'bolt_item':


                #print(data)

                item_code = data.get('item_code').replace("\n", "")

                group_pk = data.get('group')
                subgroup_pk = data.get('sub_group')

                image = data.get('image')
                entity_id = data.get('entity')
                entity = BusinessEntityTypes.objects.get(id = entity_id)
                #print(item_code)
                if Products.objects.filter(barcode=item_code,entity=entity).count() == 0:
                    item_code = f"0{item_code}"
                pd = Products.objects.get(barcode=item_code,entity=entity)
                price = pd.price
                item_code = pd.code


                stock = get_stock(item_code)
                #print(stock)
                nia_stock = stock.get('nia',0)
                spintex_stock = stock.get('spintex',0)
                osu_stock = stock.get('osu',0)

                # BoltItems.objects.filter(product=pd).delete()

                if BoltGroups.objects.filter(name=group_pk,entity=entity).count() == 0:
                    BoltGroups.objects.create(name=group_pk,entity=entity)

                group = BoltGroups.objects.get(name=group_pk,entity=entity)

                if BoltSubGroups.objects.filter(name=subgroup_pk,group=group,entity=entity).count() == 0:
                    BoltSubGroups.objects.create(name=subgroup_pk,group=group,entity=entity)

                subgroup = BoltSubGroups.objects.get(name=subgroup_pk,entity=entity,group=group)




                if BoltItems.objects.filter(product=pd,menu=entity).count() == 0:
                    BoltItems(product=pd,price=price,stock_nia=nia_stock,stock_spintex=spintex_stock,stock_osu=osu_stock,group=group,subgroup=subgroup,image=image,menu=entity).save()
                    success_response['message'] = "Product Added"
                else:
                    raise Exception(f"Product Exist with barcode {pd.barcode}")

            elif module == 'moves':
                move_type = data.get('type','')
                print(data)
                if move_type == 'CB':
                    from datetime import datetime, timedelta
                    # closing balance
                    target_date = data.get('target_date')
                    # add one day from target date to get next Opening Balance
                    date_obj = datetime.strptime(target_date, "%Y-%m-%d")
                    ob_date = date_obj + timedelta(days=1)

                    transactions = data.get('transactions')
                    for transaction in transactions:
                        barcode = transaction['barcode']
                        quantity = transaction['quantity']

                        product = Products.objects.get(barcode=barcode)
                        ProductMoves.objects.create(
                            move_type = move_type,
                            quantity=quantity,
                            product=product,
                            date=date_obj
                        )
                        ProductMoves.objects.create(
                            move_type='OB',
                            quantity=quantity,
                            product=product,
                            date=ob_date
                        )

                elif move_type == 'SR':
                    transactions = data.get('transactions')
                    entry_date = data.get('entry_date')
                    loc_id = data.get('loc_id')
                    location = Locations.objects.get(code=loc_id)
                    remarks = data.get('remarks')

                    for transaction in transactions:
                        # print(transaction)
                        barcode = transaction['barcode']
                        counted_stock = Decimal(transaction['quantity'])  # The physically counted stock

                        product = Products.objects.get(barcode=barcode)

                        # Get the current stock from ProductMoves
                        current_stock = \
                        ProductMoves.objects.filter(product=product, location=location,date__lt=entry_date).aggregate(Sum('quantity'))[
                            'quantity__sum'] or 0

                        # Calculate the adjustment needed
                        adjustment = counted_stock - current_stock  # Ensure total matches counted stock

                        # Only create adjustment if there's a difference
                        if adjustment != 0:
                            ProductMoves.objects.create(
                                product=product,
                                location=location,
                                move_type=move_type,
                                quantity=adjustment,
                                remark=remarks,
                                ref="SR",
                                date=entry_date,
                                line_no=1
                            )

                elif move_type == 'IN':
                    conn = ret_cursor()
                    cursor = conn.cursor()
                    print("INVOICES")

                    # get invoice tran
                    query = "select hd.entry_no,hd.entry_date,tr.total_units,tr.line_no,hd.loc_id,'',tr.item_code,hd.remark from inv_tran tr join inv_hd hd on hd.entry_no = tr.entry_no where ocean is NULL and hd.valid = 1 and hd.posted = 1"
                    cursor.execute(query)
                    print(query)
                    rows = cursor.fetchall()
                    over_lines = len(rows)
                    compare_line = 1
                    for row in rows:
                        entry_no,entry_date,quantity,line_no,loc_id,loc_to,item_code,remarks = row
                        log = f"{move_type} : {compare_line} / {over_lines} {item_code}"
                        # get product
                        product =  Products.objects.filter(code=item_code)
                        if product.exists():
                            location = Locations.objects.get(code=loc_id)
                            pd = product.last()
                            quantity = quantity * -1
                            log += " Exist "
                            log += str(quantity)
                            try:
                                ProductMoves.objects.get_or_create(
                                    line_no=line_no,
                                    product=pd,
                                    quantity=quantity,
                                    date=entry_date,
                                    move_type=move_type,
                                    remark=remarks,
                                    ref=entry_no,
                                    location=location
                                )

                                update_q = f"UPDATE inv_tran set ocean = 1 where item_code = '{item_code}' and entry_no = '{entry_no}' and line_no = {line_no}"
                                cursor.execute(update_q)
                                cursor.commit()
                            except Exception as e:
                                pass
                        else:
                            log += " Does Not Exists"
                        compare_line += 1
                        print(log)

                    conn.close()

                elif move_type == 'PIXXX':
                    conn = ret_cursor()
                    cursor = conn.cursor()
                    print("PI")

                    # get invoice tran
                    query = "select hd.entry_no,hd.entry_date,tr.tran_qty, '' as 'line_no',hd.location_id,'' as 'loc_to',tr.prod_id,CONCAT('POS SALES',hd.ref_no) from pos_tran tr join pos_tran_hd hd on hd.entry_no = tr.entry_no where tr.ocean is null and hd.inv_upd = 1 and hd.inv_valid = 1 and hd.location_id = '001'"
                    cursor.execute(query)
                    print(query)
                    rows = cursor.fetchall()
                    over_lines = len(rows)
                    compare_line = 1
                    line_no = 1
                    for row in rows:
                        entry_no, entry_date, quantity, x, loc_id, loc_to, item_code, remarks = row
                        log = f"{move_type} : {compare_line} / {over_lines} {item_code}"
                        # get product
                        product = Products.objects.filter(code=item_code)
                        if product.exists():
                            location = Locations.objects.get(code=loc_id)
                            pd = product.last()
                            quantity = quantity * -1
                            log += " Exist "
                            log += str(quantity)
                            try:
                                ProductMoves.objects.get_or_create(
                                    line_no=line_no,
                                    product=pd,
                                    quantity=quantity,
                                    date=entry_date,
                                    move_type=move_type,
                                    remark=remarks,
                                    ref=entry_no,
                                    location=location
                                )

                                update_q = f"UPDATE pos_tran set ocean = 1 where prod_id = '{item_code}' and entry_no = '{entry_no}'"
                                cursor.execute(update_q)
                                cursor.commit()
                                line_no += 1
                            except Exception as e:
                                pass
                        else:
                            log += " Does Not Exists"
                        compare_line += 1
                        print(log)

                    conn.close()

                elif move_type == 'GR':
                    conn = ret_cursor()
                    cursor = conn.cursor()
                    print("GRN")

                    # get invoice tran
                    query = ("select hd.entry_no,hd.grn_date,tr.total_units,tr.line_no,hd.loc_id,'',tr.item_code,"
                             "hd.remark from grn_tran tr join grn_hd hd on hd.entry_no = tr.entry_no where ocean is NULL and hd.valid = 1 and hd.posted = 1")
                    cursor.execute(query)
                    print(query)
                    rows = cursor.fetchall()
                    over_lines = len(rows)
                    compare_line = 1
                    for row in rows:
                        entry_no, entry_date, quantity, line_no, loc_id, loc_to, item_code, remarks = row
                        log = f"{move_type} : {compare_line} / {over_lines} {item_code}"
                        # get product
                        product = Products.objects.filter(code=item_code)
                        if product.exists():
                            location = Locations.objects.get(code=loc_id)
                            pd = product.last()

                            log += " Exist "
                            log += str(quantity)
                            try:
                                ProductMoves.objects.get_or_create(
                                    line_no=line_no,
                                    product=pd,
                                    quantity=quantity,
                                    date=entry_date,
                                    move_type=move_type,
                                    remark=remarks,
                                    ref=entry_no,
                                    location=location
                                )

                                update_q = f"UPDATE grn_tran set ocean = 1 where item_code = '{item_code}' and entry_no = '{entry_no}' and line_no = {line_no}"
                                cursor.execute(update_q)
                                cursor.commit()
                            except Exception as e:
                                log += f" {e}"
                        else:
                            log += " Does Not Exists"
                        compare_line += 1
                        print(log)

                    conn.close()

                elif move_type == 'AD':
                    conn = ret_cursor()
                    cursor = conn.cursor()
                    print("GRN")

                    # get invoice tran
                    query = ("select hd.entry_no,hd.entry_date,tr.total_units,tr.line_no,hd.loc_id,'',tr.item_code,"
                             "hd.reason from adj_tran tr join adj_hd hd on hd.entry_no = tr.entry_no where ocean is NULL and hd.valid = 1 and hd.posted = 1")
                    cursor.execute(query)
                    print(query)
                    rows = cursor.fetchall()
                    over_lines = len(rows)
                    compare_line = 1
                    for row in rows:
                        entry_no, entry_date, quantity, line_no, loc_id, loc_to, item_code, remarks = row
                        log = f"{move_type} : {compare_line} / {over_lines} {item_code}"
                        # get product
                        product = Products.objects.filter(code=item_code)
                        if product.exists():
                            location = Locations.objects.get(code=loc_id)
                            pd = product.last()

                            log += " Exist "
                            log += str(quantity)
                            try:
                                ProductMoves.objects.get_or_create(
                                    line_no=line_no,
                                    product=pd,
                                    quantity=quantity,
                                    date=entry_date,
                                    move_type=move_type,
                                    remark=remarks,
                                    ref=entry_no,
                                    location=location
                                )

                                update_q = f"UPDATE adj_tran set ocean = 1 where item_code = '{item_code}' and entry_no = '{entry_no}' and line_no = {line_no}"
                                cursor.execute(update_q)
                                cursor.commit()
                            except Exception as e:
                                pass
                        else:
                            log += " Does Not Exists"
                        compare_line += 1
                        print(log)

                    conn.close()

                elif move_type == 'TR':
                    conn = ret_cursor()
                    cursor = conn.cursor()
                    print("TRANSFER")

                    # get invoice tran
                    query = "select hd.entry_no,hd.entry_date,tr.total_units,tr.line_no,hd.loc_from,loc_to,tr.item_code,hd.remark,tr.item_ref,tr.item_des from tran_tr tr join tran_hd hd on hd.entry_no = tr.entry_no where ocean is NULL and hd.valid = 1 and hd.posted = 1;"
                    cursor.execute(query)
                    print(query)
                    rows = cursor.fetchall()
                    over_lines = len(rows)
                    compare_line = 1
                    for row in rows:
                        entry_no, entry_date, quantity, line_no, loc_id, loc_to, item_code, remarks,barcode,name = row
                        print([entry_no,item_code,barcode,name])
                        log = f"{move_type} : {compare_line} / {over_lines} {item_code}"
                        # get product
                        product = Products.objects.filter(code=item_code)
                        cds = [item_code]
                        if product.exists():
                            location = Locations.objects.get(code=loc_to)
                            pd = product.last()

                            log += " Exist "
                            log += str(quantity)

                            try:

                                ProductMoves.objects.get_or_create(
                                    line_no=line_no,
                                    product=pd,
                                    quantity=quantity,
                                    date=entry_date,
                                    move_type=move_type,
                                    remark=remarks,
                                    ref=entry_no,
                                    location=location
                                )

                                locfrom = Locations.objects.get(code=loc_id)

                                ProductMoves.objects.get_or_create(
                                    line_no=line_no,
                                    product=pd,
                                    quantity=quantity * -1,
                                    date=entry_date,
                                    move_type=move_type,
                                    remark=remarks,
                                    ref=entry_no,
                                    location=locfrom
                                )
                                log += f"{quantity * -1} to {locfrom.code}"

                                update_q = f"UPDATE tran_tr set ocean = 1 where item_code = '{item_code}' and entry_no = '{entry_no}' and line_no = {line_no}"
                                cursor.execute(update_q)
                                cursor.commit()
                            except Exception as e:
                                log += f" {e}"
                        else:
                            log += " Does Not Exists"
                        compare_line += 1
                        print(log)

                    conn.close()

                elif move_type == 'SRT':
                    conn = ret_cursor()
                    cursor = conn.cursor()
                    print("GRN")

                    # get invoice tran
                    query = "select hd.entry_no,hd.entry_date,tr.total_units,tr.line_no,hd.loc_id,'',tr.item_code,hd.remark from return_tran tr join return_hd hd on hd.entry_no = tr.entry_no where ocean is NULL and hd.valid = 1 and hd.posted = 1;"
                    cursor.execute(query)
                    print(query)
                    rows = cursor.fetchall()
                    over_lines = len(rows)
                    compare_line = 1
                    for row in rows:
                        entry_no, entry_date, quantity, line_no, loc_id, loc_to, item_code, remarks = row
                        log = f"{move_type} : {compare_line} / {over_lines} {item_code}"
                        # get product
                        product = Products.objects.filter(code=item_code)
                        if product.exists():
                            location = Locations.objects.get(code=loc_id)
                            pd = product.last()

                            log += " Exist "
                            log += str(quantity)
                            try:
                                ProductMoves.objects.get_or_create(
                                    line_no=line_no,
                                    product=pd,
                                    quantity=quantity,
                                    date=entry_date,
                                    move_type=move_type,
                                    remark=remarks,
                                    ref=entry_no,
                                    location=location
                                )

                                update_q = f"UPDATE return_tran set ocean = 1 where item_code = '{item_code}' and entry_no = '{entry_no}' and line_no = {line_no}"
                                cursor.execute(update_q)
                                cursor.commit()
                            except Exception as e:
                                pass
                        else:
                            log += " Does Not Exists"
                        compare_line += 1
                        print(log)

                    conn.close()

                elif move_type == 'PRT':
                    conn = ret_cursor()
                    cursor = conn.cursor()
                    print("PURCHASE RETURN")

                    # get invoice tran
                    query = "select hd.entry_no,hd.entry_date,tr.total_units * -1 as 'total_units',tr.line_no,hd.loc_id,'',tr.item_code,hd.remark from purch_ret_tran tr join purch_ret_hd hd on hd.entry_no = tr.entry_no where ocean is NULL and hd.valid = 1 and hd.posted = 1;;"
                    cursor.execute(query)
                    print(query)
                    rows = cursor.fetchall()
                    over_lines = len(rows)
                    compare_line = 1
                    for row in rows:
                        entry_no, entry_date, quantity, line_no, loc_id, loc_to, item_code, remarks = row
                        log = f"{move_type} : {compare_line} / {over_lines} {item_code}"
                        # get product
                        product = Products.objects.filter(code=item_code)
                        if product.exists():
                            location = Locations.objects.get(code=loc_id)
                            pd = product.last()

                            log += " Exist "
                            log += str(quantity)
                            try:
                                ProductMoves.objects.get_or_create(
                                    line_no=line_no,
                                    product=pd,
                                    quantity=quantity,
                                    date=entry_date,
                                    move_type=move_type,
                                    remark=remarks,
                                    ref=entry_no,
                                    location=location
                                )

                                update_q = f"UPDATE purch_ret_tran set ocean = 1 where item_code = '{item_code}' and entry_no = '{entry_no}' and line_no = {line_no}"
                                cursor.execute(update_q)
                                cursor.commit()
                            except Exception as e:
                                pass
                        else:
                            log += " Does Not Exists"
                        compare_line += 1
                        print(log)

                    conn.close()


                else:
                    raise Exception("No Document Type")
                    products = Products.objects.all()[:10]
                    lg_ct = 1
                    all_ct = products.count()
                    for product in products:#.order_by('name'):
                        print(lg_ct,"/",all_ct, product.name)
                        lg_ct = lg_ct + 1
                        # GRN
                        if move_type == 'GR':
                            g_query = f"select hd.entry_no,hd.grn_date,total_units,tr.line_no,hd.loc_id,'' from grn_tran tr join grn_hd hd on hd.entry_no = tr.entry_no  where tr.ocean is null and hd.posted = 1 and tr.item_ref = '{product.barcode}'"

                        elif move_type == 'TR':
                            g_query = f"select hd.entry_no,hd.entry_date,total_units,tr.line_no,hd.loc_from,hd.loc_to  from tran_tr tr join tran_hd hd on hd.entry_no = tr.entry_no  where tr.ocean is null and hd.posted = 1 and tr.item_ref = '{product.barcode}'"

                        elif move_type == 'AD':
                            g_query = f"select hd.entry_no,hd.entry_date,total_units,tr.line_no,hd.loc_id,'' from adj_tran tr join adj_hd hd on hd.entry_no = tr.entry_no  where tr.ocean is null and hd.posted = 1 and tr.item_ref = '{product.barcode}'"

                        conn = ret_cursor()
                        cursor = conn.cursor()
                        cursor.execute(g_query)
                        for grn_row in cursor.fetchall():
                            ref,ent_date,quantity,line_no,loc_id,loc_to = grn_row
                            # li = [ref,ent_date,quantity,line_no]
                            location = Locations.objects.get(code=loc_id)
                            # print(move_type,li)
                            sent_qty = 0
                            if move_type == 'TR':
                                sent_qty = quantity
                                quantity = quantity * -1
                            ProductMoves.objects.create(
                                product=product,
                                ref=ref,
                                quantity=quantity,
                                date=ent_date,
                                move_type=move_type,
                                location=location,
                            )


                            update_q = ""
                            if move_type == "GR":
                                upd_query = f"UPDATE grn_tran set ocean = 1 where entry_no = '{ref}' and item_ref = '{product.barcode}' and line_no = '{line_no}'"
                            if move_type == "TR":
                                location = Locations.objects.get(code=loc_to)
                                ProductMoves.objects.create(
                                    product=product,
                                    ref=ref,
                                    quantity=sent_qty,
                                    date=ent_date,
                                    move_type=move_type,
                                    location=location,
                                )
                                upd_query = f"UPDATE tran_tr set ocean = 1 where entry_no = '{ref}' and item_ref = '{product.barcode}' and line_no = '{line_no}'"
                            if move_type == 'AD':
                                upd_query = f"UPDATE adj_tran set ocean = 1 where entry_no = '{ref}' and item_ref = '{product.barcode}' and line_no = '{line_no}'"

                            # print(upd_query)
                            cursor.execute(upd_query)
                            cursor.commit()

                        conn.close()
            elif module == 'bill':
                header = data.get('header')
                transactions = data.get('transactions')

                location = Locations.objects.get(code=header.get('loc_id'))
                bill_no = header.get('bill_no')
                bill_ref = header.get('bill_ref')
                pay_mode = header.get('pay_mode')
                bill_date = header.get('bill_date')
                bill_time = header.get('bill_time')
                bill_amt = header.get('bill_amt')

                BillHeader.objects.filter(bill_ref=bill_ref).delete()
                # save header
                BillHeader.objects.create(loc=location,bill_no=bill_no,bill_ref=bill_ref,pay_mode=pay_mode,bill_date=bill_date,bill_time=bill_time,bill_amt=bill_amt)
                # get the just created
                bill = BillHeader.objects.get(bill_ref=bill_ref)

                # save transactions
                for tran in transactions:
                    try:
                        time = tran.get('time')
                        barcode = tran.get('barcode')
                        quantity = Decimal(tran.get('quantity'))
                        price = tran.get('price')
                        prod_id = tran.get('prod_id')
                        tran_type = tran.get('tran_type')
                        line_no = tran.get('line_no')
                        # print(tran_type)

                        if tran_type == 'S':
                            product = Products.objects.get(barcode=barcode)

                            quantity = quantity * -1

                            print(prod_id)
                            BillTrans.objects.create(bill=bill, product=product, quantity=quantity, time=time,
                                                     price=price)
                            print("Saved Start")

                            # Print before creating ProductMoves to see if the code reaches here
                            # print(f"Creating ProductMove for product: {product} with quantity: {quantity}")

                            ProductMoves.objects.create(
                                move_type='PI',
                                product=product,
                                ref=bill_ref,
                                quantity=quantity,
                                date=bill_date,
                                remark=f"POS Sales from {location.descr}",
                                line_no=line_no,
                                location=location
                            )
                            print("Saved Moved")

                    except Exception as e:
                        print(f"Error processing transaction: {e}")





            elif module == 'sample':
                loc_id = data.get('location')
                bill_ref = data.get('bill_ref')
                mypk = data.get('mypk')

                owner = User.objects.get(pk=mypk)
                location = Locations.objects.get(pk=loc_id)

                SampleHd(
                    location=location,
                    owner=owner,
                    bill_ref=bill_ref
                ).save()

                success_response['message'] = "Sampel Recoeded"
                response = success_response

            elif module == 'sample_sync':
                samples = SampleHd.objects.filter(is_sync=False)
                for sample in samples:
                    loc = sample.location
                    host = loc.ip_address
                    db = loc.db
                    db_us = loc.db_user
                    db_ps = loc.db_password

                    # connect to database
                    conn = ret_cursor(host,'',db,db_us,db_ps)
                    cur = conn.cursor()

                    bill_ref = sample.bill_ref
                    qu = (f"SELECT RTRIM(tran_code),RTRIM(tran_desc),mech_no,bill_no,tran_qty,unit_price,tran_amt,tran_type "
                          f"FROM bill_tran where tran_type = 'S' and billRef = '{bill_ref}'")
                    cur.execute(qu)
                    for row in cur.fetchall():
                        tran_code,tran_desc,mech_no,bill_no,tran_qty,unit_price,tran_amt,tran_type= row
                        SampleTran(
                            sample=sample,
                            barcode=tran_code,
                            item_name = tran_desc,
                            mech_no = mech_no,
                            bill_no = bill_no,
                            tran_qty = tran_qty,
                            unit_price=unit_price,
                            tran_amt=tran_amt
                        ).save()


                    conn.close()

                    sample.is_sync = True
                    sample.save()

            elif module =='sync_transfer':
                conn = ret_cursor()
                cursor = conn.cursor()
                t_q = "select entry_no,RTRIM(loc_from),RTRIM(loc_to),entry_date,RTRIM(remark),tot_amt,tot_retail,RTRIM(user_id) from tran_hd where valid = 1 and ocean_sync is null"
                cursor.execute(t_q)
                for row in cursor.fetchall():
                    #print()
                    entry_no,loc_from,loc_to,entry_date,remark,tot_amt,tot_retail,user_id=row
                    #print(entry_no,loc_from,loc_to,entry_date,remark,tot_amt,tot_retail,user_id)

                    l_f = Locations.objects.get(code=loc_from)
                    l_t = Locations.objects.get(code=loc_to)

                    TranHd(
                        entry_no=entry_no.strip(),
                        loc_fr=l_f,
                        loc_to=l_t,
                        entry_date=entry_date,
                        remark=remark,
                        tot_amt=tot_amt,
                        tot_retail=tot_retail,
                        user_id=user_id

                    ).save()

                    hd = TranHd.objects.get(
                        entry_no=entry_no,
                    )

                    tr_q = f"select line_no,RTRIM(item_ref) as 'BARCODE',item_code,RTRIM(item_des) as 'name',unit_qty,packing,um_qty,total_units from tran_tr where entry_no = '{entry_no}'"
                    #print(tr_q)

                    for trow in cursor.execute(tr_q):
                        line_no,barcode,item_code,name,qty,packing,pack_qty,total_qy = trow
                        ti = [line_no,barcode,item_code,name,qty,packing,pack_qty,total_qy]
                        TranTr(
                            transfer=hd,
                            line_no=line_no,
                            barcode=barcode,
                            item_code=item_code,
                            item_descr=name,
                            quantity=qty,
                            packing=packing,
                            pack_qty=pack_qty,
                            total_qty=total_qy,
                        ).save()
                        #print(ti)
                    #print()

                    cursor.execute(f"UPDATE tran_hd set ocean_sync = 1 where entry_no = '{entry_no}'")
                    cursor.commit()

                success_response['message'] = "DONE"
                response = success_response
            elif module == 'mark_butch':
                barcodes = data.get('barcode').split(',')
                for barcode in barcodes:
                    if Products.objects.filter(barcode=barcode).exists():
                        Products.objects.filter(barcode=barcode).update(is_butch=True)
                response = success_response
            elif module == 'kofi_ghana':
                #print(data)
                location = data.get('location')
                barcode = data.get('barcode')
                quantity = data.get('quantity')

                ButchSales(barcode=barcode, quantity=quantity,location_id=location).save()
                success_response['message'] = "Product Added"
                response = success_response

            elif module == 'butch_live_sales':
                pending = ButchSales.objects.filter(is_checked=False)
                for xitem in pending:
                    barcode = xitem.barcode
                    location = xitem.location
                    quantity = xitem.quantity

                    cursor = ret_cursor(location.ip_address,'',location.db,location.db_user,location.db_password)
                    qu = f"SELECT top(1) bill_no,billRef,tran_time,bill_date,tran_desc,tran_qty FROM bill_tran where tran_code = '{barcode}' and tran_qty like '{quantity}%'"
                    cursor.execute(qu)

                    rows = cursor.fetchall()
                    if len(rows) == 0:
                        pass

                    else:
                        first = rows[0]
                        bill_no, bill_ref, tran_time, bill_date, name, tran_qty = first
                        ButcheryLiveTransactions(barcode=barcode, name=name, quantity=tran_qty, time=tran_time,
                                                 bill_no=bill_no, bill_ref=bill_ref,butch_log=xitem).save()

                        xitem.is_checked = True
                        xitem.save()




                # bill_no = data.get('bill_no')
                # bill_ref = data.get('bill_ref')
                # barcode = data.get('barcode')
                # name = data.get('name')
                # quantity = data.get('quantity')
                # time = data.get('time')
                #
                # #print(data)
                #
                # ButcheryLiveTransactions(barcode=barcode, name=name, quantity=quantity, time=time,
                #                          bill_no=bill_no,bill_ref=bill_ref).save()
                response = success_response

            elif module == 'bolt_group':  # add bolt group
                name = data.get('name')
                entity = data.get('entity')
                ent = BusinessEntityTypes.objects.get(entity_type_name=entity)
                subs = data.get('subs')
                print(subs)
                try:
                    if BoltGroups.objects.filter(name=name,entity=ent).count() == 0:
                        BoltGroups(name=name,entity=ent).save()

                    group = BoltGroups.objects.get(name=name,entity=ent)
                    for sub in subs:
                        print(sub)
                        if BoltSubGroups.objects.filter(name=sub,group=group,entity=ent).count() == 0:
                            BoltSubGroups.objects.create(group=group, name=sub,entity=ent)

                    success_response['message'] = BoltGroups.objects.get(name=name,entity=ent).pk

                except Exception as e:
                    success_response['status_code'] = 500
                    success_response['message'] = str(e)




            elif module == 'sync_retail_suppliers':
                query = "select supp_code as 'code',supp_name as 'name',contact as 'person',phone1 as 'phone',address1 as 'email',address2 as 'city',country_id as 'country' from supplier"
                conn = ret_cursor()
                cursor = conn.cursor()
                cursor.execute(query)
                counts = 0
                not_counted = 0
                for supplier in cursor.fetchall():
                    code = supplier[0].strip()
                    name = supplier[1].strip()
                    person = supplier[2].strip()
                    phone = '' if supplier[3] is None else supplier[3].strip()
                    email = '' if supplier[4] is None else supplier[4].strip()
                    city = '' if supplier[5] is None else supplier[5].strip()
                    country = '' if supplier[6] is None else supplier[6].strip()

                    try:
                        supp, create = ProductSupplier.objects.get_or_create(code=code, name=name, phone=phone,
                                                                             person=person,
                                                                             email=email, city=city, country=country)
                        counts = counts + 1
                    except Exception as e:
                        not_counted = not_counted + 1

                conn.close()
                success_response['message'] = f"{counts} / {counts + not_counted} suppliers synced"
                response = success_response

            elif module == 'sync_retail_groups':
                query = f"select group_code,group_des from group_mast"
                conn = ret_cursor()
                cursor = conn.cursor()
                cursor.execute(query)
                count = 0
                not_counted = 0

                for group in cursor.fetchall():
                    code, name = group[0].strip(), group[1].strip()
                    try:
                        get, add = ProductGroup.objects.get_or_create(code=code, name=name)
                        count = count + 1
                    except Exception as e:
                        not_counted = not_counted + 1

                conn.close()
                success_response['message'] = f"{count} / {count + not_counted} groups synced"
                response = success_response

            elif module == 'sync_retail_sub_groups':
                conn = ret_cursor()
                cursor = conn.cursor()
                query = f"SELECT group_code,(select group_des from group_mast where group_code = sub_group.group_code) as 'group', sub_group,sub_group_des from sub_group"
                cursor.execute(query)
                count = 0
                not_counted = 0
                errors = []

                for sub_group in cursor.fetchall():

                    group_code, group_des, sub_group_code, sub_group_des = sub_group[0], sub_group[1].strip(), \
                        sub_group[2].strip(), sub_group[3].strip()
                    group, add = ProductGroup.objects.get_or_create(code=group_code, name=group_des)

                    if add:
                        group = ProductGroup.objects.get(code=group_code, name=group_des)

                    try:
                        sub, create = ProductSubGroup.objects.get_or_create(group=group, code=sub_group_code,
                                                                            name=sub_group_des)
                        count = count + 1
                    except Exception as e:
                        errors.append(e)
                        not_counted = not_counted + 1

                conn.close()
                success_response['message'] = f"{count} / {count + not_counted} sub groups synced with errors {errors}"
                response = success_response

            elif module == 'sync_retail_products':
                entity = data.get('entity',BusinessEntityTypes.objects.get(entity_type_name='retail').pk)
                conn = ret_cursor()
                cursor = conn.cursor()
                query = ("SELECT item_code, barcode, item_des, (SELECT group_des FROM group_mast WHERE group_mast.group_code = prod_mast.group_code) AS 'group', (SELECT sub_group_des FROM sub_group WHERE sub_group.group_code = prod_mast.group_code AND sub_group.sub_group = prod_mast.sub_group) "
                         "AS 'sub_group', (SELECT supp_name FROM supplier WHERE supplier.supp_code = prod_mast.supp_code) AS 'supplier', retail1 FROM prod_mast WHERE item_type != 0 order by item_code desc")
                cursor.execute(query)
                saved = 0
                not_synced = 0
                error = []
                for product in cursor.fetchall():
                    code = product[0]
                    barcode = str(product[1]).strip()
                    item_des = product[2].strip()
                    group = product[3]


                    try:
                        sub_group = product[4].strip()
                    except Exception as e:
                        sub_group = product[4]



                    supplier = product[5]
                    retail1 = product[6]

                    # add to products
                    # subgroup = ProductSubGroup.objects.get(code='999')
                    if ProductSubGroup.objects.filter(name=sub_group).count() == 1:
                        subgroup = ProductSubGroup.objects.get(name=sub_group)
                        # delete product

                    if Products.objects.filter(code=code).exists():
                        # update
                        prod = Products.objects.get(code=code)
                        prod.barcode = barcode
                        prod.name = item_des
                        prod.price = retail1
                        prod.subgroup = subgroup
                        prod.entity_id = entity

                        prod.save()
                        print("Okay")

                    else:
                        # save new
                        Products.objects.get_or_create(subgroup=subgroup, name=item_des, barcode=barcode,
                                                           code=code, price=retail1,is_active=True,entity_id = entity)
                        saved = saved + 1
                        print("CREATED", code, barcode, product[2])
                    # else:
                    #     not_synced = not_synced + 1
                    #     error.append(f"{barcode} - {item_des} # sub group {sub_group} / does not exist")

                conn.close()
                success_response['message'] = {
                    "message":f"{saved} / {saved + not_synced} products synced",
                    "errors":error
                }
                response = success_response

            elif module == 'sync_retail_barcodes':
                conn = ret_cursor()
                cursor = conn.cursor()
                cursor.execute("SELECT RTRIM(item_code),RTRIM(barcode) FROM barcode")
                for row in cursor.fetchall():
                    item_code,barcode = row

                    try:
                        Barcode.objects.create(
                            product=Products.objects.get(code=item_code),
                            barcode=barcode
                        )
                        print([item_code, barcode])
                    except Exception as e:
                        print(e)
                        pass

            elif module == 'build_barcode':
                for product in Products.objects.all():
                    bars = ""
                    for barcodes in Barcode.objects.filter(product=product):
                        barcode = f"{barcodes.barcode},"
                        bars += barcode
                    if len(bars) > 0:
                        bars = bars[:-1]
                        print(bars)
                        product.barcode = bars
                        product.save()

            elif module == 'update_stock':
                products = Products.objects.all()
                count = products.count()

                for product in products:
                    barcode = product.barcode.strip()
                    item_code = product.code

                    stock = get_stock(item_code)
                    #print(item_code, stock)
                    spintex = stock.get('spintex')
                    nia = stock.get('nia')
                    osu = stock.get('osu')
                    warehouse = stock.get('warehouse')
                    kitchen = stock.get('kitchen')

                    if Stock.objects.filter(product=product,location='001').exists():
                        sp_stk = Stock.objects.get(product=product,location='001')
                        sp_stk.quantity = spintex
                    else:
                        Stock(product=product, quantity=spintex, location='001').save()


                    if Stock.objects.filter(product=product,location='202').exists():
                        sp_stk = Stock.objects.get(product=product,location='202')
                        sp_stk.quantity = nia
                    else:
                        Stock(product=product, quantity=nia, location='202').save()


                    if Stock.objects.filter(product=product,location='205').exists():
                        sp_stk = Stock.objects.get(product=product,location='205')
                        sp_stk.quantity = osu
                    else:
                        Stock(product=product, quantity=osu, location='205').save()

                    if Stock.objects.filter(product=product,location='999').exists():
                        sp_stk = Stock.objects.get(product=product,location='999')
                        sp_stk.quantity = warehouse
                    else:
                        Stock(product=product, quantity=warehouse, location='999').save()

                    if Stock.objects.filter(product=product, location='201').exists():
                        sp_stk = Stock.objects.get(product=product, location='201')
                        sp_stk.quantity = kitchen
                    else:
                        Stock(product=product, quantity=kitchen, location='201').save()


                    #print(product.name)


                    count -= 1
                    #print(f"Ramaining {count}")
                success_response['message'] = "Stock Updated"

            elif module == 'recipe_group':
                name = data.get('name')
                us = data.get('mypk')
                owner = User.objects.get(pk=us)
                RecipeGroup(name=name, owner=owner).save()

                success_response['message'] = "Group Added!!"
            elif module == 'recipe_product':
                gk = data.get('gk')
                group = RecipeGroup.objects.get(pk=gk)
                us = data.get('mypk')
                owner = User.objects.get(pk=us)
                name = data.get('name')
                barcode = data.get('barcode')
                si_unit = data.get('si_unit')

                RecipeProduct(group=group, name=name, barcode=barcode, si_unit=si_unit, owner=owner).save()
                recipe = RecipeProduct.objects.get(group=group, name=name, barcode=barcode, si_unit=si_unit,
                                                   owner=owner)
                success_response['message'] = recipe.pk

            elif module == 'recipe_items':
                pro_key = data.get('product')
                product = RecipeProduct.objects.get(pk=pro_key)

                us = data.get('mypk')
                owner = User.objects.get(pk=us)
                items = data.get('items')

                # delete all recipes
                Recipe.objects.filter(product=product).delete()

                for item in items:
                    #print(item)
                    name = item['name']
                    qty = item['qty']
                    si_unit = item['si_unit']

                    Recipe(product=product, name=name, owner=owner, quantity=qty, si_unit=si_unit).save()

                success_response['message'] = "Recipe Saved Successfully"

            # retrieve frozen stock
            elif module == 'retrieve_frozen_stock':
                mypk = data.get('mypk')
                entry = data.get('')

                # db query
                conn = ret_cursor()
                cursor = conn.cursor()
                query = f"select loc_id,ref_no,ld_date,remarks,grp,grp_from,grp_to from stock_keep_hd where ref_no = '{entry}'"
                cursor.execute(query)
                hd = cursor.fetchone()

                if hd:
                    loc_id, ref_no, date_kept, remarks, is_group, st_grp, end_grp = hd
                    owner = User.objects.get(pk=mypk)
                    loc = Locations.objects.get(loc_id=loc_id)

                    StockHd(
                        loc=loc_id, ref_no=ref_no, date_kept=date_kept,
                        remarks=remarks, is_group=is_group, st_grp=st_grp, end_grp=end_grp, owner=owner
                    ).save()

                    stock_hd = StockHd.objects.get(
                        loc=loc_id, ref_no=ref_no, date_kept=date_kept,
                        remarks=remarks, is_group=is_group, st_grp=st_grp, end_grp=end_grp, owner=owner
                    )

                    # save transactions

                    success_response['status_code'] = 200
                    success_response['message'] = loc_id
                else:
                    success_response['status_code'] = 404
                    success_response['message'] = f"No stock keep for entry number {entry}"

                conn.close()
            elif module == 'sync_locations':
                conn = ret_cursor()
                cursor = conn.cursor()
                cursor.execute("SELECT br_code,RTRIM(br_name) from branch")

                for br in cursor.fetchall():
                    code, name = br
                    if Locations.objects.filter(code=code).count() == 0:
                        # create location
                        Locations(code=code, descr=name, owner_id=1).save()
                conn.close()
            elif module == 'bulk_monitor':
                barcodes = data.get('barcodes')
                enable = data.get('enable')
                yes = 0
                no = 0
                #print(enable)
                for barcode in barcodes:

                    # get product
                    if Products.objects.filter(barcode=barcode).count() == 1:
                        product = Products.objects.get(barcode=barcode)
                        product.stock_monitor = enable
                        product.save()
                        yes += 1
                    else:
                        no += 1

                success_response['message'] = f"Bulk Monitor Flag set to {enable}  for ({yes}/{yes+no}) products"

            elif module == 'update_raw_stock':
                r_cur = ret_cursor()
                r_cur.execute("SELECT item_code,item_des from prod_mast order by item_des")
                for row in r_cur.fetchall():
                    item_code,name = row
                    #print(name)
                    updateStock(item_code)

                success_response['message'] = "Operation Successful"


            else:
                success_response['status_code'] = 404
                success_response['message'] = f"no {method} method with module called {module}"

        elif method == 'VIEW':
            arr = []
            if module == 'location_master':
                locations = Locations.objects.all()
                for location in locations:
                    arr.append(location.obj())

                success_response['message'] = arr
                response = success_response

            elif module == 'butch_items':
                target_date = data.get('target_date',timezone.now().date())
                for item in Products.objects.filter(is_butch=True).order_by('name'):
                    obj = item.obj(target_date)

                    arr.append(obj)
                success_response['message'] = arr
                response = success_response

            elif module == 'suppliers':
                suppliers = ProductSupplier.objects.all().order_by('name')
                for supplier in suppliers:
                    arr.append(supplier.obj())

                success_response['message'] = arr
                response = success_response

            elif module == 'sample':
                samples = SampleHd.objects.all()

                for sample in samples:
                    arr.append(sample.obj())

                success_response['message'] = arr
                response = success_response

            elif module == 'price_change':
                from datetime import datetime
                current_date = datetime.now().strftime('%Y-%m-%d')
                date_from = data.get('from',current_date)
                date_to = data.get('to',current_date)
                query = (f"select pl.item_code as 'item_code',RTRIM(pm.barcode) as 'BARCODE',RTRIM(pm.item_des) as 'NAME',"
                         f"retail as 'to',prev_price1 as 'from',pchange_date,pl.user_id from price_lev pl join prod_mast pm "
                         f"on pm.item_code = pl.item_code "
                         f"where pchange_date between '{date_from}' and '{date_to}'")
                #print(query)
                conn = ret_cursor()
                cursor = conn.cursor()
                cursor.execute(query)
                from openpyxl import Workbook
                from openpyxl.styles import Font
                doc = data.get('doc','json')
                arr = []
                book = Workbook()
                sheet = book.active
                sheet.merge_cells('A1:E1')
                sheet['A1'] = f"OCEAN GENERATED PRICE CHANGE"
                header = ['BARCODE','NAME','OLD PRICE','NEW PRICE',"% Margin",'BY']
                sheet.append(header)
                for row in cursor:
                    item_code,barcode,name,retail,prev_price1,pchange_date,user_id=row
                    percentage_margin = str(format_currency(percentage_difference(prev_price1,retail))).replace('$','') + "%"
                    #print(item_code,barcode,name,retail,prev_price1,pchange_date,user_id)
                    if doc == 'excel':
                        li = [barcode,name,prev_price1,retail,percentage_margin,user_id]
                        sheet.append(li)
                        #print(percentage_margin)
                        if ")" in percentage_margin:
                            row_index = sheet.max_row
                            for cell in sheet[row_index]:
                                cell.font = Font(color='FF0000')


                    else:
                        arr.append({
                            'barcode':barcode,
                            'name':name,
                            'old_price':prev_price1,
                            'new_price':retail,
                            'by':user_id,
                            'date':pchange_date,
                            'margin':percentage_margin
                        })

                if doc == 'excel':
                    file_name = f'static/general/tmp/price_change_between_{date_from}_to_{date_to}.xlsx'
                    book.save(file_name)
                    arr = file_name

                conn.close()
                success_response['message'] = arr
                response = success_response

            elif module == 'entity_prod_mast':
                entity_name = data.get('entity_name')
                filter_flag = data.get('filter_flag','*')
                entity = BusinessEntityTypes.objects.get(entity_type_name=entity_name)
                items = Products.objects.filter(entity=entity)


                for item in items:
                    if filter_flag == 'not_on_bolt':
                        if not item.is_on_bolt():
                            arr.append(item.obj())
                    else:
                        arr.append(item.obj())


                success_response['message'] = arr
                response = success_response

            elif module == 'menu':
                entity = data.get('entity','*')
                to = data.get('to')
                if entity == '*':
                    items = BoltItems.objects.all()
                else:
                    items = BoltItems.objects.filter(entity_id=entity)

                if to:
                    items = items.exclude(menu_id=to)
                tot = items.count()
                x = 1
                for item in items:
                    print(f"{x} / {tot}")
                    arr.append(item.obj())
                    x += 1

                success_response['message'] = arr
                response = success_response

            elif module == 'butch_live_monitor':
                records = ButchSales.objects.filter(is_checked=False)
                t_date = data.get('date',timezone.now())
                if t_date:
                    records.filter(date_added=t_date)
                for record in records:
                    arr.append(record.obj())

                success_response['message'] = arr
                response = success_response

            elif module == 'expiry':
                days = data.get('days')
                doc = data.get('document','json')
                if doc == 'excel':
                    import openpyxl
                    book = openpyxl.Workbook()
                    sheet = book.active

                    header = ['EXPIRY DATE','GRN','BARCODE','ITEM CODE','ITEM NAME','DAYS REMAINING','WAREHOUSE',"SPINTEX","NIA","OSU"]
                    sheet.append(header)
                arr = []
                query = f"""
                    WITH DistinctBarcodes AS (
                        SELECT
                            item_ref AS barcode,
                            MIN(exp_date) AS exp_date
                        FROM grn_tran
                        WHERE exp_date BETWEEN GETDATE() AND DATEADD(DAY, {days}, GETDATE())
                        GROUP BY item_ref
                    )
                    SELECT
                        t.exp_date,
                        t.entry_no,
                        t.item_ref AS barcode,
                        t.item_code,
                        t.item_des AS name,
                        DATEDIFF(day,GETDATE(),t.exp_date) AS days_to_expire
                    FROM grn_tran t
                    INNER JOIN DistinctBarcodes d
                    ON t.item_ref = d.barcode
                    AND t.exp_date = d.exp_date
                    WHERE t.exp_date BETWEEN GETDATE() AND DATEADD(DAY, 90, GETDATE())
                    ORDER BY t.item_des;

                """
                #print(query)
                conn = ret_cursor()
                cursor = conn.cursor()
                cursor.execute(query)
                for row in cursor.fetchall():
                    exp_date,entry_no,barcode,item_code,item_des,days_to_expire = row
                    stock = 0
                    warehouse_stock,spintex_stock,nia_stock,osu_stock = [0,0,0,0]
                    if Products.objects.filter(code=item_code).count() == 1:
                        product = Products.objects.get(code=item_code)
                        stock = Stock.objects.filter(product=product).aggregate(Sum('quantity'))['quantity__sum']
                        stock = get_stock(item_code)

                        warehouse_stock = stock.get('warehouse')
                        spintex_stock =  stock.get('spintex')
                        nia_stock =  stock.get('nia')
                        osu_stock =  stock.get('osu')
                        
                        # warehouse_stock = Stock.objects.filter(product=product,location='999').aggregate(Sum('quantity'))['quantity__sum']
                        # spintex_stock = Stock.objects.filter(product=product,location='001').aggregate(Sum('quantity'))['quantity__sum']
                        # nia_stock = Stock.objects.filter(product=product,location='202').aggregate(Sum('quantity'))['quantity__sum'],
                        # osu_stock = Stock.objects.filter(product=product,location='205').aggregate(Sum('quantity'))['quantity__sum']

                    li = [exp_date,entry_no,barcode,item_code,item_des,days_to_expire,str(warehouse_stock),str(spintex_stock),str(nia_stock),str(osu_stock)]
                    if doc == 'json':
                        arr.append({
                            'expiry_date': exp_date,
                            'grn':entry_no,
                            'barcode':barcode,
                            'item_code':item_code,
                            'item_des':item_des,
                            'stock':stock,
                            'warehouse_stock':warehouse_stock,
                            'spintex_stock':spintex_stock,
                            'osu_stock':osu_stock,
                            'nia_stock':nia_stock
                        })

                    if doc == 'excel':
                        sheet.append(li)

                if doc == 'excel':
                    file = f'static/general/tmp/{days}_days_expiry.xlsx'
                    book.save(file)
                    arr = file

                conn.close()
                success_response['message'] = arr
                response = success_response

            elif module == 'transactions_today':
                from datetime import datetime
                today = datetime.now().strftime('%Y-%m-%d')
                limit = data.get('limit',0)
                # get some alphas
                transactions = BillTrans.objects.filter(bill__bill_date=today).order_by('-time')
                if limit > 0:
                    transactions = transactions[:limit]
                arr = []
                for tran in transactions:
                    arr.append(tran.obj())

                success_response['message'] = arr

            elif module == 'analysis_for_transfer':
                loc_id = data.get('location')
                loc = Locations.objects.get(code=loc_id)
                ripe = data.get('ripe')
                out = data.get('view','json')

                if out == 'excel':
                    import openpyxl
                    book = openpyxl.Workbook()
                    sheet = book.active

                    sheet.title = loc.descr
                    header = ['SN',"ITEM CODE","BARCODE","NAME","TRANSFER", "TRANSFER QUANTITY","SOLD","SOLD %","SOLD(2 weeks)","HEALTH","SYSTEM STOCK"]
                    sheet.append(header)

                #print("JESUS")

                #print(data)
                #print("JESUS")
                if ripe == 'YES':
                    #print("YES RIPE")
                    to_sends = StockToSend.objects.filter(location=loc,healthy=True).order_by('name')
                elif ripe == 'NO':
                    #print("NO RIPE")
                    to_sends = StockToSend.objects.filter(location=loc,healthy=False).order_by('name')
                else:
                    to_sends = StockToSend.objects.filter(location=loc).order_by('name')



                line = 1
                for to_send in to_sends:
                    if out == 'json':
                        arr.append(to_send.obj())

                    if out == 'html':
                        arr.append(to_send.html(line))

                    if out == 'excel':
                        header = ['SN', "ITEMCODE", "BARCODE", "NAME", "TRANSFER", "TRANSFER QUANTITY", "SOLD",
                                  "HEALTH"]

                        # if Products.objects.filter(code=to_send.item_code).count() == 1:
                        #     pd = Products.objects.get(code=to_send.item_code)
                        #     stk  = Stock.objects.filter(product=pd,location=loc_id).aggregate(Sum('quantity'))['quantity__sum']
                        # else:
                        #     stk = get_stock(item_code=to_send.item_code).get(loc_id)
                        stk = RawStock.objects.filter(loc_id=loc_id,prod_id=to_send.item_code).aggregate(Sum('qty'))['qty__sum']
                        tr_en = f"{to_send.last_transfer_entry} ({to_send.last_transfer_date})"
                        li = [line,to_send.item_code,to_send.barcode,to_send.name,tr_en,to_send.last_transfer_quantity,to_send.sold_quantity,to_send.percentage_sold,to_send.cust_sold,to_send.healthy,stk]
                        #print(li)
                        sheet.append(li)

                    line+=1

                if out == 'excel':
                    file_name = f'static/general/tmp/{loc.descr}.xlsx'
                    book.save(file_name)
                    arr = file_name
                success_response['message'] = arr
                response = success_response

            elif module == 'sync_analyse_for_transfer':
                arr = []
                loc_id = data.get('loc','*')
                if loc_id == '*':
                    locations = Locations.objects.filter(type='retail')
                else:
                    locations = Locations.objects.filter(type='retail',code=loc_id)
                for loc in locations:
                    loc_id = loc.code
                    loc_code = loc.code
                    healthy = 0
                    not_healthy = 0
                    conn = ret_cursor()
                    cursor = conn.cursor()
                    qq = "SELECT RTRIM(barcode),RTRIM(item_des),item_code from prod_mast where item_type != 0 and supp_code != 'S00008' order by item_des"
                    # qq = "SELECT RTRIM(barcode),RTRIM(item_des),item_code from prod_mast where item_type != 0 and barcode = 'SD34316-7'"
                    cursor.execute(qq)

                    for product in cursor.fetchall():
                        barcode,name,item_code = product

                        # check last sent
                        cursor.execute(
                            f"select top(1) th.entry_no,th.entry_date,tr.total_units from tran_tr tr  right join tran_hd th on th.entry_no = tr.entry_no where tr.item_code = '{item_code}' and th.loc_from = '999' and th.loc_to = '{loc_id}' order by th.entry_date desc ")
                        last_tran_row = cursor.fetchone()

                        last_tr_entry, last_tran_date, last_tran_qty, sold_qty,two_weeks_sales = ['none', 'none', 0, 0,0]
                        total_out = 0
                        adjusted = 0
                        transfered = 0
                        if last_tran_row is not None:
                            last_tr_entry, last_tran_date, last_tran_qty = last_tran_row
                            inv_cursor = cursor
                            inv_sales = (f"select sum(tran_qty) from pos_tran where prod_id = '{item_code}' and entry_no in "
                                         f"(select entry_no from pos_tran_hd where location_id = '{loc_id}' and entry_date > '{last_tran_date}')")
                            inv_cursor.execute(inv_sales)
                            sold_resp = inv_cursor.fetchone()
                            if sold_resp is not None:
                                if sold_resp[0] is not None:
                                    sold_qty = sold_resp[0]

                                today = timezone.now().date()
                                from datetime import timedelta

                                two_weeks_ago = today - timedelta(days=14)

                                sold_in_2_q = f"SELECT sum(tran_qty) as 'sold_qty' FROM history_tran where bill_date between '{two_weeks_ago}' and '{today}' and prod_id = '{item_code}'"
                                sold_in_2_q = f"""
                                    select sum(tran_qty) from pos_tran where prod_id = '{item_code}' and entry_no in 
                                    (select entry_no from pos_tran_hd where location_id = '{loc_id}' and entry_date between '{two_weeks_ago}' and '{today}')
                                """
                                # #print("WEEK")
                                # #print(sold_in_2_q)
                                # #print("WEEK")
                                two_weks_resp = inv_cursor.execute(sold_in_2_q)
                                tw_fetch = two_weks_resp.fetchone()
                                two_weeks_sales = 0
                                if tw_fetch is not None:
                                    if tw_fetch[0] is not None:
                                        two_weeks_sales = tw_fetch[0]

                                #print("2WEEKS",two_weeks_sales,"2WEEKS")

                                # adjusted
                                inv_cursor.execute(
                                    f"SELECT SUM(total_units) FROM adj_tran where item_code = '{item_code}'and  entry_no in (SELECT entry_no from adj_hd where entry_date > '{last_tran_date}' and loc_id = '{loc_id}' )")
                                adjusted = inv_cursor.fetchone()[0] or 0

                                # transferd
                                inv_cursor.execute(
                                    f"SELECT SUM(total_units) FROM tran_tr where item_code = '{item_code}'  and entry_no in (SELECT entry_no from tran_hd where entry_date > '{last_tran_date}' and loc_from = '{loc_id}' )")
                                transfered = inv_cursor.fetchone()[0] or 0

                                total_out = sold_qty + adjusted

                        health = False
                        if total_out > Decimal(0.5) * last_tran_qty:
                            health = True
                            healthy += 1
                        else:
                            not_healthy += 1

                        percentage_sold = 0
                        if last_tran_qty != 0 and sold_qty != 0:  # Ensure we don't divide by zero
                            percentage_sold = (Decimal(sold_qty) / Decimal(last_tran_qty)) * 100

                        if percentage_sold > 35:
                            health = True

                        obj = {
                                    "location":loc.descr,
                                    'item_code': item_code,
                                    'name': name,
                                    'barcode':barcode,
                                    'last_tran_entry': last_tr_entry,
                                    'last_tran_date': last_tran_date,
                                    "last_tran_qty": last_tran_qty,
                                    'sold_qty': sold_qty,
                                    "percentage_sold": percentage_sold,
                                    'health': health,
                                    "two_weeks_sales":two_weeks_sales,

                                }

                        StockToSend.objects.filter(item_code=item_code,location=loc).delete()
                        #print("SAVING POINT")
                        #print(two_weeks_sales)
                        #print("SAVING POINT")

                        StockToSend(location=loc,item_code=item_code, name=name, barcode=barcode,last_transfer_entry=last_tr_entry,
                                    last_transfer_date=last_tran_date,last_transfer_quantity=last_tran_qty,sold_quantity=sold_qty,
                                    percentage_sold=percentage_sold,healthy=health,cust_sold=two_weeks_sales).save()

                        arr.append(obj)
                        #print(name)
                        # #print(obj)


                conn.close()
                success_response['message'] = arr
                response = success_response

            elif module == 'sync_sales':
                loc_code = data.get('loc','*')
                if loc_code == '*':
                    locs = Locations.objects.filter(type='retail')
                else:
                    locs = Locations.objects.filter(type='retail',code=loc_code)
                for loc in locs:
                    host = loc.ip_address
                    db = loc.db
                    db_pass = loc.db_password
                    db_user = loc.db_user

                    # address = loc.get('server')
                    # user = loc.get('user')
                    # key = loc.get('key')
                    # db = loc.get('db')
                    # loc = loc.get('loc')

                    conn = ret_cursor(host,'',db,db_user,db_pass)
                    cursor = conn.cursor()

                    cursor.execute(
                        "SELECT bill_no,billRef,bill_date,bill_end_time,bill_amt FROM bill_header where bill_type = 'S' and ocean_sync = 0")

                    for row in cursor.fetchall():
                        bill_no, billRef, bill_date, bill_end_time, bill_amt = row

                        pay_method = "Unknonw"
                        transactions = []
                        cursor.execute(
                            f"SELECT RTRIM(tran_code),RTRIM(tran_desc),tran_qty,unit_price,tran_type,tran_time,prod_id FROM bill_tran where billRef = '{billRef}' and tran_valid = 'Y'")

                        for trow in cursor.fetchall():
                            barcode, desc, qty, unit_price, tran_type, tran_time, prod_id = trow

                            if tran_type == 'S':
                                ti = {"barcode": barcode, "name": desc, "quantity": str(qty), "price": str(unit_price),
                                      "time": tran_time, "prod_id": prod_id, 'tran_type': tran_type}
                                transactions.append(ti)

                            cursor.execute(
                                f"SELECT tran_desc from bill_tran where billRef = '{billRef}' and tran_type = 'P'")

                            pay_method = cursor.fetchone()[0]

                        li = [bill_no, billRef, bill_date, bill_end_time, pay_method]
                        print(li)

                        payload = {}
                        header = {}
                        header['loc_id'] = loc.code
                        header['bill_no'] = str(bill_no)
                        header['bill_ref'] = billRef
                        header['pay_mode'] = pay_method
                        header['bill_date'] = str(bill_date).split(' ')[0]
                        header['bill_end_time'] = str(bill_end_time)
                        header['bill_time'] = str(bill_end_time)
                        header['bill_amt'] = str(bill_amt)

                        data = {}

                        payload['module'] = 'bill'
                        data['header'] = header
                        data['transactions'] = transactions
                        payload['data'] = data


                        url = "http://192.168.2.60/retail/api/"

                        payloadx = json.dumps(payload)
                        headers = {
                            'Content-Type': 'application/json'
                        }

                        response = requests.request("PUT", url, headers=headers, data=payloadx)
                        json_response = json.loads(response.text)
                        status_code = json_response.get('status_code')
                        message = json_response.get('message')
                        # print(status_code,message)
                        if status_code == 200:
                            print(pay_method, billRef, message, "SUCCESS")
                            # update bill
                            update_query = f"UPDATE bill_header set ocean_sync = 1 where billRef = '{billRef}'"
                            print(update_query)
                            cursor.execute(update_query)
                            cursor.commit()
                        else:
                            print(billRef, message)





                    conn.close()

            elif module == 'sales':
                loc_id = data.get('loc','*')
                date_from = data.get('date_from',timezone.now().date())
                date_to = data.get('date_to',timezone.now().date())

                if loc_id == '*':
                    sales = RetailSales.objects.filter(bill_date__range=[date_from,date_to])
                else:
                    location = Locations.objects.get(code=loc_id)
                    sales = RetailSales.objects.filter(location=location,bill_date__range=[date_from,date_to])

                sales.order_by('-tran_time')

                arr = []
                for sale in sales:
                    arr.append(sale.obj())

                success_response['message'] = arr
                response = success_response

            elif module == 'bolt_products':
                pk = data.get('key') or '*'

                if pk == '*':

                    items = BoltItems.objects.all()
                else:
                    items = BoltItems.objects.filter(Q(pk=pk) | Q(barcode=pk))

                item_list = []
                for item in items:
                    item_list.append({
                        'pk': item.pk,
                        'barcode': item.product.barcode,
                        'item_des': item.product.name,
                        'price': item.price,
                        'stock_nia': item.stock_nia,
                        'stock_spintex': item.stock_spintex,
                        'stock_osu': item.stock_osu,
                        'group': {
                            'pk': item.group.pk,
                            'name': item.group.name
                        },
                        'subgroup': {
                            'pk': item.subgroup.pk,
                            'name': item.subgroup.name
                        }
                    })

                success_response['message'] = item_list

            elif module == 'bolt_group':
                pk = data.get('key') or '*'
                entity = BusinessEntityTypes.objects.get(pk=data.get('entity'))
                grps = []
                if pk == '*':
                    groups = BoltGroups.objects.all().order_by('name')
                else:
                    groups = BoltGroups.objects.filter(pk=pk).order_by('name')

                groups = groups.filter(entity=entity)

                for group in groups:
                    grps.append({
                        'pk': group.pk,
                        'name': group.name
                    })

                success_response['message'] = grps

            elif module == 'bolt_sub_group':
                pk = data.get('key') or '*'
                entity_pk = data.get('entity')
                entity = BusinessEntityTypes.objects.get(pk=entity_pk)
                grps = []
                if pk == '*':
                    groups = BoltSubGroups.objects.filter(entity=entity).order_by('name')
                else:
                    groups = BoltSubGroups.objects.filter(group_id=pk,entity=entity).order_by('name')

                for group in groups:
                    grps.append({
                        'pk': group.pk,
                        'name': group.name
                    })

                success_response['message'] = grps

            elif module == 'bolt_price_change':
                send = data.get('send_mail') or 'no'
                rate_inc = data.get('rate_at', 20)
                entity_id = data.get('entity_id',1)
                conn = ret_cursor()
                cursor = conn.cursor()
                import openpyxl
                worksheet = openpyxl.Workbook()
                sheet = worksheet.active
                sheet['A1'] = "SKU"
                sheet['B1'] = "NAME"
                sheet['C1'] = "PRICE"
                sheet_row = 2
                providers_id = ""

                from datetime import datetime
                current_datetime = datetime.now()
                formatted_datetime = current_datetime.strftime("%Y_%m_%d_%H_%M_%S")

                price_change_file = f"static/general/tmp/price_changes_as_of_{formatted_datetime}.csv"

                for bolt_id in BOLT_PROVIDER_ID:
                    print(bolt_id)
                    row = BOLT_PROVIDER_ID[bolt_id]
                    providers_id += f"{row.get('id')},"

                # remove last comma
                providers_id = providers_id[:-1]

                import csv
                with open(price_change_file,mode='w',newline="") as file:
                    writer = csv.writer(file)
                    hd = ['SKU','Price','Location']
                    writer.writerow(["SKU","Selling price","Provier IDs"])
                    all = BoltItems.objects.all().count()
                    x = 1
                    for item in BoltItems.objects.filter(menu_id=entity_id):
                        print(f"{x} / {all}")
                        x += 1
                        barcode = item.product.barcode.replace('.0', '')
                        # print(data)
                        # get product detail

                        product = cursor.execute(
                            f"SELECT retail1,item_code,item_type from prod_mast where barcode = '{barcode}'").fetchone()
                        if product is not None:
                            price, item_code, item_type = product
                            new_price = Decimal(product[0]) / Decimal((1 - rate_inc / 100))
                            pl = [barcode,new_price,providers_id]
                            writer.writerow(pl)
                            # sheet[f"A{sheet_row}"] = barcode
                            # sheet[f"B{sheet_row}"] = item.product.name
                            # sheet[f"C{sheet_row}"] =
                            # sheet[f"D{sheet_row}"] = providers_id
                            # sheet_row += 1
                            item.price = price




                price_count = sheet_row - 2


                from datetime import datetime
                current_datetime = datetime.now()
                formatted_datetime = current_datetime.strftime("%Y_%m_%d_%H_%M_%S")

                # price_change_file = f"static/general/tmp/price_changes_as_of_{formatted_datetime}.xlsx"


                # worksheet.save(price_change_file)



                success_response['message'] = {
                    'price_change': price_change_file
                }

            elif module == 'export_items':
                key = data.get('key')
                format = data.get('format')
                if key == '*':
                    items = BoltItems.objects.all()
                    file_name = f"static/general/bolt/bolt-items.xlsx"
                else:
                    items = BoltItems.objects.filter(group_id=key)
                    group = BoltGroups.objects.get(pk=key)
                    file_name = f"static/general/bolt/{group.name}.xlsx"

                import openpyxl
                workbook = openpyxl.Workbook()
                sheet = workbook.active

                if format == 'excel':
                    sheet['A1'] = "CATEGORY"
                    sheet['B1'] = "BARCODE"
                    sheet['C1'] = "NAME"
                    sheet['D1'] = "PRICE"
                    sheet['E1'] = "SPINTEX"
                    sheet['F1'] = "NIA"
                    sheet['G1'] = "OSU"
                    row = 2

                    for item in items:
                        sheet[f'A{row}'] = f"{item.group.name} / {item.subgroup.name}"
                        sheet[f'B{row}'] = item.product.barcode
                        sheet[f'C{row}'] = item.product.name
                        sheet[f'D{row}'] = item.price
                        sheet[f'E{row}'] = item.stock_spintex
                        sheet[f'F{row}'] = item.stock_nia
                        sheet[f'G{row}'] = item.stock_osu
                        row += 1

                    workbook.save(file_name)
                    success_response['message'] = file_name

            elif module == 'slow_moving_items':
                print(data)
                loc = data.get('location') or ''
                if loc == '*':
                    loc = ""

                act_loc = f"%{loc}%"
                print(act_loc)
                days = int(data.get('days'))
                export = data.get('export')
                quantity = data.get('quantity',1)
                is_active = data.get('stock_type',1)
                operator = data.get('operator','=')
                focus = data.get('focus','moved')
                supplier = data.get('supplier','')

                from datetime import timedelta
                date_from = timezone.now().date() - timedelta(days=days)
                date_to = timezone.now().date()
                if export == 'excel':
                    import openpyxl
                    book = openpyxl.Workbook()
                    sheet = book.active
                    hd = ["ITEMCODE", "BARCODE", "NAME", "MOVE_QTY", "SOLD"]
                    sheet.merge_cells('A1:E1')
                    sheet['A1'] = f"{str(date_from)} - {str(date_to)} ({days} days) for location {loc} above {quantity} units"
                    sheet.append(hd)

                query = f"exec dbo.Sp_slow_moving_rept N'%',N'%',N'%',N'',N'Zade',N'AA001',N'SOO216',N'',N'YWS195','2023-11-13 00:00:00',N'30',N'%',N'%',N'%','2023-08-15 00:00:00','2023-11-13 00:00:00',N'ALL'"
                #print(query)


                conn = ret_cursor()
                cursor = conn.cursor()

                cursor.execute(f"""
                            DECLARE @ls_doc_type varchar(20) = ''
                DECLARE @ldt_dtfrom varchar(20) = '{date_from}'
                DECLARE @ls_loc varchar(20) = '{act_loc}'
                DECLARE @ldt_dtto varchar(20) = '{date_to}'
                DECLARE @ls_group varchar(20) = '%%'
                DECLARE @ls_subgroup varchar(20) = '%%'
                DECLARE @supp_code varchar(20) = '%{supplier}%'
                DECLARE @barcode varchar(20) = '%%'


                SELECT 
                    prod_mast.item_code,
                    prod_mast.item_ref, 
                    prod_mast.barcode, 
                    prod_mast.item_des, 
                    prod_mast.last_net_cost,     
                    prod_mast.packing, 
                    prod_mast.group_code, 
                    prod_mast.sub_group, 
                    prod_mast.family_code, 
                    group_mast.group_des AS group_name,   
                    sub_group.sub_group_des AS sub_group_name, 
                    prod_mast.sub_subgroup,   
                    sub_subgroup.sub_subgroup_des AS sub_subgroup_name,

                    -- Conditional Columns Based on @ls_doc_type
                    ISNULL(
                        CASE 
                            WHEN @ls_doc_type = 'Whole Sales' THEN (
                                SELECT SUM(ABS(qty)) 
                                FROM exp_tran2 
                                WHERE doc_date BETWEEN @ldt_dtfrom AND @ldt_dtto
                                  AND item_code = prod_mast.item_code 
                                  AND doc_type IN ('IN', 'TR', 'AD', 'PR') 
                                  AND loc_id LIKE @ls_loc 
                                  AND tran_io = 'O'
                            )
                            WHEN @ls_doc_type = 'POS Sales' THEN (
                                SELECT SUM(ABS(qty)) 
                                FROM exp_tran2 
                                WHERE doc_date BETWEEN @ldt_dtfrom AND @ldt_dtto
                                  AND item_code = prod_mast.item_code 
                                  AND doc_type IN ('TR', 'PI', 'AD', 'PR') 
                                  AND loc_id LIKE @ls_loc 
                                  AND tran_io = 'O'
                            )
                            ELSE (
                                SELECT SUM(ABS(qty)) 
                                FROM exp_tran2 
                                WHERE doc_date BETWEEN @ldt_dtfrom AND @ldt_dtto
                                  AND item_code = prod_mast.item_code 
                                  AND doc_type IN ('IN', 'TR', 'PI', 'AD', 'PR') 
                                  AND loc_id LIKE @ls_loc 
                                  AND tran_io = 'O'
                            )
                        END, 0) AS move_qty,

                    -- First Receive Date
                    ISNULL((
                        SELECT MIN(doc_date) 
                        FROM exp_tran2 
                        WHERE item_code = prod_mast.item_code 
                          AND doc_type IN ('GR', 'TR', 'OB', 'AD', 'PR') 
                          AND loc_id LIKE @ls_loc 
                          AND tran_io = 'I'
                    ), NULL) AS first_rec_date,

                    -- Latest Receive Date
                    ISNULL((
                        SELECT MAX(doc_date) 
                        FROM exp_tran2 
                        WHERE item_code = prod_mast.item_code 
                          AND doc_type IN ('GR', 'TR', 'OB', 'AD', 'PR') 
                          AND loc_id LIKE @ls_loc 
                          AND tran_io = 'I'
                    ), NULL) AS rec_date,

                    -- Latest Sales Date
                    ISNULL((
                        SELECT MAX(doc_date) 
                        FROM exp_tran2 
                        WHERE item_code = prod_mast.item_code 
                          AND doc_type IN (
                            CASE 
                                WHEN @ls_doc_type = 'Whole Sales' THEN 'IN'
                                WHEN @ls_doc_type = 'POS Sales' THEN 'TR'
                                ELSE 'IN'
                            END, 'TR', 'PI', 'AD', 'PR'
                          ) 
                          AND loc_id LIKE @ls_loc 
                          AND tran_io = 'O'
                    ), NULL) AS sales_date,

                    -- Quantity
                    ISNULL((
                        SELECT SUM(qty) 
                        FROM exp_tran2 
                        WHERE item_code = prod_mast.item_code 
                          AND loc_id LIKE @ls_loc 
                          AND doc_date <= @ldt_dtto
                    ), 0) AS q_qty,

                    -- Retail
                    prod_mast.retail1 AS retail,

                    -- Latest Issue Date
                    ISNULL((
                        SELECT MAX(doc_date) 
                        FROM exp_tran2 
                        WHERE item_code = prod_mast.item_code 
                          AND doc_type IN (
                            CASE 
                                WHEN @ls_doc_type = 'Whole Sales' THEN 'IN'
                                WHEN @ls_doc_type = 'POS Sales' THEN 'TR'
                                ELSE 'IN'
                            END, 'TR', 'PI', 'AD', 'PR'
                          ) 
                          AND loc_id LIKE @ls_loc 
                          AND qty < 0
                    ), NULL) AS iss_date,

                    -- Move Quantity (Current)
                    ISNULL((
                        SELECT SUM(ABS(qty)) 
                        FROM exp_tran2 
                        WHERE item_code = prod_mast.item_code 
                          AND doc_type IN (
                            CASE 
                                WHEN @ls_doc_type = 'Whole Sales' THEN 'IN'
                                WHEN @ls_doc_type = 'POS Sales' THEN 'TR'
                                ELSE 'IN'
                            END, 'TR', 'PI', 'AD', 'PR'
                          ) 
                          AND loc_id LIKE @ls_loc
                    ), 0) AS move_qty_curr,

                    -- S Quantity
                    ISNULL((
                        SELECT SUM(ABS(qty)) 
                        FROM exp_tran2 
                        WHERE doc_date BETWEEN @ldt_dtfrom AND @ldt_dtto 
                          AND item_code = prod_mast.item_code 
                          AND doc_type IN (
                            CASE 
                                WHEN @ls_doc_type = 'Whole Sales' THEN 'IN'
                                WHEN @ls_doc_type = 'POS Sales' THEN 'PI'
                                ELSE 'IN'
                            END, 'PI'
                          ) 
                          AND loc_id LIKE @ls_loc
                    ), 0) AS s_qty

                FROM 
                    prod_mast
                LEFT JOIN group_mast 
                    ON group_mast.group_code = prod_mast.group_code
                LEFT JOIN sub_group 
                    ON sub_group.group_code = prod_mast.group_code 
                   AND sub_group.sub_group = prod_mast.sub_group
                LEFT JOIN sub_subgroup 
                    ON sub_subgroup.group_code = prod_mast.group_code 
                   AND sub_subgroup.sub_group = prod_mast.sub_group 
                   AND sub_subgroup.sub_subgroup = prod_mast.sub_subgroup
                WHERE 
                    (prod_mast.group_code LIKE @ls_group)
                    AND (prod_mast.sub_group LIKE @ls_subgroup)
                    AND (prod_mast.supp_code LIKE @supp_code)
                    AND (prod_mast.barcode LIKE @barcode) 
                    AND (prod_mast.item_type LIKE '%{is_active}%')





                            """)

                # Fetch column names for the query
                column_names = [desc[0] for desc in cursor.description]

                results = cursor.fetchall()
                arr = []
                for row in results:
                    print(row)
                    row_data = dict(zip(column_names, row))  # Pair column names with their values
                    item_code = row_data['item_code']
                    barcode = row_data['barcode'].strip()
                    name = row_data['item_des'].rstrip() if row_data['item_des'] and len(row_data['item_des']) > 0 else \
                    row_data['item_des']
                    packing = row_data['packing']
                    group_name = row_data['group_name'].strip() if row_data['group_name'] and len(
                        row_data['group_name']) > 0 else row_data['group_name']
                    sub_group_name = row_data['sub_group_name'].strip() if row_data['sub_group_name'] and len(
                        row_data['sub_group_name']) > 0 else row_data['sub_group_name']
                    sub_subgroup_name = row_data['sub_subgroup_name'].strip() if row_data['sub_subgroup_name'] and len(
                        row_data['sub_subgroup_name']) > 0 else row_data['sub_subgroup_name']
                    move_qty = row_data['move_qty']
                    first_rec_date = row_data['first_rec_date']
                    rec_date = row_data['rec_date']
                    sales_date = row_data['sales_date']
                    s_qty = row_data['s_qty']

                    # operators = {
                    #     "==": operator.eq,
                    #     "!=": operator.ne,
                    #     "<": operator.lt,
                    #     "<=": operator.le,
                    #     ">": operator.gt,
                    #     ">=": operator.ge,
                    # }

                    ev = f"{quantity} {operator} {move_qty}"
                    if focus == 'sold':
                        ev = f"{quantity} {operator} {s_qty}"

                    print(ev)
                    if eval(ev):
                        li = [item_code, barcode, name, move_qty, s_qty]

                        if export == 'json':
                            arr.append({
                                'barcode':barcode,'itemcode':item_code,'name':name,'moved':move_qty,'sold':s_qty
                            })

                        if export == 'excel':
                            sheet.append(li)

                conn.close()
                if export == 'json':
                    success_response['message'] = arr
                elif export == 'excel':
                    file_name = f"static/general/tmp/SLOW_MOVING_ITEMS.xlsx"
                    book.save(file_name)
                    success_response['message'] = file_name

                response = success_response

            elif module == 'retail_categories':
                import openpyxl
                document = data.get('doc')
                group_id = data.get('group_id') or '*'

                if group_id == '*':
                    groups = ProductGroup.objects.all()
                else:
                    groups = ProductGroup.objects.filter(pk=group_id)

                arr = []

                book = openpyxl.Workbook()
                sheet = book.active
                sheet.title = 'CATEGORIES'
                sheet['A1'] = 'CODE'
                sheet['B1'] = 'NAME'
                sheet['C1'] = 'SUBS'
                sheet_count = 2

                for group in groups:
                    if document == 'json':
                        arr.append({
                            'code': group.code,
                            'name': group.name,
                            'subs': group.subgroups().count()
                        })
                    elif document == 'excel':
                        sheet[f"A{sheet_count}"] = group.code
                        sheet[f"B{sheet_count}"] = group.name
                        sheet[f"C{sheet_count}"] = group.subgroups().count()
                        sheet_count += 1

                if document == 'json':
                    success_response['message'] = arr
                elif document == 'excel':
                    file_name = 'static/general/tmp/categories.xlsx'
                    book.save(file_name)
                    success_response['message'] = file_name

            elif module == 'retail_sub_categories':
                import openpyxl
                document = data.get('doc')
                sub_group_id = data.get('sub_group_id') or '*'
                if sub_group_id == '*':
                    sub_groups = ProductSubGroup.objects.all()
                else:
                    sub_groups = ProductSubGroup.objects.filter(pk=sub_group_id)

                arr = []
                book = openpyxl.Workbook()
                sheet = book.active
                sheet.title = "SUB CATEGORIES"
                sheet['A1'] = "GROUP"
                sheet['B1'] = "CODE"
                sheet['C1'] = "NAME"
                sheet['D1'] = "PRODUCTS"

                sheet_count = 2

                for sub_group in sub_groups:
                    group = sub_group.group
                    code = sub_group.code
                    name = sub_group.name
                    products = sub_group.products().count()

                    if document == 'json':
                        arr.append(
                            {'pk': sub_group.pk, 'group': group.name, 'code': code, 'name': name, 'products': products})
                    elif document == 'excel':
                        sheet[f"A{sheet_count}"] = group.name
                        sheet[f"B{sheet_count}"] = code
                        sheet[f"C{sheet_count}"] = name
                        sheet[f"D{sheet_count}"] = products

                if document == 'json':
                    success_response['message'] = arr
                elif document == 'excel':
                    file_name = 'static/general/tmp/sub_categories.xlsx'
                    book.save(file_name)
                    success_response['message'] = file_name

            elif module == 'retail_products':
                import openpyxl
                product = data.get('product') or '*'
                document = data.get('doc') or 'json'

                if product == '*':
                    products = Products.objects.all()
                else:
                    products = Products.objects.filter(pk=product)

                book = openpyxl.Workbook()
                sheet = book.active
                sheet.title = 'Products'

                sheet['A1'] = 'Group'
                sheet['B1'] = 'Sub Group'
                sheet['C1'] = 'Code'
                sheet['D1'] = 'Barcode'
                sheet['E1'] = 'Name'
                sheet['F1'] = 'Price'
                sheet['G1'] = 'IS_ON_BOLT'
                sheet_count = 2
                arr = []

                for product in products:
                    sub_group = product.subgroup
                    group = sub_group.group.name
                    sub_name = sub_group.name
                    code = product.code
                    barcode = product.barcode.strip()
                    name = product.name
                    price = product.price
                    pk = product.pk
                    bolt = product.is_on_bolt()

                    # check if on bolt

                    if document == 'json':
                        arr.append({
                            'pk': pk,
                            'group': group,
                            'sub_group': sub_name,
                            'code': code,
                            'barcode': barcode,
                            'name': name,
                            'price': price,
                            'is_on_bolt': bolt
                        })
                    elif document == 'excel':
                        sheet['A' + str(sheet_count)] = group
                        sheet['B' + str(sheet_count)] = sub_name
                        sheet['C' + str(sheet_count)] = code
                        sheet['D' + str(sheet_count)] = barcode
                        sheet['E' + str(sheet_count)] = name
                        sheet['F' + str(sheet_count)] = price
                        sheet['G' + str(sheet_count)] = bolt
                        sheet_count += 1

                if document == 'json':
                    success_response['message'] = arr
                elif document == 'excel':
                    file_name = 'static/general/tmp/products.xlsx'
                    book.save(file_name)
                    success_response['message'] = file_name

            elif module == 'recipe_group':
                key = data.get('key', '*')

                if key == '*':
                    groups = RecipeGroup.objects.all()
                else:
                    groups = RecipeGroup.objects.filter(pk=key)

                for group in groups:
                    prod = []
                    products = group.products()
                    p_count = products.count()
                    for product in products:
                        prod.append({
                            'pk': product.pk,
                            'name': product.name,
                            'barcode': product.barcode,
                            'si_unit': product.si_unit,
                            'recipe_items': product.recipe_items(),
                            'image': product.image.url or 'none'
                        })

                    arr.append({
                        'name': group.name,
                        'pk': group.pk,
                        'is_open': group.is_open,
                        'products': {
                            'counts': p_count,
                            'list': prod
                        }
                    })

                success_response['message'] = arr

            elif module == 'recipe_product':
                key = data.get('key', '*')
                search_type = data.get('s_type', 'normal')

                if search_type == 'by_name':
                    products = RecipeProduct.objects.filter(
                        name__icontains=key) if key != '*' else RecipeProduct.objects.all()
                else:
                    products = RecipeProduct.objects.filter(pk=key)

                if products.count() > 0:
                    for product in products:
                        print(product)
                        # #print(product)
                        rec_list = []
                        for r in product.recipe():
                            rec_list.append({
                                'pk': r.pk,
                                'name': r.name,
                                'quantity': r.quantity,
                                'si_unit': r.si_unit
                            })

                        arr.append({
                            'pk': product.pk,
                            'name': product.name,
                            'barcode': product.barcode,
                            'si_unit': product.si_unit,
                            'recipe_items': {
                                'count': product.recipe_items(),
                                'list': rec_list
                            },
                            'is_open': product.is_open,
                            'image': product.img_url(),
                            'next': product.next(),
                            'previous': product.prev()
                        })

                        print(len(arr))
                        if len(arr) < 1:
                            print("NO FISH HERE")
                            raise Exception(f"Item not synced")
                        success_response['message'] = arr
                else:
                    success_response['status_code'] = 404
                    success_response['message'] = "Product Not Found"

            elif module == 'recipe_item':
                key = data.get('key', '*')

                if key == '*':
                    recipe_item = Recipe.objects.all()
                else:
                    recipe_item = Recipe.objects.filter(product_id=key)

                for item in recipe_item:
                    arr.append({
                        'name': item.name,
                        'quantity': item.quantity,
                        'si_unit': item.si_unit
                    })

                success_response['message'] = arr

            elif module == 'export_recipe_group':
                group_pk = data.get('group')
                group = RecipeGroup.objects.get(pk=group_pk)
                file_name = f"static/general/tmp/{group.name}.xlsx"
                products = group.products()
                import openpyxl

                book = openpyxl.Workbook()

                for product in products:

                    name = product.name
                    sheet = book.create_sheet(title=name.replace('/', ''))
                    sheet['A1'] = 'NAME'
                    sheet['B1'] = "UNIT"
                    sheet['C1'] = "QUANTITY"

                    sheet_row = 2
                    l_qty = 0
                    recipes = product.recipe()
                    #print(name)
                    for recipe in recipes:
                        r_name = recipe.name
                        si_unit = recipe.si_unit
                        quantity = recipe.quantity
                        sheet[f"A{sheet_row}"] = r_name
                        sheet[f"B{sheet_row}"] = si_unit
                        sheet[f"C{sheet_row}"] = quantity
                        sheet_row += 1
                        l_qty += Decimal(quantity)
                        #print(r_name)

                    sheet[f"A{sheet_row}"] = 'TOTAL'
                    sheet[f"C{sheet_row}"] = l_qty

                    #print()

                book.save(file_name)
                success_response['message'] = f'/{file_name}'

            elif module == 'stock_monitor':
                # get enabled stocks
                products = Products.objects.filter(stock_monitor=True)
                for product in products:
                    pk = product.pk
                    code = product.code
                    pm = ProdMaster()
                    stk = pm.get_stock(code)

                    nia = stk['nia']
                    osu = stk['osu']
                    spintex = stk['spintex']
                    warehouse = stk['warehouse']
                    vegetables = stk['vegetables']
                    kitchen = stk['kitchen']
                    total = stk['total']
                    sales = stk['sales']

                    # spintex
                    sp = False
                    if spintex >= 0:
                        sp = True

                    if StockMonitor.objects.filter(location='001', product=product).exists():
                        stp = StockMonitor.objects.get(location='001', product=product)
                        stp.stock_qty = spintex
                        stp.valid = sp
                        stp.save()
                    else:
                        StockMonitor(location='001', product=product, stock_qty=spintex, valid=sp).save()

                    ## NIA
                    ni = False
                    if nia >= 0:
                        ni = True

                    if StockMonitor.objects.filter(location='202', product=product).exists():
                        stp = StockMonitor.objects.get(location='202', product=product)
                        stp.stock_qty = nia
                        stp.valid = ni
                        stp.save()
                    else:
                        StockMonitor(location='202', product=product, stock_qty=nia, valid=ni).save()

                    os = False
                    if osu >= 0:
                        os = True

                    if StockMonitor.objects.filter(location='205', product=product).exists():
                        stp = StockMonitor.objects.get(location='205', product=product)
                        stp.stock_qty = os
                        stp.valid = sp
                        stp.save()
                    else:
                        StockMonitor(location='205', product=product, stock_qty=osu, valid=os).save()

                    # kitchen
                    kt = False
                    if kitchen >= 0:
                        kt = True

                    if StockMonitor.objects.filter(location='201', product=product).exists():
                        stp = StockMonitor.objects.get(location='201', product=product)
                        stp.stock_qty = kitchen
                        stp.valid = kt
                        stp.save()
                    else:
                        StockMonitor(location='201', product=product, stock_qty=kitchen, valid=kt).save()

                    # warehouse
                    wh = False
                    if warehouse >= 0:
                        wh = True

                    if StockMonitor.objects.filter(location='999', product=product).exists():
                        stp = StockMonitor.objects.get(location='999', product=product)
                        stp.stock_qty = warehouse
                        stp.valid = wh
                        stp.save()
                    else:
                        StockMonitor(location='999', product=product, stock_qty=warehouse, valid=wh).save()

            elif module == 'see_stock_monitor':
                arr = []
                print(data)
                filter = data.get('filter')
                loc_code = data.get('loc_code','*')
                #print(filter)
                doc = data.get('doc', 'json')
                if doc == 'excel':
                    import openpyxl
                    book = openpyxl.Workbook()
                    sheet = book.active
                    sheet.title = "Stock Check"
                    sheet.append(['LOCATION', 'BARCODE', 'NAME', 'STOCK'])
                # if filter == 'not_accurate':
                #     all_m = StockMonitor.objects.filter(valid=False)
                # elif filter == 'accurate':
                #     all_m = StockMonitor.objects.filter(valid=True)
                # else:
                #     all_m = StockMonitor.objects.all()
                all_m = Products.objects.all()
                locations = Locations.objects.filter(type='retail')
                if loc_code != '*':
                    locations = locations.filter(code=loc_code)
                for m in all_m:
                    for location in locations:
                        code = location.code
                        loc_stock = \
                        RawStock.objects.filter(prod_id=m.code, loc_id=location.code).aggregate(sum=Sum('qty'))[
                            'sum'] or Decimal('0.00')
                        row = [location.descr, m.name, m.name, loc_stock]
                        mark = 0
                        if loc_stock >= mark:
                            pass
                            #print(f"{loc_stock} is positive for ",filter,location.code,m.code)
                        elif loc_stock < mark:
                            pass
                            #print(mark, "is negative for",filter,location.code,m.code)
                        else:
                            pass
                            #print(mark, "Is neither negative or positive",location.code,m.code)

                        if filter == 'negative' and loc_stock < mark:
                            #print("YES NEGATIVE")

                            if doc == 'excel':
                                sheet.append(row)
                            arr.append({'loc_code': code, 'loc_name': location.descr, 'barcode': m.barcode,
                                        'item_name': m.name,
                                        'stock': loc_stock})

                        elif filter == 'positive' and loc_stock > mark:
                            arr.append({'loc_code': code, 'loc_name': location.descr, 'barcode': m.barcode,
                                        'item_name': m.name,
                                        'stock': loc_stock})
                            if doc == 'excel':
                                sheet.append(row)
                                #print("YES POSITIVE")

                        elif filter == 'neutral' and loc_stock == mark:
                            arr.append({'loc_code': code, 'loc_name': location.descr, 'barcode': m.barcode,
                                        'item_name': m.name,
                                        'stock': loc_stock})
                            if doc == 'excel':
                                sheet.append(row)


                    # arr.append(m.stock())

                if doc == 'excel':
                    filename = 'static/general/tmp/stock.xlsx'
                    book.save(filename)
                    arr = filename
                success_response['message'] = arr
                response = success_response

            elif module == 'mr_check':
                mr_no = data.get('mr_no')
                doc_type = data.get('doc', "JSON")
                healthy = 0
                not_healthy = 0

                conn = ret_cursor()
                cursor = conn.cursor()
                cursor.execute(
                    f"SELECT entry_no, entry_date, remark, (SELECT RTRIM(user_name) FROM user_file WHERE user_id = "
                    f"request_hd.user_id) AS 'requested_by', (SELECT RTRIM(br_name) FROM branch WHERE br_code = request_hd.loc_id) "
                    f"AS 'loc_requested', (SELECT RTRIM(br_name) FROM branch WHERE br_code = request_hd.loc_from) "
                    f"AS 'loc_to_send',loc_id FROM request_hd where entry_no = '{mr_no}';")
                mr_hd = cursor.fetchone()
                if mr_hd is not None:
                    if doc_type == 'excel':
                        import openpyxl
                        book = openpyxl.Workbook()
                        sheet = book.active
                        sheet.title = "RECORDS"

                        header = ["BARCODE","NAME","REQ QTY","LAST TREANSFER","LAST TRANSFER DATE","LAST TRANSF QTY","ADJUSTED","SOLD","OUT","OUT PERCENTAGE","HEALTH"]
                        sheet.append(header)

                        # sheet[f'A1'] = "BARCODE"
                        # sheet[f"B1"] = "NAME"
                        # sheet[f"C1"] = "REQ. QTY"
                        # sheet[f'D1'] = "LAST TRAN. ENTRY"
                        # sheet[f"E1"] = "LAST TRAN. DATE"
                        # sheet[f'F1'] = "LAST TRAN. QTY"
                        # sheet[f"G1"] = "SOLD BETWEEN"
                        # sheet[f"H1"] = "SOLD PERENTAGE"
                        # sheet[f'I1'] = "HEALTH"

                        sheet_row = 2

                    arr = []
                    entry_no, entry_date, remark, requested_by, loc_requested, loc_to_send, loc_id = mr_hd
                    header = {
                        "entry_no": entry_no.strip(),
                        "entry_date": entry_date,
                        'remark': remark.strip() if remark is not None else None,
                        'requested_by': requested_by,
                        'loc_requested': loc_requested,
                        'loc_to_send': loc_to_send,
                        'loc_id': loc_id
                    }

                    # connect to branch and get sold
                    loc = Locations.objects.get(code=loc_id)
                    pos_cursor = ret_cursor(loc.ip_address, '', loc.db, loc.db_user, loc.db_password).cursor()
                    inv_cursor = ret_cursor('192.168.2.4', '', 'SMSEXPV17', 'sa', 'sa@123456').cursor()
                    # get transactions
                    trans_query = cursor.execute(
                        f"select item_code,RTRIM(item_des) as 'name',total_units,barcode from request_tran where entry_no = '{mr_no}'")
                    tr_arr = []
                    for tran in trans_query.fetchall():
                        item_code, item_name, request_qty,barcode = tran
                        cursor.execute(
                            f"select top(1) th.entry_no,th.entry_date,tr.total_units from tran_tr tr  right join tran_hd th on th.entry_no = tr.entry_no where tr.item_code = '{item_code}' and th.loc_to='{loc_id}' order by th.entry_date desc")
                        last_tran_row = cursor.fetchone()

                        last_tr_entry, last_tran_date, last_tran_qty, sold_qty = ['none', 'none', 0, 0]
                        total_out = 0
                        adjusted = 0
                        transfered = 0
                        if last_tran_row is not None:
                            last_tr_entry, last_tran_date, last_tran_qty = last_tran_row

                            pos_cursor.execute(
                                f"SELECT sum(tran_qty) as 'sold_qty' FROM history_tran where bill_date between '{last_tran_date}' and '{entry_date}' and prod_id = '{item_code}'")
                            sold_resp = pos_cursor.fetchone()
                            if sold_resp is not None:
                                if sold_resp[0] is not None:
                                    sold_qty = sold_resp[0]

                                # adjusted
                                inv_cursor.execute(
                                    f"SELECT SUM(total_units) FROM adj_tran where item_code = '{item_code}'and  entry_no in (SELECT entry_no from adj_hd where entry_date between '{last_tran_date}' and '{entry_date}' and loc_id = '{loc_id}' )")
                                adjusted = inv_cursor.fetchone()[0] or 0


                                # transferd
                                inv_cursor.execute(
                                    f"SELECT SUM(total_units) FROM tran_tr where item_code = '{item_code}'  and entry_no in (SELECT entry_no from tran_hd where entry_date between '{last_tran_date}' and '{entry_date}' and loc_from = '{loc_requested}' )")
                                transfered = inv_cursor.fetchone()[0] or 0

                                total_out = sold_qty + adjusted

                        health = False
                        if total_out > Decimal(0.5) * last_tran_qty:
                            health = True
                            healthy += 1
                        else:
                            not_healthy += 1

                        percentage_sold = 0
                        if last_tran_qty != 0 and sold_qty != 0:  # Ensure we don't divide by zero
                            percentage_sold = (Decimal(sold_qty) / Decimal(last_tran_qty)) * 100
                        if doc_type == 'JSON':
                            tr_arr.append({
                                'item_code': tran[0],
                                'name': tran[1],
                                'request_qty': tran[2],
                                'last_tran_entry': last_tr_entry,
                                'last_tran_date': last_tran_date,
                                "last_tran_qty": last_tran_qty,
                                'sold_qty': sold_qty,
                                "percentage_sold": percentage_sold,
                                'health': health
                            })
                        elif doc_type == 'excel':
                            header = [barcode, item_name,  tran[2], last_tr_entry, last_tran_date,
                                      last_tran_qty, adjusted, sold_qty, total_out, percentage_sold,
                                      health]
                            sheet.append(header)
                            # sheet[f'A{sheet_row}'] = barcode
                            # sheet[f"B{sheet_row}"] = item_name
                            # sheet[f"C{sheet_row}"] = tran[2]
                            # sheet[f'D{sheet_row}'] = last_tr_entry
                            # sheet[f"E{sheet_row}"] = last_tran_date
                            # sheet[f'F{sheet_row}'] = last_tran_qty
                            # sheet[f"G{sheet_row}"] = total_out
                            # sheet[f"H{sheet_row}"] = percentage_sold
                            # sheet[f'I{sheet_row}'] = health

                            sheet_row += 1

                    if doc_type == "JSON":
                        resp = {
                            'header': header,
                            'transactions': tr_arr
                        }

                        arr.append(resp)
                    elif doc_type == 'excel':
                        # Write "yes" and "no" to cells
                        sheet['K1'] = "HEALTHY"
                        sheet['L1'] = "UNHEALTHY"
                        sheet['K2'] = healthy  # Example value for "yes"
                        sheet['L2'] = not_healthy  # Example value for "no"

                        # Create a pie chart
                        chart = PieChart()
                        data = Reference(sheet, min_col=11, min_row=2, max_col=12,
                                         max_row=2)  # Range of data values (yes and no counts)
                        labels = Reference(sheet, min_col=11, min_row=1, max_col=12,
                                           max_row=1)  # Range of category labels (yes and no)
                        chart.add_data(data, titles_from_data=True)
                        # chart.title = "STOCK HEALTH"
                        chart.set_categories(labels)

                        # Set the title of the chart
                        chart.title = "Yes vs No"

                        # Add the chart to the worksheet
                        sheet.add_chart(chart, "K5")

                        # Set the dimensions of the chart
                        chart.width = 15  # Adjust width
                        chart.height = 10  # Adjust height

                        file_name = f"static/general/tmp/mr_check{mr_no}.xlsx"
                        book.save(file_name)
                        resp = file_name
                    else:
                        resp = f"Unknown Document To Render {doc_type}"
                    success_response['message'] = resp
                    response = success_response
                else:
                    success_response['message'] = f"no request with entry {mr_no}"
                    success_response['status_code'] = 404
                    response = success_response

                conn.close()

            elif module == 'tr_check':
                entry_no = data.get('entry_no')
                export = data.get('doc')
                if export == 'excel':
                    import openpyxl
                    book = openpyxl.Workbook()

                    sheet = book.active
                    header = ['BARCODE','ITEM CODE','QUANTITY','NAME','PACKING','PACK QUANTITY','TOTAL UNITS','GRN PACK QUANTITY',"HEALTH"]
                    sheet.append(header)
                conn = ret_cursor()
                cursor = conn.cursor()
                q = f"""
                    SELECT
                    item_ref as 'BARCODE',
                    item_code as 'ITEM CODE',
                    unit_qty as 'QUANTITY',
                    LTRIM(RTRIM(item_des)) as 'NAME',
                    packing as 'PACKING',
                    um_qty as 'PACK QUANTITY',
                    total_units as 'TOTAL SENT',
                    CASE
                        WHEN (SELECT count(*) from grn_tran gtr where gtr.item_code = trr.item_code) > 0 then 
                        (SELECT top(1) um_qty from grn_tran gtrr join grn_hd grh on grh.entry_no= gtrr.entry_no  where gtrr.item_code = trr.item_code order by grh.grn_date desc  )
                        ELSE 0
                    END as 'GRN_PACK'
                
                from tran_tr trr
                where entry_no = '{entry_no}'
                
                """
                cursor.execute(q)
                arr = []
                for row in cursor.fetchall():

                    barcode, item_code, quantity, name, packing, um_qty, total_units, grn_um = row
                    if um_qty != grn_um:
                        if grn_um != 0:
                            health = False
                        else:
                            health = True
                    else:
                        health = True

                    li = [barcode, item_code, quantity, name, packing, um_qty, total_units, grn_um,health]

                    if export == 'excel':
                        sheet.append(li)

                    if export == 'json':
                        arr.append({
                            'barcode':barcode,"item_code":item_code, "quantity":quantity,
                            'name':name,'packing':packing,"um_qty":um_qty,"total_units":total_units,'grn_um':grn_um,'health':health
                        })

                if export == 'excel':
                    filename = 'static/general/tmp/stock.xlsx'
                    book.save(filename)
                    arr = filename

                conn.close()
                success_response['message'] = arr
                response = success_response

            elif module == 'documents':
                start_date = data.get('start_date')
                end_date = data.get('end_date')
                #print(start_date,end_date)
                query = f"""
                    DECLARE @start_date varchar(20) = '{start_date}'
                    DECLARE @end_date varchar(20) = '{end_date}'
                    
                    SELECT 'GRN' AS document, (
                        SELECT 
                            (SELECT COUNT(*) FROM grn_hd where grn_date between @start_date and @end_date) AS total_entries,
                            (SELECT COUNT(*) FROM grn_hd WHERE grn_date between @start_date and @end_date and valid = 1 AND posted = 1) AS posted,
                            (SELECT COUNT(*) FROM grn_hd WHERE grn_date between @start_date and @end_date and valid = 1 AND posted = 0) AS not_posted,
                            (SELECT COUNT(*) FROM grn_hd WHERE grn_date between @start_date and @end_date and valid = 0) AS deleted,
                            (SELECT SUM(inv_amt) FROM grn_hd where grn_date between @start_date and @end_date) AS total_value,
                            (SELECT SUM(inv_amt) FROM grn_hd where grn_date between @start_date and @end_date and valid = 1 and posted = 1) AS posted_value,
                            (SELECT SUM(inv_amt) FROM grn_hd where grn_date between @start_date and @end_date and valid = 1 and posted = 0) AS pending_value,
                            (SELECT SUM(inv_amt) FROM grn_hd where grn_date between @start_date and @end_date and valid = 0) AS deleted_value
                        FOR JSON PATH, WITHOUT_ARRAY_WRAPPER
                    ) AS Data
                    
                    
                    
                    -- INVOICE
                    UNION ALL
                    SELECT 'BACK OFFICE SALES' AS document, (
                        SELECT 
                            (SELECT COUNT(*) FROM inv_hd where entry_date between @start_date and @end_date) AS total_entries,
                            (SELECT COUNT(*) FROM inv_hd where entry_date between @start_date and @end_date and valid = 1 AND posted = 1) AS posted,
                            (SELECT COUNT(*) FROM inv_hd where entry_date between @start_date and @end_date and valid = 1 AND posted = 0) AS not_posted,
                            (SELECT COUNT(*) FROM inv_hd where entry_date between @start_date and @end_date and valid = 0) AS deleted,
                            (SELECT SUM(inv_amt) FROM inv_hd where entry_date between @start_date and @end_date) AS total_value,
                            (SELECT SUM(inv_amt) FROM inv_hd where entry_date between @start_date and @end_date and valid = 1 and posted = 1) AS posted_value,
                            (SELECT SUM(inv_amt) FROM inv_hd where entry_date between @start_date and @end_date and valid = 1 and posted = 0) AS pending_value,
                            (SELECT SUM(inv_amt) FROM inv_hd where entry_date between @start_date and @end_date and valid = 0) AS deleted_value
                        FOR JSON PATH, WITHOUT_ARRAY_WRAPPER
                    ) AS Data
                    
                    -- pos sales
                    UNION ALL
                    SELECT 'POS SALES' AS document, (
                        SELECT 
                            (SELECT COUNT(*) FROM pos_tran_hd where entry_date between @start_date and @end_date) AS total_entries,
                            (SELECT COUNT(*) FROM pos_tran_hd where entry_date between @start_date and @end_date and inv_valid = 1 AND inv_upd = 1) AS posted,
                            (SELECT COUNT(*) FROM pos_tran_hd where entry_date between @start_date and @end_date and inv_valid = 1 AND inv_upd = 0) AS not_posted,
                            (SELECT COUNT(*) FROM pos_tran_hd where entry_date between @start_date and @end_date and inv_valid = 0) AS deleted,
                            (SELECT SUM(inv_amt) FROM pos_tran_hd where entry_date between @start_date and @end_date) AS total_value,
                            (SELECT SUM(inv_amt) FROM pos_tran_hd where entry_date between @start_date and @end_date and inv_valid = 1 and inv_upd = 1) AS posted_value,
                            (SELECT SUM(inv_amt) FROM pos_tran_hd where entry_date between @start_date and @end_date and inv_valid = 1 and inv_upd = 0) AS pending_value,
                            (SELECT SUM(inv_amt) FROM pos_tran_hd where entry_date between @start_date and @end_date and inv_valid = 0) AS deleted_value
                        FOR JSON PATH, WITHOUT_ARRAY_WRAPPER
                    ) AS Data
                    
                    -- ADJUSTMENT
                    UNION ALL
                    SELECT 'ADJUSTMENT' AS document, (
                        SELECT 
                            (SELECT COUNT(*) FROM adj_hd where entry_date between @start_date and @end_date) AS total_entries,
                            (SELECT COUNT(*) FROM adj_hd where entry_date between @start_date and @end_date and valid = 1 AND posted = 1) AS posted,
                            (SELECT COUNT(*) FROM adj_hd where entry_date between @start_date and @end_date and valid = 1 AND posted = 0) AS not_posted,
                            (SELECT COUNT(*) FROM adj_hd where entry_date between @start_date and @end_date and valid = 0) AS deleted,
                            (SELECT SUM(tot_retail) FROM adj_hd where entry_date between @start_date and @end_date) AS total_value,
                            (SELECT SUM(tot_retail) FROM adj_hd where entry_date between @start_date and @end_date and valid = 1 and posted = 1) AS posted_value,
                            (SELECT SUM(tot_retail) FROM adj_hd where entry_date between @start_date and @end_date and valid = 1 and posted = 0) AS pending_value,
                            (SELECT SUM(tot_retail) FROM adj_hd where entry_date between @start_date and @end_date and valid = 0) AS deleted_value
                        FOR JSON PATH, WITHOUT_ARRAY_WRAPPER
                    ) AS Data

                    

                """
                arr = []
                conn = ret_cursor()
                cursor = conn.cursor()
                cursor.execute(query)
                for document in cursor.fetchall():
                    ##print(document)
                    doc,daita = document
                    ##print(daita)
                    obj = {}
                    json_data = json.loads(daita)
                    obj['document'] = doc
                    obj['total_entries'] = json_data["total_entries"]
                    obj['posted'] = json_data["posted"]
                    obj['not_posted'] = json_data["not_posted"]
                    obj['deleted'] = json_data["deleted"]
                    obj['total_value'] = format_currency(json_data.get('total_value'))
                    obj['posted_value'] = format_currency(json_data.get('posted_value'))
                    obj['pending_value'] = format_currency(json_data.get("pending_value"))


                    arr.append(obj)
                    # #print(doc)
                    # #print(json_data)
                    ##print()

                conn.close()
                success_response['message'] = arr
                response = success_response

            elif module == 'bolt_graph_week':
                locations = Locations.objects.filter(type='retail')
                from datetime import timedelta
                # Calculate the start date (7 days ago)
                from django.utils.timezone import now
                start_date = now().date() - timedelta(days=7)
                end_date = now().date()
                date_range = [start_date + timedelta(days=x) for x in range(0, (end_date - start_date).days + 1)]
                # date_range.append('Total')
                labels = []
                sales_data = {}

                total_arr = []
                total_sales = 0


                for dr in date_range:
                    labels.append(dr)
                    total_sales = BillHeader.objects.filter(bill_date=dr, pay_mode='BOLT').aggregate(Sum('bill_amt'))['bill_amt__sum'] or 0
                    total_arr.append(total_sales)


                for location in locations:

                    sales_arr = []

                    for d in labels:

                        sales = BillHeader.objects.filter(loc=location, bill_date=d, pay_mode='BOLT').aggregate(
                                Sum('bill_amt'))
                        sales_arr.append(sales['bill_amt__sum'] or 0)


                    sales_data.update({location.descr:sales_arr})






                # sales_data.update({"Total":total_arr})
                arr = {
                    "labels": labels,
                    "dataset": sales_data,
                    "totalset":{"Total Sales":total_arr}
                }

                print(arr)


                success_response['message'] = arr
                response = success_response


            elif module == 'revenue':
                target_date = data.get('date')
                entities = BusinessEntityTypes.objects.all()
                revenue = Decimal(0.00)
                for entity in entities:
                    arr.append(entity.obj())
                    revenue += entity.obj()['revenue']

                success_response['message'] = {
                    'revenue':format_currency(revenue),
                    'breakdown':arr
                }
                response = success_response

            elif module == 'sales_graph_week':
                from django.utils.timezone import now
                loc_q = "SELECT br_code,RTRIM(br_name) from branch"
                conn = ret_cursor()
                cursor = conn.cursor()
                cursor.execute(loc_q)
                from datetime import timedelta
                # Calculate the start date (7 days ago)
                start_date = now().date() - timedelta(days=7)
                end_date = now().date()

                date_range = [start_date + timedelta(days=x) for x in range(0, (end_date - start_date).days + 1)]
                labels = []
                sales_data = {}
                for dr in date_range:
                    labels.append(dr)

                for location in cursor.fetchall():
                    name = location[1]
                    code = location[0]
                    # locations data set

                    # get sales
                    sales_arr = []
                    for d in labels:

                        s_q = f"SELECT inv_amt FROM pos_tran_hd where location_id = '{code}' and entry_date = '{d}'"
                        #print(s_q)
                        cursor.execute(s_q)
                        row = cursor.fetchone()

                        # dheck if d is today yyyy-mm-dd
                        if d == now().date():
                            sales = BillHeader.objects.filter(loc__code=code,bill_date=d).aggregate(Sum('bill_amt'))['bill_amt__sum'] or 0
                        else:

                            if row is None:
                                sales = 0.00
                            else:
                                sales = row[0]
                        sales_arr.append(sales)

                    obj = {
                        "branch": name,
                        "sales": sales_arr
                    }

                    sales_data.update({name: sales_arr})
                arr = {
                    "labels": labels,
                    "dataset": sales_data
                }

                conn.close()
                success_response['message'] = arr
                response = success_response

            elif module == 'prod_master':
                item_code = data.get('item_code','')
                arr = []
                # query = f"""
                #
                #     SELECT
                #     item_code,
                #     group_code,(SELECT RTRIM(LTRIM(group_des)) from group_mast gm where gm.group_code = pm.group_code)as 'group_name',
                #     sub_group,(SELECT RTRIM(LTRIM(sub_group_des)) from sub_group sg where sg.group_code = pm.group_code and sub_group = pm.sub_group) as 'sub_group_name',
                #     sub_subgroup,(SELECT sub_subgroup_des FROM sub_subgroup ssg where ssg.group_code = pm.group_code and ssg.sub_subgroup = pm.sub_subgroup and ssg.sub_group = pm.sub_group) as 'sub_sub_name',
                #     RTRIM(barcode) as 'barcode',
                #     RTRIM(item_des) as 'name',
                #     retail1,
                #     (select SUM(tran_qty) from pos_tran where prod_id = pm.item_code and entry_no in (SELECT entry_no from pos_tran_hd where entry_date >= '2024-01-01')) as 'sold'
                #      from prod_mast pm where pm.item_code like '%{item_code}%' or barcode like '%{item_code}%' and is_gp_chg = 0 and item_type != 0 order by pm.item_des
                #
                # """

                query = f"""
                    
                    WITH SoldData AS (
    SELECT
        prod_id,
        SUM(tran_qty) AS sold
    FROM
        pos_tran pt
    JOIN
        pos_tran_hd pth ON pt.entry_no = pth.entry_no
    WHERE
        pth.entry_date >= '2024-01-01'
    GROUP BY
        prod_id
)
SELECT 
    pm.item_code,
    pm.group_code,
    RTRIM(LTRIM(gm.group_des)) AS group_name,
    pm.sub_group,
    RTRIM(LTRIM(sg.sub_group_des)) AS sub_group_name,
    pm.sub_subgroup,
    RTRIM(ssg.sub_subgroup_des) AS sub_sub_name,
    RTRIM(pm.barcode) AS barcode,
    RTRIM(pm.item_des) AS name,
    pm.retail1,
    COALESCE(sd.sold, 0) AS sold
FROM 
    prod_mast pm
LEFT JOIN 
    group_mast gm ON gm.group_code = pm.group_code
LEFT JOIN 
    sub_group sg ON sg.group_code = pm.group_code AND sg.sub_group = pm.sub_group
LEFT JOIN 
    sub_subgroup ssg ON ssg.group_code = pm.group_code AND ssg.sub_group = pm.sub_group AND ssg.sub_subgroup = pm.sub_subgroup
LEFT JOIN 
    SoldData sd ON sd.prod_id = pm.item_code
WHERE 
    (pm.item_code LIKE '%{item_code}%' OR pm.barcode LIKE '%{item_code}%') 
    AND pm.is_gp_chg = 0 
    AND pm.item_type != 0
ORDER BY 
    pm.item_des;

            
                """

                conn = ret_cursor()
                cursor = conn.cursor()
                cursor.execute(query)
                for item in cursor.fetchall():
                    item_code,group_code,group_name,sub_group_code,sub_group_name,sub_subsub_group_code,sub_subsub_group_name,barcode,name,retail1,sold = item
                    if sold == 'null':
                        sold = 0
                    obj = {
                        "item_code":item_code,
                        "group_code":group_code,
                        "group_name":group_name,
                        "name":name,
                        "sub_group_code":sub_group_code,
                        "sub_group_name":sub_group_name,
                        "sub_subsub_group_code":sub_subsub_group_code,
                        "sub_subsub_group_name":sub_subsub_group_name,
                        "barcode":barcode,
                        "retail1":retail1,
                        "is_on_bolt":True if BoltItems.objects.filter(product__barcode=barcode).exists() else False,
                        'sold':sold

                    }

                    arr.append(obj)
                    #print(item_code,group_code,group_name,sub_group_code,sub_group_name,sub_subsub_group_code,sub_subsub_group_name,barcode,name,retail1)

                conn.close()
                success_response['message'] = arr
                response = success_response


            elif module == 'prod':
                pk = data.get('pk','*')
                filter = data.get('filter','pk')
                entity_pk = data.get('entity')
                ini = data.get('ini','ini')

                # print(data)
                entity = BusinessEntityTypes.objects.get(pk=entity_pk)
                # if ini == 'ini':
                #     pk = Products.objects.filter(entity=entity).last().pk

                if pk == '*':
                    prods = Products.objects.filter(entity=entity).order_by('-pk')[:1]
                else:
                    prods = Products.objects.filter(pk=pk,entity=entity) if filter == 'pk' else Products.objects.filter(barcode=pk,entity=entity)

                for prod in prods:
                    print(prod.obj())
                    arr.append(prod.obj())
                if len(arr) > 0:
                    success_response['message'] = arr
                else:
                    raise Exception("Item Not Synced")


            elif module == 'group_master':
                query = "select RTRIM(group_code),RTRIM(LTRIM(group_des)) as 'group_name' from group_mast order by group_name"
                conn = ret_cursor()
                cursor = conn.cursor()
                cursor.execute(query)
                for group in cursor.fetchall():
                    group_code,group_name = group
                    arr.append({
                        'code':group_code,
                        'name':group_name
                    })

                conn.close()
                success_response['message'] = arr
                response = success_response

            elif module == 'sub_group_master':
                group = data.get('group')
                query = f"select RTRIM(sub_group),RTRIM(LTRIM(sub_group_des)) as 'sub_group_des' from sub_group where group_code = {group} order by sub_group_des"
                conn = ret_cursor()
                cursor = conn.cursor()
                cursor.execute(query)
                for group in cursor.fetchall():
                    group_code, group_name = group
                    arr.append({
                        'code': group_code,
                        'name': group_name
                    })

                conn.close()
                success_response['message'] = arr
                response = success_response

            elif module == 'sub_subgroup_master':
                group = data.get('group')
                sub_group_id = data.get('sub_group_id')
                query = f"select RTRIM(sub_subgroup),RTRIM(LTRIM(sub_subgroup_des)) as 'sub_subgroup_des' from sub_subgroup where group_code = '{group}' and sub_group = '{sub_group_id}' order by sub_subgroup_des"
                conn = ret_cursor()
                cursor = conn.cursor()
                cursor.execute(query)
                for group in cursor.fetchall():
                    sub_subgroup, sub_subgroup_des = group
                    arr.append({
                        'code': sub_subgroup,
                        'name': sub_subgroup_des
                    })

                conn.close()
                success_response['message'] = arr
                response = success_response

            elif module == 'detailed_stock':
                loc_id = data.get('loc_id','*')
                status = data.get('status','*')
                category = data.get('category','*')
                doc_type = data.get('export','json')

                print(data)

                if doc_type == 'excel':
                    import openpyxl
                    book = openpyxl.Workbook()
                    sheet = book.active
                    sheet.title = "STOCK"

                    sheet.append(['BARCODE',"NAME","STOCK"])

                prod_query = f"SELECT item_code,RTRIM(barcode),RTRIM(item_des) from prod_mast prod_mast where group_code = '{category}'"
                if category != '*':
                    prod_query = f"SELECT item_code,RTRIM(barcode),RTRIM(item_des) from prod_mast where group_code = '{category}'"

                print(prod_query)

                conn = ret_cursor()
                cursor = conn.cursor()
                arr = []
                cursor.execute(prod_query)
                for item in cursor.fetchall():
                    item_code,barcode,item_name = item



                    # get stock
                    stock = get_stock(item_code)
                    if loc_id == '*':
                        stock = stock.get('total')
                    else:
                        stock = stock.get(loc_id)

                    li = None
                    if status == '*':
                        li = [barcode,item_name,stock]
                    if status == '-':
                        if stock < 0:
                            li = [barcode,item_name,stock]

                    if status == '+':
                        if stock > 0:
                            li = [barcode,item_name,stock]

                    if doc_type == 'json':
                        if li is not None:
                            arr.append(li)

                    if doc_type == 'excel':
                        if li is not None:
                            sheet.append(li)
                    print(li)


                conn.close()
                if doc_type == 'excel':

                    file = 'static/general/tmp/retail_stock.xlsx'

                    book.save(file)
                    arr = file

                success_response['message'] = arr
                success_response['statuc_code'] = 200

                response = success_response

            elif module == 'update_bolt_price':
                products = BoltItems.objects.all()

                arr = []
                for product in products:
                    price = product.price
                    inv_price = product.product.price

                    # if price !=  inv_price:
                    if True:
                        product.price = inv_price
                        new_price = inv_price + inv_price * Decimal(0.30)
                        print(new_price)
                        arr.append({
                            "sku":product.product.barcode,
                            "base_selling_price":format_currency(new_price),
                            "price_measure_unit":'piece'
                        })

                legend_payload = {
                    "region_id": 0,
                    "provider_ids": ["string"],
                    "price_list": {
                        "currency": "GHS",
                        "skus":arr
                    }
                }

                success_response['message'] = legend_payload
                response = success_response

            elif module == 'update_bolt_stock':
                products = BoltItems.objects.all()
                id = data.get('loc_id')
                arr = []
                for product in products:
                    arr.append({
                        "sku":product.product.barcode,
                        "quantity":product.stock().get(id),
                        'selling_unit':"pieces"

                    })

                legend_payload = {
                    "provider_ids": ["string"],
                    "sku_quantities": arr
                }

                success_response['message'] = legend_payload
                response = success_response

            elif module == 'send2bolt':
                items = BoltItems.objects.filter(is_sync=False).order_by('product__name')
                count = items.count()
                countx = count
                import csv

                out = []
                # out.append(["SKU","NAME","GROUP","PRICE"])
                # for item in items:
                #     # print(f"{str(count) }"f"/{ str(countx)}")
                #     li = [item.product.barcode,item.product.name,f"{item.group.name}/{item.subgroup.name}",item.product.price]
                #     print(li)
                #     out.append(li)
                #
                #     count -= 1

                providers_id = ""
                for bolt_id in BOLT_PROVIDER_ID:
                    print(bolt_id)
                    row = BOLT_PROVIDER_ID[bolt_id]
                    providers_id += f"{row.get('id')},"

                # remove last comma
                providers_id = providers_id[:-1]
                print(providers_id)

                import os
                import zipfile
                import shutil

                time_ext = str(timezone.now()).replace(':', '_').replace(' ', '_')
                folder = f"static/general/tmp/{time_ext}"
                # Create the folder if it doesn't exist
                if not os.path.exists(folder):
                    os.makedirs(folder)


                csv_file = f"static/general/tmp/bot_{time_ext}_file.csv"
                prices_file = f"static/general/tmp/bot_{time_ext}_prices.csv"
                spintex_stock_file = f"static/general/tmp/bot_{time_ext}_spintex_stock.csv"
                nia_stock_file = f"static/general/tmp/bot_{time_ext}_nia_stock.csv"
                osu_stock_file = f"static/general/tmp/bot_{time_ext}_osu_stock.csv"

                with open(csv_file,mode='w',newline='', encoding="utf-8") as px,\
                        open(prices_file,mode='w',newline='', encoding="utf-8") as prices,\
                    open(spintex_stock_file, mode='w', newline='', encoding="utf-8") as spintex_stock,\
                        open(nia_stock_file, mode='w', newline='', encoding="utf-8") as nia_stock,\
                        open(osu_stock_file, mode='w', newline='', encoding="utf-8") as osu_stock:
                    writer = csv.writer(px)
                    pw = csv.writer(prices)
                    spintex = csv.writer(spintex_stock)
                    nia = csv.writer(nia_stock)
                    osu = csv.writer(osu_stock)



                    print("Writing to file")
                    # Controlled loop to write each row
                    header = [
                        "SKU",
                        "Barcode",
                        "Name en-GH",
                        "Name en-US",
                        "Level 1 - Category",
                        "Level 2 - Subcategory",
                        "Provider IDs",
                        "Sub-category code",
                        "Description en-US",
                        "Description en-GH"
                    ]
                    writer.writerow(header)
                    pw.writerow(["SKU","Selling price","Provier IDs"])
                    stock_header = ['sku','quantity','address']
                    nia.writerow(stock_header)
                    osu.writerow(stock_header)
                    spintex.writerow(stock_header)


                    for item in items:
                        barcodes = item.product.barcodes if item.product.barcodes else item.product.barcode
                        li = [
                            str(item.product.barcode),
                            str(barcodes),
                            item.product.name,
                            item.product.name,
                            item.group.name,
                            item.subgroup.name,
                            str(providers_id),
                            f"{item.group.name}/{item.subgroup.name}",
                            item.description,
                            item.description
                        ]

                        # stock = get_stock(item.product.code)
                        stock = stock_by_moved(item.product.pk,'*')
                        print(stock)
                        sp_stk = [
                            item.product.barcode,
                            stock.get('001') if stock.get('001') > 0 else 0,
                            BOLT_PROVIDER_ID.get('001')['address']
                        ]
                        ni_stk = [
                            item.product.barcode,
                            stock.get('202')  if stock.get('202') > 0 else 0,
                            BOLT_PROVIDER_ID.get('202')['address']
                        ]
                        os_stk = [
                            item.product.barcode,
                            stock.get('205')  if stock.get('205') > 0 else 0,
                            BOLT_PROVIDER_ID.get('205')['address']
                        ]

                        margin = BOLT_MARGIN / 100
                        cap = Decimal(margin) * Decimal(item.product.price)
                        price_with_margin = item.product.price + cap
                        pw.writerow([str(item.product.barcode),price_with_margin,providers_id])
                        nia.writerow(ni_stk)
                        spintex.writerow(sp_stk)
                        osu.writerow(os_stk)


                        original_file =  item.image.path
                        file_name = os.path.basename(original_file)

                        name, ext = os.path.splitext(file_name)
                        new_file_name = f"{name}_image{ext}"

                        destination_file = os.path.join(folder, f"{new_file_name}")
                        shutil.copy(original_file, destination_file)

                        # copy file


                        print(li)
                        out.append(li)
                        writer.writerow(li)  # Writes each row once





                original_file = csv_file
                file_name = os.path.basename(original_file)
                destination_file = os.path.join(folder, f"{file_name}")
                shutil.copy(original_file, destination_file)

                zip_filename = folder + ".zip"
                # Compress the folder into a zip file
                with zipfile.ZipFile(zip_filename, 'w', zipfile.ZIP_DEFLATED) as zipf:
                    # Walk the directory tree and add all files to the zip
                    for root, dirs, files in os.walk(folder):
                        for file in files:
                            zipf.write(os.path.join(root, file), os.path.relpath(os.path.join(root, file), folder))

                success_response['message'] = {
                    "images":zip_filename,"csv":csv_file,'prices':prices_file,'stock':{
                        'nia':nia_stock_file,'osu':osu_stock_file,'spintex':spintex_stock_file
                    }
                }

            elif module == 'update_on_bolt':
                import csv
                time_ext = str(timezone.now()).replace(':', '_').replace(' ', '_')
                csv_file = f"static/general/tmp/bot_{time_ext}_file.csv"
                with open(csv_file, mode='w', newline='', encoding="utf-8") as px:
                    writer = csv.writer(px)
                    header = [
                        "SKU",
                        "Barcode",
                        "Name en-GH",
                        "Name en-US",
                        "Level 1 - Category",
                        "Level 2 - Subcategory",
                        "Provider IDs",
                        "Sub-category code",
                        "Description en-US",
                        "Description en-GH"
                    ]
                    writer.writerow(header)
                    items = BoltItems.objects.all().order_by('product__name')
                    providers_id = ""
                    for bolt_id in BOLT_PROVIDER_ID:
                        print(bolt_id)
                        row = BOLT_PROVIDER_ID[bolt_id]
                        providers_id += f"{row.get('id')},"

                    # remove last comma
                    providers_id = providers_id[:-1]
                    print(items.count())
                    for item in items:
                        barcodes = item.product.barcodes if item.product.barcodes else item.product.barcode
                        li = [
                            str(f"{item.product.barcode.split(',')[0]}"),
                            str(f"{barcodes}"),
                            item.product.name,
                            item.product.name,
                            item.group.name,
                            item.subgroup.name,
                            str(providers_id),
                            f"{item.group.name}/{item.subgroup.name}",
                            item.description,
                            item.description
                        ]
                        print(item.product.barcode)
                        writer.writerow(li)

                success_response['message'] = csv_file
                response = success_response
            elif module == 'bolt_stock_update':
                import os
                import zipfile
                import shutil
                import csv

                items = BoltItems.objects.all().order_by('product__name')
                time_ext = str(timezone.now()).replace(':', '_').replace(' ', '_')
                folder = f"static/general/tmp/{time_ext}"

                spintex_stock_file = f"static/general/tmp/bot_{time_ext}_spintex_stock.csv"
                nia_stock_file = f"static/general/tmp/bot_{time_ext}_nia_stock.csv"
                osu_stock_file = f"static/general/tmp/bot_{time_ext}_osu_stock.csv"
                providers_id = ""
                for bolt_id in BOLT_PROVIDER_ID:
                    print(bolt_id)
                    row = BOLT_PROVIDER_ID[bolt_id]
                    providers_id += f"{row.get('id')},"

                # remove last comma
                providers_id = providers_id[:-1]
                print(providers_id)




                # Create the folder if it doesn't exist
                if not os.path.exists(folder):
                    os.makedirs(folder)

                with open(spintex_stock_file, mode='w', newline='', encoding="utf-8") as spintex_stock,\
                        open(nia_stock_file, mode='w', newline='', encoding="utf-8") as nia_stock,\
                        open(osu_stock_file, mode='w', newline='', encoding="utf-8") as osu_stock:
                    stock_header = ['sku', 'quantity', 'address']

                    spintex = csv.writer(spintex_stock)
                    nia = csv.writer(nia_stock)
                    osu = csv.writer(osu_stock)
                    
                    nia.writerow(stock_header)
                    osu.writerow(stock_header)
                    spintex.writerow(stock_header)

                    for item in items:
                        print(item.product.barcode)
                        # stock = stock_by_prod(item.product.pk)
                        # print(stock)

                        # sp_stk = item.stock_spintex
                        # ni_stk = item.stock_nia
                        # os_stk = item.stock_osu

                        sp_stk = [
                            item.product.barcode,
                            item.stock_spintex,
                            BOLT_PROVIDER_ID.get('001')['address']
                        ]
                        ni_stk = [
                            item.product.barcode,
                            item.stock_nia,# if stock.get('nia') > 0 else 0,
                            BOLT_PROVIDER_ID.get('202')['address']
                        ]
                        os_stk = [
                            item.product.barcode,
                            item.stock_osu,# 5) if stock.get('osu') > 0 else 5,
                            BOLT_PROVIDER_ID.get('205')['address']
                        ]

                        nia.writerow(ni_stk)
                        spintex.writerow(sp_stk)
                        osu.writerow(os_stk)
                
                success_response['message'] = {
                    "spintex":spintex_stock_file,
                    "nia":nia_stock_file,
                    "osu":osu_stock_file
                }

            elif module == 'stock':
                item_code = data.get('item_code')
                success_response['message'] = get_stock(item_code)

            elif module == 'mark_send2bold':
                items = BoltItems.objects.filter(is_sync=False)
                import openpyxl
                wb = openpyxl.Workbook()
                sh = wb.active
                sh.title = "Sent To Bolt"
                header = ["BARCODE","NAME","PRICE"]
                sh.append(header)
                for item in items:
                    product = item.product
                    margin = BOLT_MARGIN / 100
                    cap = Decimal(margin) * Decimal(item.product.price)
                    price_with_margin = item.product.price + cap
                    li = [product.barcode,product.name,price_with_margin]
                    sh.append(li)

                time_ext = str(timezone.now()).replace(':', '_').replace(' ', '_')
                file_name = f"static/general/tmp/sent_to_bolt_as_of_{time_ext}.xlsx"
                wb.save(file_name)
                # Filter the BoltItems having is_sync=False
                BoltItems.objects.filter(is_sync=False).update(is_sync=True)

                success_response['message'] = file_name

            elif module == 'mark_stock_update':
                for product in Products.objects.all():
                    stock = stock_by_moved(product.pk,'*')
                    for key, value in stock.items():
                        loc = Locations.objects.get(code=key)
                        if MoveStock.objects.filter(product=product,location=loc).exists():
                            # update
                            MoveStock.objects.filter(product=product,location=loc).update(quantity=value)
                            print(f"{product.name} updated")
                        else:
                            MoveStock.objects.create(
                                product=product,
                                location=loc,
                                quantity=value
                            )
                            print(f"{product.name} created")



            elif module == 'check_bolt_expiry':
                ent_pk = data.get('entity')
                entity = BusinessEntityTypes.objects.get(pk=ent_pk)
                items = BoltItems.objects.filter(menu=entity,is_expired=True)
                import openpyxl
                book = openpyxl.Workbook()
                sheet = book.active
                hd = ["BARCODE","NAME","EXPIRY DATE"]
                sheet.append(hd)
                for item in items:
                    arr.append(item.obj())
                    li = [item.product.barcode,item.product.name,item.exp_date]
                    sheet.append(li)

                book.save('static/general/tmp/bolt_expity.xlsx')

                success_response['message'] = {
                    'excel':'static/general/tmp/bolt_expity.xlsx',
                    'json':arr
                }
                response = success_response
            else:
                raise Exception("No View Module")


        elif method == 'PATCH':
            if module == 'price_update':
                items = BoltItems.objects.all()

                for item in items:
                    item.price = item.product.price
                    item.save()

            elif module == 'hide_bolt':
                reason = data.get('hide_reason')
                pk = data.get('pk')

                item = BoltItems.objects.get(pk=pk)
                item.is_hidden = True
                item.hide_reason = reason
                item.save()

            elif module == 'check_bolt_expiry':
                ent_pk = data.get('entity',1)
                entity = BusinessEntityTypes.objects.get(pk=ent_pk)
                conn = ret_cursor()
                cursor = conn.cursor()
                for item in BoltItems.objects.filter(menu=entity):
                    product = item.product
                    barcode = product.barcode
                    itemcode = product.code

                    qu = f"select top(1) grn_date,hd.entry_no,tr.exp_date from grn_tran tr join grn_hd hd on hd.entry_no = tr.entry_no where tr.item_ref = '{barcode}' order by hd.grn_date desc"
                    cursor.execute(qu)
                    row = cursor.fetchone()
                    exp_date = '2099-01-01'
                    grn_date = exp_date
                    entry_no = ''
                    is_expired = False
                    today = timezone.now().date()
                    from datetime import datetime

                    if row:
                        grn_date,entry_no,exp_date = row
                        # get difference between dates

                        if not exp_date or exp_date == 'NULL':
                            exp_date = datetime.strptime('2099-01-01', '%Y-%m-%d').date()
                        else:
                            #print(str(exp_date).split(' ')[0])
                            exp_date = datetime.strptime(str(exp_date).split(' ')[0], '%Y-%m-%d').date()
                        
                        if exp_date < today:
                            is_expired = True
                    li = [barcode,grn_date,entry_no,exp_date,is_expired]
                   
                    
                    if item.is_expired != is_expired:
                        item.chng = True

                    if item.hide_reason == "EXP":
                        item.exp_last_check = timezone.now()

                    #print(item.hide_reason)

                    # check stock
                    stock_moved = stock_by_moved(item.product.pk,'*')
                    nia_stock = stock_moved.get('202',0)
                    osu_stock = stock_moved.get('205',0)
                    spi_stock = stock_moved.get('001',0)
                    print(stock_moved)

                    item.is_expired = is_expired
                    item.exp_date = exp_date
                    item.stock_nia = nia_stock
                    item.stock_spintex = spi_stock
                    item.stock_osu = osu_stock
                    item.save()


                    # li = [barcode,itemcode]

                conn = ret_cursor()


            elif module == 'menu_transfer':
                to_pk = data.get('to')
                items = data.get('items')
                entity = BusinessEntityTypes.objects.get(pk=to_pk)
                for item in items:
                    BoltItems.objects.filter(product__barcode=item).update(menu=entity)
                    
                success_response['message'] = f"Entity updated for {len(items)} products"
                response = success_response

            elif module == 'shelf':
                barcode = data.get('barcode')
                product = Products.objects.get(barcode=barcode)
                item_code = product.code
                product.shelf = data.get('shelf')
                product.save()

                # update shelf in product master
                # conn = ret_cursor()
                # cursor = conn.cursor()
                # cursor.execute(f"UPDATE prod_mast set shelf = '{shelf}' where item_code = '{item_code}'")
                # conn.commit()
                # conn.close()

                success_response['message'] = "Shelf Changed"


            elif module == 'mark2bolt':
                category = data.get('category')
                sub_category = data.get('sub_category')
                barcode = data.get('barcode')
                description = data.get('description')
                menu_pk = data.get('menu')
                menu = BusinessEntityTypes.objects.get(pk=menu_pk)

                group = BoltGroups.objects.get(pk=category)
                sub = BoltSubGroups.objects.get(pk=sub_category)
                product = Products.objects.get(barcode=barcode)

                stock = get_stock(barcode)
                stock_nia = 0
                stock_spintex = 0
                stock_osu = 0


                BoltItems.objects.create(
                    product = product,
                    group = group,
                    subgroup = sub,stock_nia = stock_nia,stock_spintex = stock_spintex,stock_osu = stock_osu,
                    description = description,menu=menu
                )

                success_response['message'] = "Product Marked"

            elif module == 'sample_adjust':
                sam_pk = data.get('sam_pk')
                adjustment_entry = data.get('adjustment_entry')

                sam = SampleHd.objects.get(pk=sam_pk)
                sam.ad = adjustment_entry
                sam.save()

            elif module == 'bolt_group':
                pk = data.get('pk')
                category = data.get('category')
                sub_category = data.get('sub_category')

                item = BoltItems.objects.get(pk=pk)
                group = BoltGroups.objects.get(pk=category)
                subgroup = BoltSubGroups.objects.get(pk=sub_category)

                item.group = group
                item.subgroup = subgroup
                item.group_changed = True

                item.save()
                success_response['message'] = "Category Changed"


            elif module == 'sample_refund':
                sam_pk = data.get('sam_pk')
                refund_entry = data.get('refund_entry')

                sam = SampleHd.objects.get(pk=sam_pk)
                sam.bill_refund_ref = refund_entry
                sam.save()

            elif module == 'sample_delete':
                sam_pk = data.get('sam_pk')

                SampleHd.objects.filter(pk=sam_pk).delete()

            elif module == 'sample_ad_check':
                samps = SampleHd.objects.filter(ad_sync=False)
                for samp in samps:
                    ad = samp.ad

                    print(ad)




            elif module == 'change_group':
                item_code = data.get('item_code')
                barcode = data.get('barcode')
                name = data.get('name')
                new_group = data.get('new_group')
                new_sub_group = data.get('new_sub_group')
                new_sub_subgroup = data.get('new_sub_subgroup')
                by = data.get('mypk')
                print(data)

                update_query = f"update prod_mast set group_code = '{new_group}', sub_group = '{new_sub_group}', sub_subgroup = '{new_sub_subgroup}',is_gp_chg = 1 where item_code = '{item_code}' "

                print(update_query)

                conn = ret_cursor()
                cursor = conn.cursor()
                cursor.execute(update_query)
                cursor.commit()
                conn.close()


                success_response['message'] = "Group Changed Successfully"
                response = success_response

            elif module == 'close_recipe':
                prod_key = data.get('prod_key')
                product = RecipeProduct.objects.get(pk=prod_key)
                items = product.recipe()

                if items.count() > 0:
                    # make recipe card
                    item_arr = []
                    for item in items:
                        item_arr.append([item.name, item.si_unit, item.quantity])

                    create_recipe_card(product.name.replace('/', ' '), item_arr)

                    success_response['message'] = "Recipe Closed"
                    product.is_open = False
                    product.save()

                else:
                    success_response['status_code'] = 404
                    success_response['message'] = "No Recipe Items"

            # flag stock monitoring
            elif module == 'flag_stock_monitoring':
                prod_pk = data.get('prod_pk')
                flag = data.get('flag')

                product = Products.objects.get(pk=prod_pk)
                name = product.name
                current = product.stock_monitor
                product.stock_monitor = flag
                product.save()

                success_response['message'] = f"Minitory flag changed for product {name} from {current} to {flag} "

        elif method == 'DELETE':
            if module == 'bolt_group':
                pk = data.get('pk')
                BoltGroups.objects.filter(pk=pk).delete()
                success_response['message'] = "Group Deleted Successfully"
            elif module == 'bolt_item':
                barcode = data.get('barcode')
                product = Products.objects.get(barcode=barcode)
                BoltItems.objects.filter(product__barcode=barcode).delete()

                success_response['message'] = "Item Deleted Successfully"
                product.allowed_on_bolt = False

        response = success_response

    except Exception as e:
        error_type, error_instance, traceback = sys.exc_info()
        tb_path = traceback.tb_frame.f_code.co_filename
        line_number = traceback.tb_lineno
        response["status_code"] = 500
        response[
            "message"] = f"An error occurred on line {line_number} in file {tb_path}. Details: {e}"

    return JsonResponse(response, safe=False)
