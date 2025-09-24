import json
import sys
from decimal import Decimal
from datetime import datetime, date, timedelta, timezone

from openpyxl.workbook import Workbook

from admin_panel.anton import create_new_user, staffs, get_week_of_month, month_by_name, md5only
from admin_panel.models import Sms, SmsApi, UserAddOns
from ocean.settings import ATTENDANCE_URL
from .modric import token
import requests

from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Q
from django.contrib.auth.models import User

from employee.models import Attendance, Leave


@csrf_exempt
def interface(request):
    response = {
        'status_code': 0,
        'message': ""
    }

    success_response = {
        'status_code': 200,
        'message': "Procedure Completed Successfully"
    }

    # get method
    method = request.method

    try:

        body = json.loads(request.body)
        module = body.get('module')
        data = body.get('data')


        if method == 'PUT':
            if module == 'area':
                import requests

                url = f"{ATTENDANCE_URL}personnel/api/areas/"  # Replace with actual domain

                headers = {
                    "Content-Type": "application/json",
                    "Authorization": f"JWT {token('solomon', 'Szczesny@411')}"
                }

                data = {
                    "area_code": data.get('area_code'),
                    "area_name": data.get('area_name'),
                    "parent_area": None
                }

                response = requests.post(url, json=data, headers=headers)


                success_response['message'] = "Area Created Successfully"

            elif module == 'leave':
                emp_code = data.get('emp_code')
                leave_type = data.get('type_of_leave')
                start_date = data.get('start_date')
                end_date = data.get('end_date')
                reason = data.get('reason')
                reliever = data.get('reliever_name')
                date_of_request = data.get('date_of_request')

                # get user
                adon = UserAddOns.objects.get(bio_id=emp_code)
                user = adon.user

                # create leave
                Leave.objects.create(
                    employee=user,
                    leave_type=leave_type,
                    start_date=start_date,
                    end_date=end_date,
                    reason=reason,
                    reliever=reliever,
                    date=date_of_request
                )



                success_response['message'] = "Leave Requested Successfully"
                response = success_response








            elif module == 'department':
                import requests

                url = f"{ATTENDANCE_URL}personnel/api/departments/"  # Replace with actual domain

                headers = {
                    "Content-Type": "application/json",
                    "Authorization": f"JWT {token('solomon', 'Szczesny@411')}"
                }

                data = {
                    "dept_code": data.get('dept_code'),
                    "dept_name": data.get('dept_name'),
                    "parent_dept": None
                }

                response = requests.post(url, json=data, headers=headers)
                success_response['message'] = "Dept Created Successfully"
                response = success_response

            elif module == 'position':
                import requests


                url = f"{ATTENDANCE_URL}personnel/api/positions/"  # Replace with actual domain

                headers = {
                    "Content-Type": "application/json",
                    "Authorization": f"JWT {token('solomon', 'Szczesny@411')}"
                }

                data = {
                    "position_code": data.get('position_code'),
                    "position_name": data.get('position_name'),
                    "parent_position": None
                }

                response = requests.post(url, json=data, headers=headers)
                success_response['message'] = "Position Created Successfully"
                response = success_response

            elif module == 'sync_attendance':
                # get workers
                tk = token('solomon', 'Szczesny@411')
                date_ini = data.get('date',datetime.now().strftime('%Y-%m-%d'))
                current_datetime = date_ini
                start = f"{str(current_datetime)} 00:00:00"
                end = f"{str(current_datetime)} 23:59:59"
                import requests
                url = f"{ATTENDANCE_URL}personnel/api/employees/?page_size=10000"

                payload = json.dumps({
                    "start_time": "2024-01-01",
                    "end_time": "2024-12-01",
                    "emp_code": 1,
                    "page_size": 500000,
                    "terminal_sn": ""
                })
                headers = {
                    'Authorization': f"JWT {tk}",
                    'Content-Type': 'application/json'
                }

                resp = requests.request("GET", url, headers=headers, data=payload)
                j_resp = json.loads(resp.text)

                gl = []
                count = j_resp['count']
                data = j_resp['data']
                js = []
                ps = 0
                abs = 0
                for i in data:
                    print(i)
                    enable_attendance = i['attemployee']['enable_attendance']
                    emp_code = i['emp_code']
                    if enable_attendance:

                        first_name = i['first_name']
                        last_name = i['last_name']
                        department = i['department']

                        dep_name = department.get('dept_name')
                        position = i['position']

                        position_name = position.get('position_name') if position else ''
                        contact = i['contact_tel']
                        mobile = i['mobile']
                        email = i['email']

                        li = [emp_code,f"{first_name} {last_name}",dep_name,position_name,contact,email]
                        # print(li)

                        # get attendance
                        url = f"{ATTENDANCE_URL}iclock/api/transactions/?start_time={start}&end_time={end}&page_size=10000&emp_code={emp_code}"

                        payload = json.dumps({
                            "start_time": "2024-01-01T00:00:00",
                            "end_time": "2024-01-30T00:00:00",
                            "emp_code": 1,
                            "page_size": 500000
                        })
                        headers = {
                            'Authorization': f"JWT {tk}",
                            'Content-Type': 'application/json'
                        }

                        att_response = requests.request("GET", url, headers=headers, data=payload)
                        attendance = json.loads(att_response.text)

                        # print(attendance)
                        att_count = attendance['count']
                        att_data = attendance['data']


                        check_in = ""
                        check_out = ""
                        minutes_diff = 0
                        present = False
                        if attendance['count'] == 0:
                            check_in = "00:00:00"
                            check_out = "00:00:00"
                            abs += 1
                        else:
                            first_record = att_data[0]
                            last_record = att_data[len(att_data) - 1]

                            first_date_time = datetime.strptime(str(first_record['punch_time']), '%Y-%m-%d %H:%M:%S')
                            first_time_only = first_date_time.time()

                            last_date_time = datetime.strptime(str(last_record['punch_time']), '%Y-%m-%d %H:%M:%S')
                            last_time_only = last_date_time.time()

                            total_time = last_date_time - first_date_time
                            minutes_diff = "{:.2f}".format(total_time.total_seconds() / 60)

                            check_in = first_time_only
                            check_out = last_time_only
                            present = True
                            ps += 1

                        st = 'absent'
                        is_late = False
                        late_duration = {
                            "SHOP": datetime.strptime('10:06', '%H:%M').time(),
                            "MOTORS": datetime.strptime('08:06', '%H:%M').time(),
                            "METERS": datetime.strptime('08:06', '%H:%M').time()
                        }

                        import calendar
                        dept_str = late_duration.get(dep_name, '00:00')
                        dept = datetime.strptime(dept_str, '%H:%M').time() if isinstance(dept_str,
                                                                                         str) else dept_str
                        day = calendar.day_name[datetime.strptime(start.split(' ')[0], '%Y-%m-%d').weekday()]


                        
                        

                        if present:
                            st = 'present'
                            print(department)
                            # dept = late_duration.get(dep_name,'00:00')
                            import calendar
                            dept_str = late_duration.get(dep_name, '00:00')
                            dept = datetime.strptime(dept_str, '%H:%M').time() if isinstance(dept_str,
                                                                                             str) else dept_str
                            if calendar.day_name[
                                datetime.strptime(start.split(' ')[0], '%Y-%m-%d').weekday()] == 'Sunday':
                                dept = (datetime.combine(datetime.today(), dept) + timedelta(hours=1)).time()

                            if check_in > dept:
                                is_late = True
                            else:
                                if(check_in == '00:00:00'):
                                    is_late = True
                                else:
                                    is_late = False



                        ali = [f"{first_name} {last_name}", dep_name, position_name, str(check_in), str(check_out), st,
                              minutes_diff,is_late]
                        js.append(ali)
                        dt = start.split(' ')[0]
                        if Attendance.objects.filter(date=dt,emp_code=emp_code).exists():
                            # update
                            Attendance.objects.filter(date=dt,emp_code=emp_code).update(
                                emp_code=emp_code,
                                date=dt,
                                time_in=check_in,
                                time_out=check_out,
                                time_diff=minutes_diff,
                                name=f"{first_name} {last_name}",
                                status=st,
                                is_late=is_late,
                                dep_text = dep_name
                            )
                        else:
                            Attendance(
                                emp_code=emp_code,
                                date=dt,
                                time_in=check_in,
                                time_out=check_out,
                                time_diff=minutes_diff,
                                name=f"{first_name} {last_name}",
                                status=st,
                                is_late=is_late,
                                dep_text=dep_name
                            ).save()

                    else:
                        Attendance.objects.filter(date=dt, emp_code=emp_code).delete()

                dat = {
                    "present":ps,
                    'absent':abs,
                    'array':js
                }
                success_response['message'] = dat
                response = success_response


            elif module == 'staff':
                import requests

                url = f"{ATTENDANCE_URL}personnel/api/employees/"  # Replace with actual domain

                headers = {
                    "Content-Type": "application/json",
                    "Authorization": f"JWT {token('solomon', 'Szczesny@411')}"
                }

                data = {
                    "emp_code": data.get('emp_code'),
                    "department": data.get('department'),
                    "area": data.get('area'),
                    "first_name": data.get('first_name'),
                    "last_name": data.get('last_name'),
                    "mobile": data.get('mobile'),
                    "email": data.get('email'),
                    "contact_tel":data.get('mobile'),
                    'position':data.get('position')

                }

                response = requests.post(url, json=data, headers=headers)
                print("RESPONSE")
                print(response.json())
                print("RESPONSE")
                success_response['message'] = "Staff Created Successfully"
                # make sms
                Sms.objects.create(
                    api=SmsApi.objects.get(is_default=True),
                    to=data.get('mobile'),
                    message=f"Dear {data.get('first_name')} {data.get('last_name')} your records have been entered into attendace / company records. Follow up to record your bio-data"
                )

                print("CREATED")
                response = success_response




        elif method == 'VIEW':
            if module == 'area':
                url = f'{ATTENDANCE_URL}personnel/api/areas/'

                tk = token('solomon','Szczesny@411')
                headers = {
                    'Authorization': f'JWT {tk}',
                    'Content-Type': 'application/json',
                    'ordering': 'dept_id'
                }
                import requests
                response = requests.get(url, headers=headers,json={})
                response = response.json()
                data = response.get('data')

                success_response['message'] = data
                response = success_response

            elif module == 'attendance':
                rg = data.get('range','week')
                mypk = data.get('mypk')
                add_on = UserAddOns.objects.get(user_id=mypk)
                emp_code = add_on.bio_id
                print(emp_code)
                res = []
                if rg == 'week':
                    today = date.today()
                    start_date = today - timedelta(days=today.weekday())
                    end_date = start_date + timedelta(days=6)

                    res = [att.obj() for att in Attendance.objects.filter(
                        emp_code=emp_code,
                        date__range=[start_date, end_date]
                    ).order_by('date')]
                    
                
                if rg == 'month':
                    today = date.today()
                    start_date = date(today.year, today.month, 1)
                    if today.month == 12:
                        end_date = date(today.year + 1, 1, 1) - timedelta(days=1)
                    else:
                        end_date = date(today.year, today.month + 1, 1) - timedelta(days=1)

                    res = [att.obj() for att in Attendance.objects.filter(
                        emp_code=emp_code,
                        date__range=[start_date, end_date]
                    ).order_by('date')]

                success_response['message'] = res


                # success_response['message'] = emp_code
                # get transactions


                response = success_response
            
            elif module == 'month_attendance':

                distinct_names = Attendance.objects.values_list('name', flat=True).distinct().order_by('name')
                mt = data.get('month', datetime.now().strftime('%Y-%m'))
                try:
                    month_year = datetime.strptime(mt, '%Y-%m')
                    start_date = month_year.replace(day=1, tzinfo=timezone.utc)
                    if month_year.month == 12:
                        end_date = month_year.replace(year=month_year.year + 1, month=1, day=1,
                                                      tzinfo=timezone.utc) - timedelta(days=1)
                    else:
                        end_date = month_year.replace(month=month_year.month + 1, day=1,
                                                      tzinfo=timezone.utc) - timedelta(days=1)
                except ValueError:
                    success_response['message'] = "Invalid month format. Use YYYY-MM"
                    success_response['status_code'] = 400
                    return JsonResponse(success_response, safe=False)

                results = []
                import openpyxl
                book = openpyxl.Workbook()
                sheet = book.active
                hd = ["NAME","DEPARTMENT","TOTAL LATE","LATE ABSENT","ABSENT","TOTAL ABSENT"]
                sheet.append(hd)

                for name in distinct_names:
                    attendance = Attendance.objects.filter(
                        name=name,
                        date__range=[start_date, end_date]
                    ).order_by('date')

                    attendance_records = []
                    total_late = 0
                    absent = 0
                    new_sheet = book.create_sheet()
                    new_sheet.title = name
                    new_sheet.append(["DATE","CHECK IN","CHECK OUT","STATUS","IS LATE"])
                    for record in attendance:
                        attendance_records.append({
                            'date': record.date,
                            'time_in': record.time_in,
                            'time_out': record.time_out,
                            'status': record.status,
                            'is_late': record.is_late
                        })
                        total_late += 1 if record.is_late else 0

                        if record.is_off() is False and record.status != 'present':
                            absent += 1

                        new_li = [record.date,record.time_in,record.time_out,record.status,record.is_late]
                        new_sheet.append(new_li)

                    # calculate total late
                    late_abcent = round(total_late / 3 if total_late > 0 else 0)
                    total_absent = absent + late_abcent
                    results.append({
                        'name': name,
                        'department': attendance.first().dep_text if attendance.exists() else None,
                        'total_late': total_late,
                        'late_absent': late_abcent,
                        'absent':absent,
                        'total_absent': total_absent,
                        'attendance': attendance_records
                    })
                    li = [
                        name,
                        attendance.first().dep_text if attendance.exists() else None,
                        total_late,
                        late_abcent,
                        absent,
                        total_absent
                    ]
                    sheet.append(li)
                    print(f"\nEmployee: {name}")
                    for record in attendance_records:
                        print(
                            f"Date: {record['date']}, Time In: {record['time_in']}, Time Out: {record['time_out']}, Status: {record['status']}")

                file = f'static/general/tmp/month_as_of_{str(datetime.now().strftime("%Y-%m-%d"))}.xlsx'

                book.save(file)
                success_response['message'] = {
                    'data':results,
                    'file':file
                }
                response = success_response

            elif module == 'department':
                import requests
                url = f"{ATTENDANCE_URL}personnel/api/departments/"
                tk   = token('solomon','Szczesny@411')
                headers = {
                    "Authorization": f"JWT {tk}",
                    "Content-Type": "application/json",
                }
                response = requests.get(url, headers=headers,params={
                    "page_size": 10000, "ordering": "dept_name"
                })
                response = response.json()
                success_response['message'] = response.get('data')
                response = success_response

            elif module == 'position':
                import requests
                url = f"{ATTENDANCE_URL}personnel/api/positions/"
                tk   = token('solomon','Szczesny@411')
                headers = {
                    "Authorization": f"JWT {tk}",
                    "Content-Type": "application/json",
                }
                response = requests.get(url, headers=headers,params={
                    "page_size": 10000, "ordering": "dept_name"
                })
                response = response.json()
                success_response['message'] = response.get('data')
                response = success_response

            elif module == 'staff':
                import requests
                url = f"{ATTENDANCE_URL}personnel/api/employees/"
                tk   = token('solomon','Szczesny@411')
                headers = {
                    "Authorization": f"JWT {tk}",
                    "Content-Type": "application/json",
                }
                response = requests.get(url, headers=headers,params={
                    "page_size": 10000, "ordering": "id"
                })
                response = response.json()
                success_response['message'] = response.get('data')
                response = success_response

            elif module == 'attendance_card':
                # loop through ann employee
                workers = staffs()
                late_duration = {
                    "SHOP": {
                        'late_by': datetime.strptime('10:06', '%H:%M').time(),
                        'off_days': 1
                    },
                    "MOTORS": {
                        'late_by': datetime.strptime('08:06', '%H:%M').time(),
                        'off_days': 1
                    },
                    "METERS": {
                        'late_by': datetime.strptime('08:06', '%H:%M').time(),
                        'off_days': 2
                    }
                }


                week_of_month = get_week_of_month()
                print(f"Current week of month: {week_of_month}")
                import openpyxl

                for worker in workers:
                    # print()
                    # print(worker)

                    enable_attendance = worker['attemployee']['enable_attendance']

                    if enable_attendance:
                        emp_id = worker.get('emp_code')
                        print(emp_id)
                        # print(emp_id)
                        mobile = worker.get('mobile').replace(' ','').replace("'",'')
                        this_week_started = datetime.now().date() - timedelta(days=(datetime.now().weekday() + 1) % 7)
                        # print(this_week_started)
                        department = worker.get('department').get('dept_name')
                        # print(department)
                        # get this week attendance

                        total_late = 0
                        total_absent = 0
                        week_attendance = Attendance.objects.filter(emp_code=emp_id,date__gte=this_week_started)
                        book = Workbook()
                        sheet = book.active
                        sheet.title = "Attendance"
                        sheet.append(["NAME","DEPARTMENT","DATE","TIME IN","TIME OUT","STATUS","IS LATE"])

                        msg_tran = ""
                        for att in week_attendance:
                            obj = att.obj()

                            li = [obj.get('name'),obj.get('dep_text'),obj.get('date'),obj.get('time_in'),obj.get('time_out'),obj.get('status'),obj.get('is_late')]
                            sheet.append(li)
                            total_late += 1 if obj.get('is_late') else 0
                            total_absent += 0 if obj.get('status') == 'present' else 1

                            msg_tran += f"""{obj.get('date')} : {obj.get('time_in')}\n"""


                        my_meta = late_duration.get(department)
                        my_off_days = my_meta.get('off_days')
                        total_absent -=my_off_days
                        file_name = f"static/general/tmp/{md5only(str(emp_id) + str(month_by_name()))}.xlsx"
                        book.save(file_name)
                        message = (f"{month_by_name()},"
                                   f"{datetime.now().year}\n\n"
                                   f"Week {week_of_month} Attendance\n"
                                   f"Total Days: {week_attendance.count()}\n"
                                   f"Total Late: {total_late}\n"
                                   f"Total Absent: {total_absent}\n\n"
                                   f"TRANSACTIONS:\n{msg_tran}\n\n")


                        # send SMS

                        if department != 'METERS':
                            print("SMS SENT TO ",emp_id,"ON",mobile)
                            Sms(
                                api=SmsApi.objects.get(is_default=True),
                                to=mobile,
                                message=message
                            ).save()

                        else:
                            print("SMS NOT SENT TO ",emp_id,"ON",mobile)

                        # print()
                        # exit()

            elif module == 'otp_auth':
                bio_id = data.get('bio_id')
                # bio_id = 2000
                import requests
                url = f"{ATTENDANCE_URL}personnel/api/employees/{bio_id}/"
                tk   = token('solomon','Szczesny@411')
                headers = {
                    "Authorization": f"JWT {tk}",
                    "Content-Type": "application/json",
                }
                rsp1 = requests.get(url, headers=headers,params={
                    "page_size": 10000, "ordering": "id"
                })
                rsp = rsp1.json()
                if rsp.get('detail'):
                    response['status_code'] = 404
                    response['message'] = rsp.get('detail')
                else:

                    # make otp
                    import random
                    otp = random.randint(100000, 999999)
                    message = f"Your verification code is {otp}. Do not share this code with anyone."
                    mobile = rsp.get('mobile')
                    if len(mobile) == 10:
                        success_response['message'] = {
                            'otp':otp,
                            'msg':f"OTP sent to {mobile}, please confirm it is you"
                        }
                        Sms.objects.create(
                            api=SmsApi.objects.get(is_default=True),
                            to=mobile,
                            message=message
                        )
                        response = success_response
                    else:
                        success_response['status_code'] = 505
                        success_response['message'] = f"{mobile} is an invalid phone number"

                    response = success_response

            elif module == 'leave':
                mypk = data.get('mypk','*')
                if mypk != '*':
                    user = User.objects.get(pk=mypk)
                    success_response['message'] = [lv.obj() for lv in Leave.objects.filter(employee=user)]
                else:
                    leaves = Leave.objects.all()
                    status = data.get('status','*')
                    if status != '*':
                        leaves = leaves.filter(status=status)
                    success_response['message'] = [lv.obj() for lv in leaves]

                response = success_response

            elif module == 'export_attendace':
                start_date = data.get('start_date')
                end_date = data.get('end_date')

            else:
                response['message'] = "Invalid Module"
                response['status_code'] = 505
    
        elif method == 'PATCH':
            if module == 'update_auth':
                bio_id = data.get('bio_id')
                bio_password = data.get('bio_password')
                mypk = data.get('mypk')

                user = User.objects.get(pk=mypk)
                ad_on = UserAddOns.objects.get(user=user)
                ad_on.bio_id = bio_id
                ad_on.bio_password = bio_password
                ad_on.save()

                response = success_response

            elif module == 'make_staff_for_ocean':
                # get all staffs
                import requests
                url = f"{ATTENDANCE_URL}personnel/api/employees/"
                tk = token('solomon', 'Szczesny@411')
                headers = {
                    "Authorization": f"JWT {tk}",
                    "Content-Type": "application/json",
                }
                dt = requests.get(url, headers=headers, params={
                    "page_size": 10000, "ordering": "id"
                })
                resp = dt.json()
                data = resp['data']
                for d in data:
                    # print(d)
                    first_name = d.get('first_name')
                    last_name = d.get('last_name')
                    emp_code = d.get('emp_code')
                    phone = d.get('mobile')
                    email = d.get('email','none@domain.com')
                    position = d['position'].get('position_name')
                    department = d['department'].get('dept_name')


                    if not email:
                        email = 'none@domain.com'

                    if phone is not None:
                        #print(first_name, last_name, emp_code, phone,email,position,department)
                        crt = create_new_user(first_name,last_name,email,phone,department,position)
                        if crt:
                            print(f"{first_name} {last_name} user CREATED")
                        else:
                            print(f"{first_name} {last_name} NOT CREATED")
                        print()


            elif module == 'leave_approval_request':
                pk = data.get('leave_id')

                leave = Leave.objects.get(pk=pk)
                user = leave.employee
                adon = UserAddOns.objects.get(user=user)
                # send sms
                sms_api = SmsApi.objects.get(is_default=True)
                # Message to employee
                employee_message = f"Dear {user.first_name}, your leave request from {leave.start_date} to {leave.end_date} has been submitted and is pending approval."
                Sms.objects.create(
                    api=sms_api,
                    to=adon.phone,
                    message=employee_message
                )

                # Notify provided phone number
                Sms.objects.create(
                    api=sms_api,
                    to="0546310011",
                    message=f"A {leave.leave_type} leave request by {adon.user.username} from {leave.start_date} to {leave.end_date} is pending for approval."
                )
                leave.status = 'requested'
                leave.save()

                success_response['message'] = "Request for Leave Approval Sent"
                response = success_response

        elif method == 'DELETE':
            if module == 'leave':
                Leave.objects.filter(pk=data.get('pk')).delete()
                success_response['message'] = "Leave Deleted Successfully"
                response = success_response
            else:
                raise Exception("Invalid Module")

    except Exception as e:
        error_type, error_instance, traceback = sys.exc_info()
        tb_path = traceback.tb_frame.f_code.co_filename
        line_number = traceback.tb_lineno
        response["status_code"] = 500
        response[
            "message"] = f"An error occurred on line {line_number} in file {tb_path}. Details: {e}"

    return JsonResponse(response, safe=False)